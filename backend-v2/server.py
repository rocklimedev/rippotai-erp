from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, Query, Body
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import io
import copy
import base64
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Any, Dict
import uuid
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
from jose import jwt, JWTError
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB
mongo_url = os.environ['MONGO_URL']
mongo_client = AsyncIOMotorClient(mongo_url)
db = mongo_client[os.environ['DB_NAME']]

# Auth setup
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
JWT_SECRET = os.environ.get('JWT_SECRET', 'buildcon-erp-secret-key-change-in-prod')
JWT_ALGO = "HS256"
JWT_EXPIRE_HOURS = 24 * 7
bearer_scheme = HTTPBearer(auto_error=False)

app = FastAPI(
    title="INOS ERP API",
    openapi_url="/api/openapi.json",
    docs_url="/api/docs",
    redoc_url="/api/redoc",
)
api = APIRouter(prefix="/api")


# ---------- Helpers ----------
def now_iso():
    return datetime.now(timezone.utc).isoformat()


def gen_id():
    return str(uuid.uuid4())


def hash_password(pw: str) -> str:
    return pwd_context.hash(pw)


def verify_password(pw: str, hashed: str) -> bool:
    try:
        return pwd_context.verify(pw, hashed)
    except Exception:
        return False


def create_token(user_id: str) -> str:
    payload = {
        "sub": user_id,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRE_HOURS),
        "iat": datetime.now(timezone.utc),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGO)


async def get_current_user(creds: Optional[HTTPAuthorizationCredentials] = Depends(bearer_scheme)):
    if not creds:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(creds.credentials, JWT_SECRET, algorithms=[JWT_ALGO])
        uid = payload.get("sub")
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")
    user = await db.users.find_one({"id": uid}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user


async def require_internal_user(current=Depends(get_current_user)):
    """Deny access to clients — quotations & vendors are internal to the firm."""
    if current.get("role") == "client":
        raise HTTPException(status_code=403, detail="Quotations are internal to the firm.")
    return current


async def require_admin(current=Depends(get_current_user)):
    """Admin-only routes."""
    if current.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return current


# ===== Phase K — plan / super-admin gating =====

PAID_PLANS = {"studio", "firm", "enterprise", "super_admin"}
FREE_TRIAL_CAPS = {"boqs": 3, "projects": 3, "quotations": 5, "documents": 20}
UPGRADE_MSG = "Upgrade to create more. This is included in Studio, Firm and Enterprise plans."


def _is_paid(user) -> bool:
    return bool(user.get("is_super_admin")) or user.get("plan") in PAID_PLANS


async def require_super_admin(current=Depends(get_current_user)):
    if not current.get("is_super_admin"):
        raise HTTPException(status_code=403, detail="Super admin access required")
    return current


async def require_paid_plan(current=Depends(get_current_user)):
    if not _is_paid(current):
        raise HTTPException(status_code=402, detail=UPGRADE_MSG)
    return current


async def enforce_free_trial_cap(user, collection_name: str):
    """If user is on free trial, enforce cap for records they've personally created."""
    if _is_paid(user):
        return
    cap = FREE_TRIAL_CAPS.get(collection_name)
    if cap is None:
        return
    count = await db[collection_name].count_documents({"created_by_id": user["id"]})
    if count >= cap:
        raise HTTPException(status_code=402, detail=UPGRADE_MSG)


ALLOWED_ROLES = {"admin", "project_manager", "architect", "estimator", "site_supervisor", "client"}


def _gen_temp_password(n: int = 10) -> str:
    import secrets, string
    alphabet = string.ascii_letters + string.digits
    return "".join(secrets.choice(alphabet) for _ in range(n))


def initials_from_name(name: str) -> str:
    parts = [p for p in name.split() if p]
    if not parts:
        return "?"
    if len(parts) == 1:
        return parts[0][:2].upper()
    return (parts[0][0] + parts[-1][0]).upper()


# ---------- Models ----------
class RegisterIn(BaseModel):
    name: str
    email: EmailStr
    password: str
    role: str = "project_manager"


class LoginIn(BaseModel):
    email: EmailStr
    password: str


class TokenOut(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: dict


# ---------- Auth Routes ----------
@api.post("/auth/register", response_model=TokenOut)
async def register(payload: RegisterIn):
    existing = await db.users.find_one({"email": payload.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    user = {
        "id": gen_id(),
        "name": payload.name,
        "email": payload.email.lower(),
        "password_hash": hash_password(payload.password),
        "role": payload.role,
        "avatar_initials": initials_from_name(payload.name),
        "created_at": now_iso(),
        "plan": "free_trial",
        "is_super_admin": False,
    }
    await db.users.insert_one(user)
    token = create_token(user["id"])
    user.pop("password_hash", None)
    user.pop("_id", None)
    return {"access_token": token, "token_type": "bearer", "user": user}


class SignupIn(BaseModel):
    name: str
    email: EmailStr
    password: str


@api.post("/auth/signup", response_model=TokenOut)
async def signup(payload: SignupIn):
    existing = await db.users.find_one({"email": payload.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    user = {
        "id": gen_id(),
        "name": payload.name,
        "email": payload.email.lower(),
        "password_hash": hash_password(payload.password),
        "role": "member",
        "avatar_initials": initials_from_name(payload.name),
        "created_at": now_iso(),
        "plan": "free_trial",
        "is_super_admin": False,
    }
    await db.users.insert_one(user)
    token = create_token(user["id"])
    user.pop("password_hash", None)
    user.pop("_id", None)
    return {"access_token": token, "token_type": "bearer", "user": user}


class DemoRequestIn(BaseModel):
    name: str
    email: EmailStr
    firm: str = ""
    phone: str = ""
    message: str


@api.post("/demo-requests")
async def create_demo_request(payload: DemoRequestIn):
    name = (payload.name or "").strip()
    message = (payload.message or "").strip()
    if not name or not message:
        raise HTTPException(status_code=422, detail="Name, work email and message are required.")
    doc = {
        "id": gen_id(),
        "name": name,
        "email": payload.email.lower(),
        "firm": (payload.firm or "").strip(),
        "phone": (payload.phone or "").strip(),
        "message": message,
        "created_at": now_iso(),
    }
    await db.demo_requests.insert_one(doc)
    doc.pop("_id", None)
    return {"ok": True, "id": doc["id"]}


# ===== Phase K — Super Admin endpoints =====

class SuperAdminUserPatch(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    plan: Optional[str] = None
    is_super_admin: Optional[bool] = None
    is_active: Optional[bool] = None


ALLOWED_PLANS = {"free_trial", "studio", "firm", "enterprise", "super_admin"}


@api.get("/super-admin/users")
async def sa_list_users(current=Depends(require_super_admin)):
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).sort("created_at", -1).to_list(1000)
    return users


@api.patch("/super-admin/users/{user_id}")
async def sa_patch_user(user_id: str, patch: SuperAdminUserPatch, current=Depends(require_super_admin)):
    updates = {}
    if patch.name is not None: updates["name"] = patch.name
    if patch.role is not None:
        if patch.role not in ALLOWED_ROLES and patch.role != "member":
            raise HTTPException(422, f"Invalid role. Allowed: {sorted(ALLOWED_ROLES | {'member'})}")
        updates["role"] = patch.role
    if patch.plan is not None:
        if patch.plan not in ALLOWED_PLANS:
            raise HTTPException(422, f"Invalid plan. Allowed: {sorted(ALLOWED_PLANS)}")
        updates["plan"] = patch.plan
    if patch.is_super_admin is not None: updates["is_super_admin"] = patch.is_super_admin
    if patch.is_active is not None: updates["is_active"] = patch.is_active
    if not updates:
        raise HTTPException(422, "No changes provided")
    res = await db.users.update_one({"id": user_id}, {"$set": updates})
    if res.matched_count == 0:
        raise HTTPException(404, "User not found")
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    return user


@api.delete("/super-admin/users/{user_id}")
async def sa_soft_delete_user(user_id: str, current=Depends(require_super_admin)):
    if user_id == current["id"]:
        raise HTTPException(400, "You cannot delete your own account.")
    res = await db.users.update_one({"id": user_id}, {"$set": {"is_active": False}})
    if res.matched_count == 0:
        raise HTTPException(404, "User not found")
    return {"ok": True}


@api.post("/super-admin/users/{user_id}/reset-password")
async def sa_reset_password(user_id: str, current=Depends(require_super_admin)):
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(404, "User not found")
    temp = _gen_temp_password(12)
    await db.users.update_one({"id": user_id}, {"$set": {"password_hash": hash_password(temp)}})
    # MOCKED email — return temp password so super admin can share it.
    return {"ok": True, "temp_password": temp, "email_sent": False, "note": "Email delivery is MOCKED; share this temp password with the user."}


# ==============================================



@api.post("/auth/login", response_model=TokenOut)
async def login(payload: LoginIn):
    user = await db.users.find_one({"email": payload.email.lower()})
    if not user or not verify_password(payload.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    token = create_token(user["id"])
    user.pop("password_hash", None)
    user.pop("_id", None)
    return {"access_token": token, "token_type": "bearer", "user": user}


@api.get("/auth/me")
async def me(current=Depends(get_current_user)):
    return current


# ---------- Dashboard ----------
@api.get("/dashboard/summary")
async def dashboard_summary(current=Depends(get_current_user)):
    active_projects = await db.projects.count_documents({"status": {"$ne": "completed"}})
    now = datetime.now(timezone.utc)
    week_end = now + timedelta(days=7)
    milestones = await db.milestones.find({"status": {"$ne": "done"}}, {"_id": 0}).to_list(500)
    due_this_week = 0
    for m in milestones:
        try:
            d = datetime.fromisoformat(m["due_at"])
            if now <= d <= week_end:
                due_this_week += 1
        except Exception:
            pass
    boqs_draft = await db.boqs.count_documents({"status": "draft"})
    docs_pending = await db.documents.count_documents({"status": "pending"})
    return {
        "active_projects": {"value": active_projects, "delta": 2, "delta_label": "vs last week"},
        "milestones_due_this_week": {"value": due_this_week, "delta": 1, "delta_label": "vs last week"},
        "boqs_in_draft": {"value": boqs_draft, "delta": -1, "delta_label": "vs last week"},
        "documents_pending": {"value": docs_pending, "delta": 3, "delta_label": "vs last week"},
    }


@api.get("/dashboard/continue-working")
async def continue_working(current=Depends(get_current_user)):
    items = await db.continue_working.find({}, {"_id": 0}).sort("updated_at", -1).to_list(10)
    return items[:3]


@api.get("/dashboard/app-badges")
async def app_badges(current=Depends(get_current_user)):
    boq_await = await db.boqs.count_documents({"status": "awaiting_review"})
    quotations_pending = await db.quotations.count_documents({"status": "pending"})
    calendar_upcoming = await db.calendar_events.count_documents({})
    chats_unread = 5
    return {
        "boq": boq_await,
        "quotations": quotations_pending,
        "calendar": calendar_upcoming,
        "chats": chats_unread,
    }


# ---------- Projects ----------
@api.get("/projects")
async def list_projects(limit: int = 6, sort: str = "priority", current=Depends(get_current_user)):
    projects = await db.projects.find({}, {"_id": 0}).to_list(200)
    # sort by priority: at_risk / delayed first, then by progress desc
    prio = {"delayed": 0, "at_risk": 1, "awaiting_input": 2, "on_track": 3, "completed": 4}
    projects.sort(key=lambda p: (prio.get(p.get("timeline_status", "on_track"), 5), -p.get("progress", 0)))
    return projects[:limit]


@api.get("/projects/{project_id}/handover-readiness")
async def handover_readiness(project_id: str, current=Depends(get_current_user)):
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    checklist = project.get("handover_checklist") or []
    if not checklist:
        checklist = [
            {"name": "Signed Agreement", "done": True},
            {"name": "Final BOQ Approved", "done": True},
            {"name": "Site Reki Report", "done": True},
            {"name": "Design Freeze Document", "done": True},
            {"name": "Vendor Quotations (All)", "done": True},
            {"name": "Purchase Orders", "done": True},
            {"name": "Execution Photos", "done": True},
            {"name": "Client Sign-off Sheet", "done": True},
            {"name": "Material Warranty Docs", "done": True},
            {"name": "Snag List Resolved", "done": True},
            {"name": "Post-Execution Reki", "done": True},
            {"name": "Final Invoice", "done": True},
            {"name": "Handover Note", "done": False},
            {"name": "Client Feedback Form", "done": False},
            {"name": "Maintenance Guide", "done": False},
        ]
    required = len(checklist)
    available = sum(1 for c in checklist if c["done"])
    return {
        "project_id": project_id,
        "project_name": project["name"],
        "required": required,
        "available": available,
        "pending": required - available,
        "percent": round(available * 100 / required) if required else 0,
        "checklist": checklist,
    }


# ---------- BOQ ----------
@api.get("/boq/productivity")
async def boq_productivity(current=Depends(get_current_user)):
    return {
        "boqs_this_month": 8,
        "avg_creation_time_minutes": 42,
        "hours_saved": 36,
        "awaiting_approval": 2,
        "top_category": "Furniture & Joinery",
        "monthly_series": [
            {"month": "Sep", "count": 4},
            {"month": "Oct", "count": 6},
            {"month": "Nov", "count": 5},
            {"month": "Dec", "count": 7},
            {"month": "Jan", "count": 9},
            {"month": "Feb", "count": 8},
        ],
    }


# ---------- Milestones ----------
@api.get("/milestones/upcoming")
async def milestones_upcoming(limit: int = 5, current=Depends(get_current_user)):
    items = await db.milestones.find({"status": {"$ne": "done"}}, {"_id": 0}).to_list(200)
    items.sort(key=lambda m: m.get("due_at", ""))
    return items[:limit]


# ---------- Documents ----------
@api.get("/documents/recent")
async def documents_recent(limit: int = 8, filter: str = "all", current=Depends(get_current_user)):
    q: dict = {}
    if filter == "generated":
        q["origin"] = "generated"
    elif filter == "uploaded":
        q["origin"] = "uploaded"
    items = await db.documents.find(q, {"_id": 0}).to_list(200)
    items.sort(key=lambda d: d.get("uploaded_at", ""), reverse=True)
    return items[:limit]


# ---------- Activity ----------
@api.get("/activity/recent")
async def activity_recent(limit: int = 10, current=Depends(get_current_user)):
    items = await db.activity.find({}, {"_id": 0}).to_list(200)
    items.sort(key=lambda a: a.get("at", ""), reverse=True)
    return items[:limit]


# ---------- Calendar ----------
@api.get("/calendar/upcoming")
async def calendar_upcoming(limit: int = 5, current=Depends(get_current_user)):
    items = await db.calendar_events.find({}, {"_id": 0}).to_list(200)
    items.sort(key=lambda e: e.get("at", ""))
    return items[:limit]


# ---------- Clients / Vendors / Quotations ----------
@api.get("/clients")
async def list_clients(current=Depends(get_current_user)):
    return await db.clients.find({}, {"_id": 0}).to_list(200)


# ============================================================
# ================ PHASE 3 — VENDORS MODULE ==================
# ============================================================
from fastapi import UploadFile, File, Form
from fastapi.responses import FileResponse, Response as FastResponse
import csv

UPLOAD_ROOT = Path(__file__).parent / "uploads"
UPLOAD_ROOT.mkdir(exist_ok=True)


class VendorIn(BaseModel):
    model_config = ConfigDict(extra="allow")


def _vendor_full(v: dict, docs=None, ratings=None) -> dict:
    if docs is not None: v["documents"] = docs
    if ratings is not None:
        v["ratings_summary"] = _rating_summary(ratings)
    return v


def _rating_summary(ratings: list) -> dict:
    if not ratings:
        return {"avg": 0, "count": 0, "breakdown": {}}
    keys = ["work_quality", "timeline", "communication", "pricing", "material_quality", "after_sales"]
    breakdown = {}
    for k in keys:
        vals = [r.get(k) for r in ratings if r.get(k) is not None]
        breakdown[k] = round(sum(vals)/len(vals), 2) if vals else 0
    avg = round(sum(breakdown.values())/max(1, len([v for v in breakdown.values() if v>0])), 2) if breakdown else 0
    return {"avg": avg, "count": len(ratings), "breakdown": breakdown}


@api.get("/vendors")
async def vendors_list(
    q: Optional[str] = None,
    vendor_type: Optional[str] = None,
    category: Optional[str] = None,
    city: Optional[str] = None,
    state: Optional[str] = None,
    price_range: Optional[str] = None,
    min_rating: Optional[float] = None,
    verified: Optional[bool] = None,
    available_now: Optional[bool] = None,
    preferred: Optional[bool] = None,
    blocked: Optional[bool] = None,
    inactive: Optional[bool] = None,
    project_type: Optional[str] = None,
    min_completed_projects: Optional[int] = None,
    documents_complete: Optional[bool] = None,
    brands: Optional[str] = None,
    materials: Optional[str] = None,
    limit: int = 200,
    current=Depends(require_internal_user),
):
    query: dict = {}
    if not blocked: query["status"] = {"$ne": "blocked"}
    if inactive is False: query["status"] = {"$nin": ["blocked", "inactive"]}
    if vendor_type: query["vendor_type"] = vendor_type
    if category: query["primary_category"] = category
    if city: query["city"] = {"$regex": f"^{city}", "$options": "i"}
    if state: query["state"] = state
    if price_range: query["price_range"] = price_range
    if verified is not None: query["verified"] = verified
    if preferred is not None: query["preferred"] = preferred
    if available_now is True: query["availability_status"] = "available"
    if min_rating is not None: query["rating"] = {"$gte": min_rating}
    if min_completed_projects is not None: query["completed_projects"] = {"$gte": min_completed_projects}
    if documents_complete is not None: query["documents_complete"] = documents_complete
    if project_type: query["project_types"] = project_type
    if brands: query["brands"] = {"$in": [b.strip() for b in brands.split(",") if b.strip()]}
    if materials: query["materials"] = {"$in": [m.strip() for m in materials.split(",") if m.strip()]}
    if q:
        ql = q.strip()
        query["$or"] = [
            {"name": {"$regex": ql, "$options": "i"}},
            {"company": {"$regex": ql, "$options": "i"}},
            {"vendor_type": {"$regex": ql, "$options": "i"}},
            {"primary_category": {"$regex": ql, "$options": "i"}},
            {"specializations": {"$regex": ql, "$options": "i"}},
            {"brands": {"$regex": ql, "$options": "i"}},
            {"materials": {"$regex": ql, "$options": "i"}},
            {"city": {"$regex": ql, "$options": "i"}},
        ]
    rows = await db.vendors.find(query, {"_id": 0}).limit(limit).to_list(limit)
    rows.sort(key=lambda r: (-1 if r.get("preferred") else 0, -(r.get("rating") or 0), r.get("name", "")))
    return rows


@api.get("/vendors/summary")
async def vendors_summary(current=Depends(require_internal_user)):
    # "Total Vendors" = all vendors regardless of status, matching the donut sum
    total = await db.vendors.count_documents({})
    verified = await db.vendors.count_documents({"verified": True, "status": {"$nin": ["blocked", "archived"]}})
    available = await db.vendors.count_documents({"availability_status": "available", "status": {"$nin": ["blocked", "archived"]}})
    active_project = await db.vendors.count_documents({"current_assignments": {"$gt": 0}})
    now = datetime.now(timezone.utc)
    cutoff = now - timedelta(days=30)
    recent = await db.vendors.count_documents({"created_at": {"$gte": cutoff.isoformat()}})
    attention = await db.vendors.count_documents({"$or": [{"documents_complete": False}, {"documents_expired": True}]})
    return {"total": total, "verified": verified, "available": available, "active_project": active_project, "recently_added": recent, "attention": attention}


@api.get("/vendors/duplicate-check")
async def vendor_duplicate_check(phone: Optional[str] = None, email: Optional[str] = None, gst: Optional[str] = None, exclude_id: Optional[str] = None, current=Depends(require_internal_user)):
    if not (phone or email or gst):
        return {"matches": []}
    q = {"$or": []}
    if phone: q["$or"].append({"phone": phone})
    if email: q["$or"].append({"email": email.lower()})
    if gst: q["$or"].append({"gst_number": gst})
    rows = await db.vendors.find(q, {"_id": 0, "id": 1, "name": 1, "company": 1, "phone": 1, "email": 1, "gst_number": 1}).to_list(20)
    if exclude_id: rows = [r for r in rows if r.get("id") != exclude_id]
    return {"matches": rows}


@api.get("/vendors/saved-searches")
async def saved_searches_list(current=Depends(require_internal_user)):
    rows = await db.vendor_saved_searches.find({"$or": [{"user_id": current["id"]}, {"scope": {"$in": ["org", "team"]}}]}, {"_id": 0}).to_list(200)
    return rows


class SavedSearchIn(BaseModel):
    name: str
    filters: dict
    scope: str = "personal"


@api.post("/vendors/saved-searches")
async def saved_search_create(payload: SavedSearchIn, current=Depends(require_internal_user)):
    doc = {"id": gen_id(), "name": payload.name, "filters": payload.filters, "scope": payload.scope, "user_id": current["id"], "created_at": now_iso()}
    await db.vendor_saved_searches.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.delete("/vendors/saved-searches/{sid}")
async def saved_search_delete(sid: str, current=Depends(require_internal_user)):
    await db.vendor_saved_searches.delete_one({"id": sid, "user_id": current["id"]})
    return {"ok": True}


@api.get("/vendors/export")
async def vendors_export(format: str = "csv", current=Depends(require_internal_user)):
    rows = await db.vendors.find({"status": {"$nin": ["blocked", "archived"]}}, {"_id": 0}).to_list(500)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Vendor Code", "Name", "Company", "Vendor Type", "Primary Category", "City", "State", "Phone", "Email", "Price Range", "Rating", "Completed Projects", "Verified", "Availability", "Preferred"])
    for r in rows:
        w.writerow([r.get("vendor_code", ""), r.get("name", ""), r.get("company", ""), r.get("vendor_type", ""), r.get("primary_category", ""), r.get("city", ""), r.get("state", ""), r.get("phone", ""), r.get("email", ""), r.get("price_range", ""), r.get("rating", ""), r.get("completed_projects", ""), r.get("verified", ""), r.get("availability_status", ""), r.get("preferred", "")])
    data = buf.getvalue().encode()
    return FastResponse(content=data, media_type="text/csv", headers={"Content-Disposition": 'attachment; filename="vendors.csv"'})


@api.get("/vendors/{vendor_id}")
async def vendor_get(vendor_id: str, current=Depends(require_internal_user)):
    v = await db.vendors.find_one({"id": vendor_id}, {"_id": 0})
    if not v: raise HTTPException(404, "Vendor not found")
    docs = await db.vendor_documents.find({"vendor_id": vendor_id}, {"_id": 0}).to_list(200)
    ratings = await db.vendor_ratings.find({"vendor_id": vendor_id}, {"_id": 0}).to_list(200)
    projects = [p async for p in db.projects.find({"id": {"$in": v.get("project_ids") or []}}, {"_id": 0, "id": 1, "name": 1, "type": 1, "current_phase": 1, "status": 1, "location": 1})]
    quotations = await db.quotations.find({"vendor_id": vendor_id}, {"_id": 0}).to_list(200)
    activity = await db.vendor_activity.find({"vendor_id": vendor_id}, {"_id": 0}).sort("at", -1).to_list(50)
    notes = await db.vendor_notes.find({"vendor_id": vendor_id}, {"_id": 0}).sort("at", -1).to_list(50)
    v["documents"] = docs
    v["ratings_summary"] = _rating_summary(ratings)
    v["ratings"] = ratings
    v["projects_worked"] = projects
    v["quotations"] = quotations
    v["activity"] = activity
    v["notes"] = notes
    # mask sensitive fields for non-admin
    if current.get("role") != "admin":
        if v.get("bank_details"): v["bank_details"] = {"_masked": True}
        if v.get("gst_number"): v["gst_number"] = "•••••" + v["gst_number"][-4:]
        if v.get("pan"): v["pan"] = "•••••" + v["pan"][-4:]
    return v


@api.post("/vendors")
async def vendor_create(payload: VendorIn, current=Depends(require_internal_user)):
    d = payload.model_dump(exclude_none=True)
    vid = gen_id()
    doc = {
        "id": vid,
        "vendor_code": d.get("vendor_code") or f"BC-V-{str(uuid.uuid4())[:6].upper()}",
        "status": d.get("status", "draft"),
        "verified": False, "preferred": False,
        "documents_complete": False, "documents_expired": False,
        "current_assignments": 0, "completed_projects": 0,
        "rating": 0.0, "availability_status": "available",
        "created_by": current.get("name"), "created_at": now_iso(), "updated_at": now_iso(),
        **d,
    }
    await db.vendors.insert_one(doc)
    await db.vendor_activity.insert_one({"id": gen_id(), "vendor_id": vid, "action": "created", "actor": current.get("name"), "at": now_iso()})
    doc.pop("_id", None)
    return doc


@api.patch("/vendors/{vendor_id}")
async def vendor_patch(vendor_id: str, payload: VendorIn, current=Depends(require_internal_user)):
    v = await db.vendors.find_one({"id": vendor_id}, {"_id": 0})
    if not v: raise HTTPException(404, "Vendor not found")
    d = payload.model_dump(exclude_unset=True)
    d["updated_at"] = now_iso()
    await db.vendors.update_one({"id": vendor_id}, {"$set": d})
    return await db.vendors.find_one({"id": vendor_id}, {"_id": 0})


@api.delete("/vendors/{vendor_id}")
async def vendor_delete(vendor_id: str, current=Depends(require_internal_user)):
    v = await db.vendors.find_one({"id": vendor_id}, {"_id": 0})
    if not v: raise HTTPException(404, "Vendor not found")
    if v.get("status") != "draft":
        raise HTTPException(400, "Only draft vendors can be deleted")
    await db.vendors.delete_one({"id": vendor_id})
    return {"ok": True}


@api.post("/vendors/{vendor_id}/set-preferred")
async def vendor_set_preferred(vendor_id: str, current=Depends(require_internal_user)):
    await db.vendors.update_one({"id": vendor_id}, {"$set": {"preferred": True, "updated_at": now_iso()}})
    return {"ok": True}


@api.post("/vendors/{vendor_id}/block")
async def vendor_block(vendor_id: str, current=Depends(require_internal_user)):
    await db.vendors.update_one({"id": vendor_id}, {"$set": {"status": "blocked", "updated_at": now_iso()}})
    return {"ok": True}


@api.post("/vendors/{vendor_id}/archive")
async def vendor_archive(vendor_id: str, current=Depends(require_internal_user)):
    await db.vendors.update_one({"id": vendor_id}, {"$set": {"status": "archived", "updated_at": now_iso()}})
    return {"ok": True}


# ---- Ratings ----
class RatingIn(BaseModel):
    project_id: Optional[str] = None
    work_quality: float
    timeline: float
    communication: float
    pricing: float
    material_quality: float
    after_sales: float
    comment: Optional[str] = ""


@api.post("/vendors/{vendor_id}/ratings")
async def vendor_rating_create(vendor_id: str, payload: RatingIn, current=Depends(require_internal_user)):
    d = payload.model_dump()
    keys = ["work_quality", "timeline", "communication", "pricing", "material_quality", "after_sales"]
    d["avg"] = round(sum(d[k] for k in keys) / len(keys), 2)
    d.update({"id": gen_id(), "vendor_id": vendor_id, "given_by": current.get("name"), "given_at": now_iso()})
    await db.vendor_ratings.insert_one(d)
    # update vendor avg rating
    ratings = await db.vendor_ratings.find({"vendor_id": vendor_id}).to_list(500)
    avg = round(sum(r["avg"] for r in ratings) / len(ratings), 2) if ratings else 0
    await db.vendors.update_one({"id": vendor_id}, {"$set": {"rating": avg}})
    d.pop("_id", None)
    return d


@api.get("/vendors/{vendor_id}/ratings")
async def vendor_ratings_list(vendor_id: str, current=Depends(require_internal_user)):
    return await db.vendor_ratings.find({"vendor_id": vendor_id}, {"_id": 0}).sort("given_at", -1).to_list(500)


# ---- Notes ----
class NoteIn(BaseModel):
    body: str


@api.post("/vendors/{vendor_id}/notes")
async def vendor_note_add(vendor_id: str, payload: NoteIn, current=Depends(require_internal_user)):
    doc = {"id": gen_id(), "vendor_id": vendor_id, "body": payload.body, "by": current.get("name"), "at": now_iso()}
    await db.vendor_notes.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.delete("/vendors/{vendor_id}/notes/{nid}")
async def vendor_note_delete(vendor_id: str, nid: str, current=Depends(require_internal_user)):
    await db.vendor_notes.delete_one({"id": nid, "vendor_id": vendor_id})
    return {"ok": True}


# ---- Attachments / File upload ----
@api.post("/vendors/{vendor_id}/attachments")
async def vendor_attachment_upload(vendor_id: str, kind: str = Form(...), file: UploadFile = File(...), expiry_date: Optional[str] = Form(None), current=Depends(require_internal_user)):
    v = await db.vendors.find_one({"id": vendor_id}, {"_id": 0})
    if not v: raise HTTPException(404, "Vendor not found")
    contents = await file.read()
    if len(contents) > 10 * 1024 * 1024:
        raise HTTPException(413, "File exceeds 10MB")
    vdir = UPLOAD_ROOT / "vendors" / vendor_id
    vdir.mkdir(parents=True, exist_ok=True)
    aid = gen_id()
    safe_name = "".join(c if c.isalnum() or c in "._-" else "_" for c in file.filename or "file")
    dest = vdir / f"{aid}-{safe_name}"
    dest.write_bytes(contents)
    rel = f"vendors/{vendor_id}/{aid}-{safe_name}"
    status_v = "valid"
    if expiry_date:
        try:
            e = datetime.fromisoformat(expiry_date).date()
            today = datetime.now(timezone.utc).date()
            if e < today: status_v = "expired"
            elif (e - today).days < 30: status_v = "expiring_soon"
        except Exception:
            pass
    doc = {
        "id": aid, "vendor_id": vendor_id, "kind": kind,
        "filename": file.filename, "path": rel, "size": len(contents),
        "mime": file.content_type or "application/octet-stream",
        "uploaded_by": current.get("name"), "uploaded_at": now_iso(),
        "expiry_date": expiry_date, "status": status_v,
    }
    await db.vendor_documents.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.delete("/vendors/{vendor_id}/attachments/{aid}")
async def vendor_attachment_delete(vendor_id: str, aid: str, current=Depends(require_internal_user)):
    d = await db.vendor_documents.find_one({"id": aid, "vendor_id": vendor_id}, {"_id": 0})
    if d:
        p = UPLOAD_ROOT / d.get("path", "")
        if p.exists(): p.unlink(missing_ok=True)
        await db.vendor_documents.delete_one({"id": aid})
    return {"ok": True}


@api.get("/files/{path:path}")
async def get_file(path: str, current=Depends(require_internal_user)):
    fp = UPLOAD_ROOT / path
    if not fp.exists() or not str(fp.resolve()).startswith(str(UPLOAD_ROOT.resolve())):
        raise HTTPException(404, "File not found")
    return FileResponse(str(fp))


# ---- Shortlists ----
class ShortlistIn(BaseModel):
    name: str
    project_id: Optional[str] = None
    work_package: Optional[str] = ""
    notes: Optional[str] = ""


@api.get("/vendor-shortlists")
async def shortlists_list(current=Depends(require_internal_user)):
    lists = await db.vendor_shortlists.find({}, {"_id": 0}).sort("updated_at", -1).to_list(200)
    for sl in lists:
        sl["vendor_count"] = await db.vendor_shortlist_items.count_documents({"shortlist_id": sl["id"]})
    return lists


@api.post("/vendor-shortlists")
async def shortlist_create(payload: ShortlistIn, current=Depends(require_internal_user)):
    proj = None
    if payload.project_id:
        proj = await db.projects.find_one({"id": payload.project_id}, {"_id": 0, "id": 1, "name": 1})
    doc = {"id": gen_id(), "name": payload.name, "project_id": payload.project_id,
           "project_name": proj["name"] if proj else None, "work_package": payload.work_package,
           "notes": payload.notes, "created_by": current.get("name"),
           "created_at": now_iso(), "updated_at": now_iso()}
    await db.vendor_shortlists.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.get("/vendor-shortlists/{sid}")
async def shortlist_get(sid: str, current=Depends(require_internal_user)):
    sl = await db.vendor_shortlists.find_one({"id": sid}, {"_id": 0})
    if not sl: raise HTTPException(404, "Shortlist not found")
    items = await db.vendor_shortlist_items.find({"shortlist_id": sid}, {"_id": 0}).sort("order", 1).to_list(200)
    vids = [i["vendor_id"] for i in items]
    vendors = {v["id"]: v async for v in db.vendors.find({"id": {"$in": vids}}, {"_id": 0})}
    for it in items:
        it["vendor"] = vendors.get(it["vendor_id"])
    sl["items"] = items
    return sl


@api.patch("/vendor-shortlists/{sid}")
async def shortlist_patch(sid: str, payload: ShortlistIn, current=Depends(require_internal_user)):
    await db.vendor_shortlists.update_one({"id": sid}, {"$set": {**payload.model_dump(exclude_none=True), "updated_at": now_iso()}})
    return await db.vendor_shortlists.find_one({"id": sid}, {"_id": 0})


@api.delete("/vendor-shortlists/{sid}")
async def shortlist_delete(sid: str, current=Depends(require_internal_user)):
    await db.vendor_shortlists.delete_one({"id": sid})
    await db.vendor_shortlist_items.delete_many({"shortlist_id": sid})
    return {"ok": True}


class ShortlistItemIn(BaseModel):
    vendor_id: str
    internal_remarks: Optional[str] = ""


@api.post("/vendor-shortlists/{sid}/items")
async def shortlist_add_item(sid: str, payload: ShortlistItemIn, current=Depends(require_internal_user)):
    existing = await db.vendor_shortlist_items.find_one({"shortlist_id": sid, "vendor_id": payload.vendor_id}, {"_id": 0})
    if existing:
        return existing
    order = await db.vendor_shortlist_items.count_documents({"shortlist_id": sid})
    doc = {"id": gen_id(), "shortlist_id": sid, "vendor_id": payload.vendor_id, "order": order, "internal_remarks": payload.internal_remarks, "added_by": current.get("name"), "added_at": now_iso()}
    await db.vendor_shortlist_items.insert_one(doc)
    await db.vendor_shortlists.update_one({"id": sid}, {"$set": {"updated_at": now_iso()}})
    doc.pop("_id", None)
    return doc


@api.patch("/vendor-shortlists/{sid}/items/{iid}")
async def shortlist_item_patch(sid: str, iid: str, payload: ShortlistItemIn, current=Depends(require_internal_user)):
    await db.vendor_shortlist_items.update_one({"id": iid, "shortlist_id": sid}, {"$set": {"internal_remarks": payload.internal_remarks}})
    return {"ok": True}


@api.delete("/vendor-shortlists/{sid}/items/{iid}")
async def shortlist_item_delete(sid: str, iid: str, current=Depends(require_internal_user)):
    await db.vendor_shortlist_items.delete_one({"id": iid, "shortlist_id": sid})
    return {"ok": True}


class VendorReorderIn(BaseModel):
    ordered_ids: List[str]


@api.post("/vendor-shortlists/{sid}/items/reorder")
async def shortlist_reorder(sid: str, payload: VendorReorderIn, current=Depends(require_internal_user)):
    for i, iid in enumerate(payload.ordered_ids):
        await db.vendor_shortlist_items.update_one({"id": iid, "shortlist_id": sid}, {"$set": {"order": i}})
    return {"ok": True}


# /api/quotations moved to Phase 4 module below.


# ---------- Search ----------
@api.get("/search")
async def search(q: str = Query(""), limit: int = 5, current=Depends(get_current_user)):
    if not q or len(q) < 2:
        return {"boqs": [], "projects": [], "vendors": [], "quotations": [], "documents": [], "tasks": [], "counts": {"boqs":0,"projects":0,"vendors":0,"quotations":0,"documents":0,"tasks":0}}
    ql = q.lower()

    def match_all(items, fields):
        out = []
        for it in items:
            for f in fields:
                v = str(it.get(f, "") or "").lower()
                if ql in v:
                    out.append(it); break
        return out

    projects_all = await db.projects.find({}, {"_id": 0}).to_list(500)
    boqs_all = await db.boqs.find({}, {"_id": 0}).to_list(500)
    vendors_all = await db.vendors.find({}, {"_id": 0}).to_list(500)
    quotations_all = await db.quotations.find({}, {"_id": 0}).to_list(500)
    documents_all = await db.documents.find({"is_archived": {"$ne": True}}, {"_id": 0}).to_list(500)
    tasks_all = await db.tasks.find({"is_archived": {"$ne": True}}, {"_id": 0}).to_list(500)

    projects_m = match_all(projects_all, ["name", "client_name", "location", "project_type"])
    boqs_m = match_all(boqs_all, ["boq_number", "version", "project_name", "status"])
    vendors_m = match_all(vendors_all, ["name", "company", "primary_category", "phone"])
    quotations_m = match_all(quotations_all, ["quotation_number", "title", "project_name", "vendor_name", "status"])
    documents_m = match_all(documents_all, ["title", "filename", "project_name", "category", "uploaded_by_name"])
    tasks_m = match_all(tasks_all, ["title", "description", "project_name", "assignee_name", "status"])

    return {
        "boqs": boqs_m[:limit],
        "projects": projects_m[:limit],
        "vendors": vendors_m[:limit],
        "quotations": quotations_m[:limit],
        "documents": documents_m[:limit],
        "tasks": tasks_m[:limit],
        "counts": {
            "boqs": len(boqs_m), "projects": len(projects_m), "vendors": len(vendors_m),
            "quotations": len(quotations_m), "documents": len(documents_m), "tasks": len(tasks_m),
        },
    }


# ---------- Notifications ----------
@api.get("/notifications")
async def notifications(current=Depends(get_current_user)):
    # Notifications include a link_url pointing to the related record.
    # Look up first BOQ / project / quotation for realistic deep links.
    first_boq = await db.boqs.find_one({"status": {"$in": ["approved", "final"]}}, {"_id": 0, "id": 1})
    first_reki_proj = await db.projects.find_one({"name": {"$regex": "Residence 24"}}, {"_id": 0, "id": 1}) or {}
    first_quotation = await db.quotations.find_one({}, {"_id": 0, "id": 1}) or {}
    return [
        {"id": "n1", "title": "BOQ V1 approved", "body": "Kohli Residence Interior BOQ was approved by client.", "at": now_iso(), "unread": True, "type": "boq",
         "entity_type": "boq", "entity_id": (first_boq or {}).get("id"),
         "link_url": f"/boq/{first_boq['id']}" if first_boq else "/boq"},
        {"id": "n2", "title": "Site Reki completed", "body": "Priya uploaded site reki report for Residence 24.", "at": now_iso(), "unread": True, "type": "site",
         "entity_type": "project", "entity_id": first_reki_proj.get("id"),
         "link_url": f"/projects/{first_reki_proj['id']}" if first_reki_proj.get("id") else "/documents"},
        {"id": "n3", "title": "Vendor quotation received", "body": "Flooring quotation received for Studio Office.", "at": now_iso(), "unread": True, "type": "quotation",
         "entity_type": "quotation", "entity_id": first_quotation.get("id"),
         "link_url": f"/quotations/{first_quotation['id']}" if first_quotation.get("id") else "/quotations"},
    ]


# ---------- Users (admin-only) ----------
class UserPatchIn(BaseModel):
    role: Optional[str] = None
    is_active: Optional[bool] = None


class UserCreateIn(BaseModel):
    name: str
    email: EmailStr
    role: str = "project_manager"


@api.get("/users")
async def users_list(current=Depends(require_admin)):
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).sort("created_at", 1).to_list(500)
    return users


@api.post("/users")
async def user_create(payload: UserCreateIn, current=Depends(require_admin)):
    if payload.role not in ALLOWED_ROLES:
        raise HTTPException(400, f"Invalid role. Allowed: {sorted(ALLOWED_ROLES)}")
    existing = await db.users.find_one({"email": payload.email.lower()})
    if existing:
        raise HTTPException(400, "Email already registered")
    temp_pw = _gen_temp_password()
    user = {
        "id": gen_id(),
        "name": payload.name,
        "email": payload.email.lower(),
        "password_hash": hash_password(temp_pw),
        "role": payload.role,
        "avatar_initials": initials_from_name(payload.name),
        "is_active": True,
        "created_at": now_iso(),
    }
    await db.users.insert_one(user)
    user.pop("_id", None)
    user.pop("password_hash", None)
    # Email is mocked — return temp password in response so admin can share it
    return {"user": user, "temp_password": temp_pw, "email_sent": False, "note": "Email delivery is mocked. Share the temporary password with the user manually."}


@api.patch("/users/{user_id}")
async def user_patch(user_id: str, payload: UserPatchIn, current=Depends(require_admin)):
    u = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not u:
        raise HTTPException(404, "User not found")
    upd: dict = {}
    if payload.role is not None:
        if payload.role not in ALLOWED_ROLES:
            raise HTTPException(400, f"Invalid role. Allowed: {sorted(ALLOWED_ROLES)}")
        upd["role"] = payload.role
    if payload.is_active is not None:
        upd["is_active"] = payload.is_active
    if not upd:
        return {k: v for k, v in u.items() if k != "password_hash"}
    upd["updated_at"] = now_iso()
    await db.users.update_one({"id": user_id}, {"$set": upd})
    fresh = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    return fresh


# ---------- BOQ Templates / Library / Activity (Phase E) ----------
@api.get("/boq/templates")
async def boq_templates_list(current=Depends(get_current_user)):
    tmpls = await db.boq_templates.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    for t in tmpls:
        cats = t.get("categories") or []
        t["category_count"] = len(cats)
        t["item_count"] = sum(len(c.get("items") or []) for c in cats)
        total = 0.0
        for c in cats:
            for it in (c.get("items") or []):
                q = float(it.get("quantity") or 0); r = float(it.get("rate", it.get("default_rate") or 0))
                total += q * r
        t["total_value"] = round(total, 2)
    # Seed 3 tier templates if missing
    if not any(t.get("template_tier") in ("essential", "premium", "luxury") for t in tmpls):
        _ = 0  # deferred inline below
    return tmpls


@api.get("/library/items")
async def library_items(q: Optional[str] = None, category_id: Optional[str] = None, current=Depends(get_current_user)):
    query: dict = {}
    if category_id: query["category_id"] = category_id
    if q:
        query["$or"] = [
            {"name": {"$regex": q, "$options": "i"}},
            {"description": {"$regex": q, "$options": "i"}},
            {"code": {"$regex": q, "$options": "i"}},
        ]
    rows = await db.library_items.find(query, {"_id": 0}).sort("name", 1).to_list(500)
    return rows


class LibraryItemIn(BaseModel):
    model_config = ConfigDict(extra="allow")


@api.post("/library/items")
async def library_item_create(payload: LibraryItemIn, current=Depends(get_current_user)):
    d = payload.model_dump(exclude_none=True)
    iid = gen_id()
    doc = {"id": iid, "name": d.get("name") or "Untitled",
           "category_id": d.get("category_id"), "category_name": d.get("category_name"),
           "unit": d.get("unit") or "Nos.", "default_rate": float(d.get("default_rate") or 0),
           "notes": d.get("notes") or "", "created_by": current.get("name"),
           "created_at": now_iso(), "updated_at": now_iso()}
    await db.library_items.insert_one(doc)
    await db.boq_activity.insert_one({"id": gen_id(), "action": "library_item_added", "target": doc["name"], "target_type": "library_item", "target_id": iid, "user": current.get("name"), "at": now_iso(), "details": f"Added library item {doc['name']} ({doc['unit']} @ ₹{doc['default_rate']})"})
    doc.pop("_id", None)
    return doc


@api.patch("/library/items/{iid}")
async def library_item_patch(iid: str, payload: LibraryItemIn, current=Depends(get_current_user)):
    existing = await db.library_items.find_one({"id": iid}, {"_id": 0})
    if not existing: raise HTTPException(404, "Library item not found")
    d = payload.model_dump(exclude_unset=True)
    d["updated_at"] = now_iso()
    await db.library_items.update_one({"id": iid}, {"$set": d})
    action = "library_item_moved" if "category_id" in d else "library_item_edited"
    await db.boq_activity.insert_one({"id": gen_id(), "action": action, "target": existing.get("name"), "target_type": "library_item", "target_id": iid, "user": current.get("name"), "at": now_iso(), "details": f"Updated {list(d.keys())}"})
    return await db.library_items.find_one({"id": iid}, {"_id": 0})


@api.delete("/library/items/{iid}")
async def library_item_delete(iid: str, current=Depends(get_current_user)):
    existing = await db.library_items.find_one({"id": iid}, {"_id": 0})
    if not existing: raise HTTPException(404, "Library item not found")
    # Warn if referenced by any unlocked BOQ (soft check — we do not block)
    ref = await db.boq_items.count_documents({"library_item_id": iid})
    await db.library_items.delete_one({"id": iid})
    await db.boq_activity.insert_one({"id": gen_id(), "action": "library_item_deleted", "target": existing.get("name"), "target_type": "library_item", "target_id": iid, "user": current.get("name"), "at": now_iso(), "details": f"Deleted (references: {ref})"})
    return {"ok": True, "referenced_by": ref}


@api.get("/library/categories")
async def library_categories(current=Depends(get_current_user)):
    rows = await db.library_categories.find({}, {"_id": 0}).sort("name", 1).to_list(200)
    for r in rows:
        r["item_count"] = await db.library_items.count_documents({"category_id": r["id"]})
    return rows


class LibraryCategoryIn(BaseModel):
    name: str
    code: Optional[str] = None


@api.post("/library/categories")
async def library_category_create(payload: LibraryCategoryIn, current=Depends(get_current_user)):
    cid = gen_id()
    doc = {"id": cid, "name": payload.name, "code": payload.code or payload.name[:2].upper(), "created_at": now_iso()}
    await db.library_categories.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.patch("/library/categories/{cid}")
async def library_category_patch(cid: str, payload: LibraryCategoryIn, current=Depends(get_current_user)):
    await db.library_categories.update_one({"id": cid}, {"$set": {"name": payload.name, "code": payload.code}})
    return await db.library_categories.find_one({"id": cid}, {"_id": 0})


@api.delete("/library/categories/{cid}")
async def library_category_delete(cid: str, current=Depends(get_current_user)):
    ref = await db.library_items.count_documents({"category_id": cid})
    await db.library_categories.delete_one({"id": cid})
    return {"ok": True, "referenced_by": ref}


@api.get("/boq/activity")
async def boq_activity_list(limit: int = 100, user: Optional[str] = None, action: Optional[str] = None,
                             date_from: Optional[str] = None, date_to: Optional[str] = None,
                             current=Depends(get_current_user)):
    q: dict = {}
    if user: q["user"] = user
    if action: q["action"] = action
    if date_from or date_to:
        q["at"] = {}
        if date_from: q["at"]["$gte"] = date_from
        if date_to: q["at"]["$lte"] = date_to
    rows = await db.boq_activity.find(q, {"_id": 0}).sort("at", -1).to_list(limit)
    return rows



# ---------- Estimates / Signature (Phase G) ----------
@api.get("/estimate/next-number")
async def estimate_next_number(project_id: str, date: str, current=Depends(get_current_user)):
    project = await db.projects.find_one({"id": project_id}, {"_id": 0, "name": 1})
    prefix = "ES"
    if project and project.get("name"):
        alpha = "".join(ch for ch in project["name"].upper() if ch.isalpha())
        prefix += "-" + (alpha[:2] if alpha else "XX")
    else:
        prefix += "-XX"
    try:
        y, m, d = date.split("-")
        date_part = f"{d}-{m}-{y}"
    except Exception:
        _t = datetime.now(timezone.utc)
        date_part = f"{_t.day:02d}-{_t.month:02d}-{_t.year}"
    count = await db.quotations.count_documents({"project_id": project_id})
    nn = f"{(count + 1):02d}"
    return {"estimate_number": f"{prefix}-{date_part}-{nn}"}


@api.post("/users/me/signature")
async def user_upload_signature(payload: dict = Body(...), current=Depends(require_admin)):
    name = (payload.get("name") or current.get("name") or "").strip()
    image_b64 = payload.get("image_b64") or ""
    mime = payload.get("mime") or "image/png"
    if not image_b64:
        raise HTTPException(400, "Signature image required")
    data_url = f"data:{mime};base64,{image_b64}"
    await db.users.update_one({"id": current.get("id")}, {"$set": {
        "estimate_signature_name": name,
        "estimate_signature_url": data_url,
        "estimate_signature_updated_at": now_iso(),
    }})
    return {"ok": True, "name": name, "signature_url": data_url}


@api.post("/quotations/{qid}/approve-with-signature")
async def quotation_approve_with_sig(qid: str, current=Depends(require_admin)):
    q = await db.quotations.find_one({"id": qid}, {"_id": 0})
    if not q: raise HTTPException(404, "Estimate not found")
    user = await db.users.find_one({"id": current.get("id")}, {"_id": 0}) or {}
    await db.quotations.update_one({"id": qid}, {"$set": {
        "status": "approved", "locked": True,
        "approved_by_user_id": current.get("id"),
        "approved_by": user.get("estimate_signature_name") or current.get("name"),
        "approved_by_name": user.get("estimate_signature_name") or current.get("name"),
        "approved_by_signature_url": user.get("estimate_signature_url"),
        "approved_at": now_iso(),
        "updated_at": now_iso(),
    }})
    fresh = await db.quotations.find_one({"id": qid}, {"_id": 0})
    return fresh


# ---------- Project Types / Phases / Activity (Phase F Track 3) ----------
PROJECT_PHASES_SPEC = [
    {"key": "documentation", "name": "Documentation", "subphases": [
        {"key": "brief", "name": "Brief", "category": "brief"},
        {"key": "pitch", "name": "Pitch", "category": "pitch"},
        {"key": "presentation", "name": "Presentation", "category": "presentation"},
        {"key": "scope_of_work", "name": "Scope of Work", "category": "scope_of_work"},
        {"key": "site_reki", "name": "Site Reki", "category": "site_reki"},
        {"key": "agreement", "name": "Agreement", "category": "agreement"},
        {"key": "boq", "name": "BOQ", "category": "boq"},
    ]},
    # Design — single-level phase; auto-completes on any finalized `design` category doc.
    {"key": "design", "name": "Design", "category": "design", "subphases": []},
    {"key": "execution", "name": "Execution", "subphases": [
        {"key": "estimates", "name": "Estimates", "category": "estimate"},
        {"key": "drawings", "name": "Drawings", "category": "drawing"},
    ]},
    # Handover — single-level phase; auto-completes on `handover_certificate`.
    {"key": "handover", "name": "Handover", "category": "handover_certificate", "subphases": []},
]


@api.get("/project-types")
async def project_types_list(current=Depends(get_current_user)):
    rows = await db.project_types.find({}, {"_id": 0}).sort("name", 1).to_list(200)
    if not rows:
        distinct = await db.projects.distinct("type")
        for t in [d for d in distinct if d]:
            await db.project_types.insert_one({"id": gen_id(), "name": t, "created_by": "System", "created_at": now_iso()})
        rows = await db.project_types.find({}, {"_id": 0}).sort("name", 1).to_list(200)
    return rows


class ProjectTypeIn(BaseModel):
    name: str


@api.post("/project-types")
async def project_type_create(payload: ProjectTypeIn, current=Depends(get_current_user)):
    if current.get("role") not in ("admin", "project_manager"):
        raise HTTPException(403, "Only admin or project_manager can add project types")
    name = payload.name.strip()
    if not name:
        raise HTTPException(400, "Name required")
    existing = await db.project_types.find_one({"name": {"$regex": f"^{name}$", "$options": "i"}}, {"_id": 0})
    if existing:
        return existing
    doc = {"id": gen_id(), "name": name, "created_by": current.get("name"), "created_at": now_iso()}
    await db.project_types.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.get("/projects/activity")
async def project_activity_list(limit: int = 100, user: Optional[str] = None, action: Optional[str] = None,
                                 date_from: Optional[str] = None, date_to: Optional[str] = None,
                                 current=Depends(get_current_user)):
    q: dict = {}
    if user: q["user"] = user
    if action: q["action"] = action
    if date_from or date_to:
        q["at"] = {}
        if date_from: q["at"]["$gte"] = date_from
        if date_to: q["at"]["$lte"] = date_to
    rows = await db.project_activity.find(q, {"_id": 0}).sort("at", -1).to_list(limit)
    if not rows:
        projects = await db.projects.find({}, {"_id": 0, "id": 1, "name": 1}).limit(4).to_list(4)
        if projects:
            base = datetime.now(timezone.utc)
            seed_rows = []
            for i, p in enumerate(projects):
                seed_rows.append({"id": gen_id(), "project_id": p["id"], "action": "created", "target": p["name"], "target_type": "project", "user": "Deepak Rao", "details": f"Project '{p['name']}' created", "at": (base - timedelta(days=6-i)).isoformat()})
                seed_rows.append({"id": gen_id(), "project_id": p["id"], "action": "document_finalized", "target": f"BOQ — {p['name']}", "target_type": "document", "user": "Priya Sharma", "details": "Finalized BOQ document", "at": (base - timedelta(days=4-i, hours=3)).isoformat()})
            for s in seed_rows:
                await db.project_activity.insert_one(s)
            rows = await db.project_activity.find({}, {"_id": 0}).sort("at", -1).to_list(limit)
    return rows


@api.get("/projects/{project_id}/phases")
async def project_phases(project_id: str, current=Depends(get_current_user)):
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(404, "Project not found")
    finalized = set()
    async for d in db.documents.find({"project_id": project_id, "is_locked": True}, {"_id": 0, "category": 1}):
        c = (d.get("category") or "").strip().lower().replace(" ", "_")
        finalized.add(c)
    state = project.get("phase_state") or {}  # {key: {manually_completed: bool}}
    result = []
    total_units = 0   # each sub-phase, OR the phase itself if no sub-phases
    done_units = 0
    for phase in PROJECT_PHASES_SPEC:
        subs_out = []
        for sub in phase["subphases"]:
            auto = sub["category"] in finalized
            manual = bool((state.get(sub["key"]) or {}).get("manually_completed"))
            subs_out.append({**sub, "auto_completed": auto, "manually_completed": manual, "complete": auto or manual})
            total_units += 1
            if auto or manual: done_units += 1
        if phase["subphases"]:
            subs_complete = sum(1 for s in subs_out if s["complete"])
            complete = subs_complete == len(subs_out)
            auto_p = False
        else:
            # single-level phase: its own category
            cat = (phase.get("category") or "").strip().lower().replace(" ", "_")
            auto_p = bool(cat) and cat in finalized
            manual_p = bool((state.get(phase["key"]) or {}).get("manually_completed"))
            complete = auto_p or manual_p
            subs_complete = 1 if complete else 0
            total_units += 1
            if complete: done_units += 1
        result.append({
            "key": phase["key"], "name": phase["name"], "subphases": subs_out,
            "subphase_count": len(subs_out), "completed_subphases": subs_complete,
            "auto_completed": auto_p if not phase["subphases"] else None,
            "manually_completed": bool((state.get(phase["key"]) or {}).get("manually_completed")),
            "category": phase.get("category") if not phase["subphases"] else None,
            "complete": complete,
        })
    pct = round((done_units / total_units) * 100) if total_units else 0
    return {"phases": result, "progress_pct": pct, "completed_subphases": done_units, "total_subphases": total_units}


class PhasePatchIn(BaseModel):
    manually_completed: bool


@api.patch("/projects/{project_id}/phases/{key}")
async def project_phase_patch(project_id: str, key: str, payload: PhasePatchIn, current=Depends(get_current_user)):
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(404, "Project not found")
    # Validate key exists in spec (phase or sub-phase)
    valid_keys = set()
    for p in PROJECT_PHASES_SPEC:
        valid_keys.add(p["key"])
        for s in p["subphases"]:
            valid_keys.add(s["key"])
    if key not in valid_keys:
        raise HTTPException(400, f"Unknown phase key '{key}'")
    # Guard: cannot un-mark auto-completed sub-phase
    if payload.manually_completed is False:
        finalized = set()
        async for d in db.documents.find({"project_id": project_id, "is_locked": True}, {"_id": 0, "category": 1}):
            c = (d.get("category") or "").strip().lower().replace(" ", "_")
            finalized.add(c)
        for p in PROJECT_PHASES_SPEC:
            if p["key"] == key and not p["subphases"] and (p.get("category") or "") in finalized:
                raise HTTPException(400, "Cannot unmark — this phase is auto-completed via a finalized document.")
            for s in p["subphases"]:
                if s["key"] == key and s["category"] in finalized:
                    raise HTTPException(400, "Cannot unmark — this sub-phase is auto-completed via a finalized document.")
    await db.projects.update_one(
        {"id": project_id},
        {"$set": {f"phase_state.{key}.manually_completed": payload.manually_completed,
                  f"phase_state.{key}.updated_by": current.get("name"),
                  f"phase_state.{key}.updated_at": now_iso()}},
    )
    await db.project_activity.insert_one({
        "id": gen_id(), "project_id": project_id,
        "action": "phase_marked_complete" if payload.manually_completed else "phase_marked_incomplete",
        "target": key, "target_type": "phase", "user": current.get("name"),
        "details": f"Manual toggle for '{key}' → {payload.manually_completed}", "at": now_iso(),
    })
    return {"ok": True, "key": key, "manually_completed": payload.manually_completed}



@api.get("/projects/{project_id}/status-checklist")
async def project_status_checklist(project_id: str, current=Depends(get_current_user)):
    order = [
        ("brief", "Brief", ["brief", "project brief"]),
        ("boq", "BOQ", ["boq", "boqs"]),
        ("estimate", "Estimate", ["estimate", "quotation", "quotations"]),
        ("pitch", "Pitch", ["pitch"]),
        ("presentation", "Presentation", ["presentation"]),
        ("3d_views", "3D", ["3d_views", "3d"]),
        ("scope_of_work", "Scope of Work", ["scope_of_work", "scope of work"]),
        ("site_reki", "Site Reki", ["site_reki", "site reki"]),
        ("agreement", "Agreement", ["agreement", "agreements"]),
        ("design", "Design", ["design"]),
        ("drawing", "Drawings", ["drawing", "drawings", "gfc drawings"]),
        ("handover_certificate", "Handover", ["handover_certificate", "handover documents", "handover"]),
    ]
    project = await db.projects.find_one({"id": project_id}, {"_id": 0, "id": 1})
    if not project:
        raise HTTPException(404, "Project not found")
    items = []
    done = 0
    for key, label, cats in order:
        normalized = [c.lower().strip() for c in cats]
        found = await db.documents.find_one({
            "project_id": project_id,
            "is_locked": True,
            "is_archived": {"$ne": True},
            "$expr": {"$in": [{"$toLower": {"$ifNull": ["$category", ""]}}, normalized]},
        }, {"_id": 0, "id": 1})
        completed = bool(found)
        if completed: done += 1
        items.append({"key": key, "label": label, "completed": completed})
    total = len(items)
    return {"items": items, "completed_count": done, "total": total, "progress_pct": round((done / total) * 100) if total else 0}



    # 3 tier templates
    if await db.boq_templates.count_documents({"template_tier": {"$exists": True}}) == 0:
        tiers = [
            ("essential", "Essential Interior Package", [
                ("A", "Civil & Masonry", [("Wall plaster", "Sqm", 250), ("Floor levelling", "Sqm", 180)]),
                ("B", "Electrical", [("Point wiring", "Point", 850), ("MCB DB (8-way)", "Nos.", 4200)]),
            ]),
            ("premium", "Premium Interior Package", [
                ("A", "Civil & Masonry", [("Wall plaster", "Sqm", 320), ("Floor levelling", "Sqm", 240)]),
                ("B", "Electrical", [("Point wiring premium", "Point", 1100), ("Modular DB", "Nos.", 6800)]),
                ("C", "Furniture & Joinery", [("Wardrobe HDMR", "Sqft", 1450), ("TV Unit", "Rft", 5200)]),
            ]),
            ("luxury", "Luxury Interior Package", [
                ("A", "Civil & Masonry", [("Wall plaster premium", "Sqm", 420), ("Italian marble laying", "Sqm", 850)]),
                ("B", "Electrical", [("Smart wiring", "Point", 1800), ("Home automation panel", "Nos.", 42000)]),
                ("C", "Furniture & Joinery", [("Veneer wardrobe", "Sqft", 2400), ("Custom TV wall", "Rft", 9800)]),
                ("D", "Lighting & Decor", [("Chandelier", "Nos.", 85000), ("LED cove lighting", "Rft", 850)]),
            ]),
        ]
        for tier, name, cats in tiers:
            cat_docs = []
            for code, cname, items in cats:
                cat_docs.append({
                    "code": code, "name": cname,
                    "items": [{"name": n, "unit": u, "quantity": 1, "rate": r, "default_rate": r, "calc_type": "M"} for (n, u, r) in items],
                })
            await db.boq_templates.insert_one({
                "id": gen_id(), "name": name, "template_tier": tier, "is_template": True,
                "categories": cat_docs, "created_by": "System", "created_at": now_iso(), "updated_at": now_iso(),
            })
    # Library — copy from catalog if empty
    if await db.library_items.count_documents({}) == 0:
        cats = await db.boq_catalog_categories.find({}, {"_id": 0}).to_list(200)
        for c in cats:
            cid = gen_id()
            await db.library_categories.insert_one({"id": cid, "name": c.get("name"), "code": c.get("code"), "created_at": now_iso()})
            for it in (c.get("items") or [])[:8]:
                await db.library_items.insert_one({
                    "id": gen_id(), "name": it.get("description") or it.get("name") or "Item",
                    "category_id": cid, "category_name": c.get("name"),
                    "unit": it.get("unit") or "Nos.",
                    "default_rate": float(it.get("rate", it.get("default_rate") or 0)),
                    "notes": "", "created_by": "System",
                    "created_at": now_iso(), "updated_at": now_iso(),
                })
    # Seed a few activity rows if empty
    if await db.boq_activity.count_documents({}) == 0:
        base = datetime.now(timezone.utc)
        seed = [
            {"action": "created", "target": "BOQ V1 — Kohli Residence", "target_type": "boq", "user": "Deepak Rao", "details": "New BOQ created from Premium template", "at": (base - timedelta(days=2)).isoformat()},
            {"action": "item_added", "target": "Wall plaster", "target_type": "boq_item", "user": "Deepak Rao", "details": "Added to Civil & Masonry (₹320/Sqm)", "at": (base - timedelta(days=1, hours=8)).isoformat()},
            {"action": "rate_changed", "target": "Point wiring", "target_type": "boq_item", "user": "Priya Sharma", "details": "Rate changed from ₹850 to ₹950", "at": (base - timedelta(days=1, hours=4)).isoformat()},
            {"action": "approved", "target": "BOQ V1 — Kohli Residence", "target_type": "boq", "user": "Deepak Rao", "details": "Approved by client", "at": (base - timedelta(hours=12)).isoformat()},
            {"action": "template_created", "target": "Luxury Interior Package", "target_type": "template", "user": "Deepak Rao", "details": "New template with 4 categories", "at": (base - timedelta(hours=6)).isoformat()},
            {"action": "library_item_added", "target": "Home automation panel", "target_type": "library_item", "user": "Deepak Rao", "details": "Added to Electrical library", "at": (base - timedelta(hours=2)).isoformat()},
            {"action": "category_moved", "target": "Modular DB", "target_type": "library_item", "user": "Priya Sharma", "details": "Moved from Electrical to Premium Electrical", "at": (base - timedelta(hours=1)).isoformat()},
        ]
        for s in seed:
            await db.boq_activity.insert_one({"id": gen_id(), **s})



async def seed_if_empty():
    # Users
    if await db.users.count_documents({}) == 0:
        demo_users = [
            {"name": "Deepak Rao", "email": "admin@buildcon.in", "password": "buildcon123", "role": "admin"},
            {"name": "Mr. Kohli", "email": "client@kohli.in", "password": "client123", "role": "client"},
        ]
        for u in demo_users:
            await db.users.insert_one({
                "id": gen_id(),
                "name": u["name"],
                "email": u["email"],
                "password_hash": hash_password(u["password"]),
                "role": u["role"],
                "avatar_initials": initials_from_name(u["name"]),
                "created_at": now_iso(),
            })

    if await db.projects.count_documents({}) == 0:
        await _seed_projects_and_docs()

    if await db.boq_catalog_categories.count_documents({}) == 0:
        await _seed_boq_catalog_and_kohli()

    # Phase E seed (idempotent)
    try:
        await _seed_boq_phase_e()
    except Exception as _e:
        logger.warning(f"Phase E seed skipped: {_e}")


async def _seed_projects_and_docs():
    clients_seed = [
        {"id": gen_id(), "name": "Mr. & Mrs. Kohli", "email": "kohli.family@example.in", "phone": "+91 98100 12345", "location": "Shalimar Bagh, New Delhi"},
        {"id": gen_id(), "name": "Anaya & Rohan Mehra", "email": "mehra@example.in", "phone": "+91 98999 23145", "location": "New Delhi"},
        {"id": gen_id(), "name": "Jain Art Press Pvt Ltd", "email": "contact@jainartpress.in", "phone": "+91 98110 55123", "location": "Okhla, New Delhi"},
        {"id": gen_id(), "name": "Studio Nine Consulting", "email": "hello@studionine.in", "phone": "+91 99100 88231", "location": "Gurugram"},
        {"id": gen_id(), "name": "Priyanka Verma", "email": "priyanka.v@example.in", "phone": "+91 98186 44112", "location": "Vasant Kunj, New Delhi"},
        {"id": gen_id(), "name": "Bansal Family", "email": "bansal@example.in", "phone": "+91 98181 99332", "location": "Noida"},
    ]
    for c in clients_seed:
        await db.clients.insert_one(c)

    # Projects
    projects_seed = [
        {
            "id": gen_id(), "name": "Kohli Residence — Interior Renovation",
            "client_id": clients_seed[0]["id"], "client_name": "Mr. & Mrs. Kohli",
            "location": "Shalimar Bagh, New Delhi", "type": "Residential Interior",
            "current_phase": "Pre-Execution", "current_milestone": "Finalise vendor selection",
            "progress": 62, "timeline_status": "on_track",
            "expected_completion": "2026-09-20", "next_deadline": "2026-03-04",
            "status": "active",
        },
        {
            "id": gen_id(), "name": "The House Within",
            "client_id": clients_seed[1]["id"], "client_name": "Anaya & Rohan Mehra",
            "location": "New Delhi", "type": "Residential Interior",
            "current_phase": "Design", "current_milestone": "Complete client brief",
            "progress": 38, "timeline_status": "on_track",
            "expected_completion": "2026-12-18", "next_deadline": "2026-03-02",
            "status": "active",
        },
        {
            "id": gen_id(), "name": "Jain Art Press",
            "client_id": clients_seed[2]["id"], "client_name": "Jain Art Press Pvt Ltd",
            "location": "Okhla, New Delhi", "type": "Commercial Interior",
            "current_phase": "Design", "current_milestone": "Submit BOQ",
            "progress": 28, "timeline_status": "at_risk",
            "expected_completion": "2026-11-10", "next_deadline": "2026-03-05",
            "status": "active",
        },
        {
            "id": gen_id(), "name": "Studio Office",
            "client_id": clients_seed[3]["id"], "client_name": "Studio Nine Consulting",
            "location": "Gurugram", "type": "Office Fit-out",
            "current_phase": "Execution", "current_milestone": "Flooring installation",
            "progress": 71, "timeline_status": "delayed",
            "expected_completion": "2026-05-14", "next_deadline": "2026-03-07",
            "status": "active",
        },
        {
            "id": gen_id(), "name": "Residence 24",
            "client_id": clients_seed[4]["id"], "client_name": "Priyanka Verma",
            "location": "Vasant Kunj, New Delhi", "type": "Residential Interior",
            "current_phase": "Pre-Design", "current_milestone": "Finalise site reki",
            "progress": 15, "timeline_status": "on_track",
            "expected_completion": "2026-10-30", "next_deadline": "2026-03-04",
            "status": "active",
        },
        {
            "id": gen_id(), "name": "Bansal Villa",
            "client_id": clients_seed[5]["id"], "client_name": "Bansal Family",
            "location": "Noida", "type": "Residential Interior",
            "current_phase": "Handover", "current_milestone": "Handover checklist",
            "progress": 94, "timeline_status": "on_track",
            "expected_completion": "2026-03-15", "next_deadline": "2026-03-12",
            "status": "active",
        },
    ]
    for p in projects_seed:
        await db.projects.insert_one(p)

    kohli = projects_seed[0]
    house_within = projects_seed[1]
    jain = projects_seed[2]
    studio = projects_seed[3]
    res24 = projects_seed[4]
    bansal = projects_seed[5]

    # BOQs — seeded by _seed_boq_catalog_and_kohli(); leave placeholder
    _projects_for_boq = {"kohli": kohli, "res24": res24, "jain": jain, "house_within": house_within, "studio": studio, "bansal": bansal}
    # (comprehensive BOQ + catalog + Kohli items seeded separately)

    # Milestones
    base = datetime.now(timezone.utc)
    milestones_seed = [
        {"id": gen_id(), "title": "Complete Client Brief", "project_id": house_within["id"], "project_name": house_within["name"], "assignee": "Deepak Rao", "assignee_initials": "DR", "due_at": (base + timedelta(days=2)).isoformat(), "status": "pending"},
        {"id": gen_id(), "title": "Finalise Site Reki", "project_id": res24["id"], "project_name": res24["name"], "assignee": "Priya Sharma", "assignee_initials": "PS", "due_at": (base + timedelta(days=4)).isoformat(), "status": "pending"},
        {"id": gen_id(), "title": "Submit BOQ", "project_id": jain["id"], "project_name": jain["name"], "assignee": "Arjun Malik", "assignee_initials": "AM", "due_at": (base + timedelta(days=5)).isoformat(), "status": "pending"},
        {"id": gen_id(), "title": "Upload Scope of Work", "project_id": studio["id"], "project_name": studio["name"], "assignee": "Neha Kapoor", "assignee_initials": "NK", "due_at": (base + timedelta(days=7)).isoformat(), "status": "pending"},
        {"id": gen_id(), "title": "Vendor Selection Freeze", "project_id": kohli["id"], "project_name": kohli["name"], "assignee": "Deepak Rao", "assignee_initials": "DR", "due_at": (base + timedelta(days=6)).isoformat(), "status": "pending"},
    ]
    for m in milestones_seed:
        await db.milestones.insert_one(m)

    # Documents
    docs_seed = [
        {"id": gen_id(), "name": "Agreement — Kohli Residence.pdf", "type": "pdf", "project_id": kohli["id"], "project_name": kohli["name"], "category": "Agreement", "uploaded_by": "Deepak Rao", "uploader_initials": "DR", "uploaded_at": (base - timedelta(hours=3)).isoformat(), "origin": "generated", "status": "final"},
        {"id": gen_id(), "name": "Client Brief — House Within.docx", "type": "doc", "project_id": house_within["id"], "project_name": house_within["name"], "category": "Brief", "uploaded_by": "Anaya Mehra", "uploader_initials": "AM", "uploaded_at": (base - timedelta(hours=8)).isoformat(), "origin": "uploaded", "status": "pending"},
        {"id": gen_id(), "name": "Site Reki Report — Residence 24.pdf", "type": "pdf", "project_id": res24["id"], "project_name": res24["name"], "category": "Site Reki", "uploaded_by": "Priya Sharma", "uploader_initials": "PS", "uploaded_at": (base - timedelta(days=1)).isoformat(), "origin": "uploaded", "status": "final"},
        {"id": gen_id(), "name": "Pitch Deck — Jain Art Press.pdf", "type": "pdf", "project_id": jain["id"], "project_name": jain["name"], "category": "Pitch", "uploaded_by": "Arjun Malik", "uploader_initials": "AM", "uploaded_at": (base - timedelta(days=1, hours=6)).isoformat(), "origin": "generated", "status": "final"},
        {"id": gen_id(), "name": "SOW — Studio Office.pdf", "type": "pdf", "project_id": studio["id"], "project_name": studio["name"], "category": "SOW", "uploaded_by": "Neha Kapoor", "uploader_initials": "NK", "uploaded_at": (base - timedelta(days=2)).isoformat(), "origin": "generated", "status": "final"},
        {"id": gen_id(), "name": "BOQ V1 — Kohli Residence.xlsx", "type": "xls", "project_id": kohli["id"], "project_name": kohli["name"], "category": "BOQ", "uploaded_by": "Deepak Rao", "uploader_initials": "DR", "uploaded_at": (base - timedelta(days=2, hours=4)).isoformat(), "origin": "generated", "status": "final"},
        {"id": gen_id(), "name": "Vendor Quotation — Flooring.pdf", "type": "pdf", "project_id": studio["id"], "project_name": studio["name"], "category": "Quotation", "uploaded_by": "Rahul Gupta", "uploader_initials": "RG", "uploaded_at": (base - timedelta(days=3)).isoformat(), "origin": "uploaded", "status": "pending"},
        {"id": gen_id(), "name": "Design Freeze — Bansal Villa.pdf", "type": "pdf", "project_id": bansal["id"], "project_name": bansal["name"], "category": "Design", "uploaded_by": "Deepak Rao", "uploader_initials": "DR", "uploaded_at": (base - timedelta(days=4)).isoformat(), "origin": "generated", "status": "final"},
        {"id": gen_id(), "name": "Purchase Order — Kohli Kitchen.pdf", "type": "pdf", "project_id": kohli["id"], "project_name": kohli["name"], "category": "PO", "uploaded_by": "Deepak Rao", "uploader_initials": "DR", "uploaded_at": (base - timedelta(days=5)).isoformat(), "origin": "generated", "status": "final"},
        {"id": gen_id(), "name": "Client Feedback — Bansal.pdf", "type": "pdf", "project_id": bansal["id"], "project_name": bansal["name"], "category": "Feedback", "uploaded_by": "Bansal Family", "uploader_initials": "BF", "uploaded_at": (base - timedelta(days=6)).isoformat(), "origin": "uploaded", "status": "pending"},
    ]
    for d in docs_seed:
        await db.documents.insert_one(d)

    # Activity
    activity_seed = [
        {"id": gen_id(), "actor": "Deepak Rao", "actor_initials": "DR", "action": "approved", "target": "BOQ V1 — Kohli Residence", "project_id": kohli["id"], "project_name": kohli["name"], "at": (base - timedelta(minutes=25)).isoformat(), "status": "success"},
        {"id": gen_id(), "actor": "Priya Sharma", "actor_initials": "PS", "action": "completed", "target": "Site Reki — Residence 24", "project_id": res24["id"], "project_name": res24["name"], "at": (base - timedelta(hours=2)).isoformat(), "status": "success"},
        {"id": gen_id(), "actor": "Rahul Gupta", "actor_initials": "RG", "action": "uploaded", "target": "Vendor Quotation — Flooring", "project_id": studio["id"], "project_name": studio["name"], "at": (base - timedelta(hours=4)).isoformat(), "status": "info"},
        {"id": gen_id(), "actor": "Arjun Malik", "actor_initials": "AM", "action": "created", "target": "BOQ Draft — Jain Art Press", "project_id": jain["id"], "project_name": jain["name"], "at": (base - timedelta(hours=6)).isoformat(), "status": "info"},
        {"id": gen_id(), "actor": "Neha Kapoor", "actor_initials": "NK", "action": "commented", "target": "SOW — Studio Office", "project_id": studio["id"], "project_name": studio["name"], "at": (base - timedelta(hours=9)).isoformat(), "status": "info"},
        {"id": gen_id(), "actor": "Deepak Rao", "actor_initials": "DR", "action": "assigned", "target": "Vendor Selection — Kohli", "project_id": kohli["id"], "project_name": kohli["name"], "at": (base - timedelta(hours=14)).isoformat(), "status": "info"},
        {"id": gen_id(), "actor": "Anaya Mehra", "actor_initials": "AM", "action": "uploaded", "target": "Client Brief — House Within", "project_id": house_within["id"], "project_name": house_within["name"], "at": (base - timedelta(days=1)).isoformat(), "status": "info"},
        {"id": gen_id(), "actor": "Priya Sharma", "actor_initials": "PS", "action": "flagged", "target": "BOQ deadline — Jain Art Press", "project_id": jain["id"], "project_name": jain["name"], "at": (base - timedelta(days=1, hours=3)).isoformat(), "status": "warning"},
        {"id": gen_id(), "actor": "Deepak Rao", "actor_initials": "DR", "action": "shared", "target": "Handover Checklist — Bansal Villa", "project_id": bansal["id"], "project_name": bansal["name"], "at": (base - timedelta(days=1, hours=8)).isoformat(), "status": "info"},
        {"id": gen_id(), "actor": "Rahul Gupta", "actor_initials": "RG", "action": "reviewed", "target": "BOQ V2 — Studio Office", "project_id": studio["id"], "project_name": studio["name"], "at": (base - timedelta(days=2)).isoformat(), "status": "success"},
    ]
    for a in activity_seed:
        await db.activity.insert_one(a)

    # Calendar
    events_seed = [
        {"id": gen_id(), "title": "Site Visit — Residence 24", "type": "site_visit", "project_name": res24["name"], "at": (base + timedelta(days=1, hours=3)).isoformat(), "duration_min": 90},
        {"id": gen_id(), "title": "Client Meeting — Kohli", "type": "meeting", "project_name": kohli["name"], "at": (base + timedelta(days=2, hours=6)).isoformat(), "duration_min": 60},
        {"id": gen_id(), "title": "BOQ Deadline — Jain Art Press", "type": "deadline", "project_name": jain["name"], "at": (base + timedelta(days=3)).isoformat(), "duration_min": 0},
        {"id": gen_id(), "title": "Vendor Discussion — Flooring", "type": "meeting", "project_name": studio["name"], "at": (base + timedelta(days=4, hours=4)).isoformat(), "duration_min": 45},
        {"id": gen_id(), "title": "Design Presentation — House Within", "type": "meeting", "project_name": house_within["name"], "at": (base + timedelta(days=5, hours=5)).isoformat(), "duration_min": 60},
    ]
    for e in events_seed:
        await db.calendar_events.insert_one(e)

    # Vendors
    vendors_seed = [
        {"id": gen_id(), "name": "Delhi Marble House", "category": "Flooring", "rating": 4.6, "city": "New Delhi"},
        {"id": gen_id(), "name": "SunLite Electricals", "category": "Electrical", "rating": 4.4, "city": "New Delhi"},
        {"id": gen_id(), "name": "Woodcraft Interiors", "category": "Furniture & Joinery", "rating": 4.7, "city": "Gurugram"},
        {"id": gen_id(), "name": "AquaFit Plumbing", "category": "Plumbing", "rating": 4.3, "city": "Noida"},
        {"id": gen_id(), "name": "Modula Kitchens", "category": "Kitchen", "rating": 4.8, "city": "New Delhi"},
        {"id": gen_id(), "name": "Prism Paints & Polish", "category": "Painting", "rating": 4.2, "city": "Noida"},
        {"id": gen_id(), "name": "Gypcore False Ceiling", "category": "False Ceiling", "rating": 4.5, "city": "Gurugram"},
        {"id": gen_id(), "name": "Lumen Studio Lighting", "category": "Lighting", "rating": 4.6, "city": "New Delhi"},
    ]
    for v in vendors_seed:
        await db.vendors.insert_one(v)

    # Quotations
    quotations_seed = [
        {"id": gen_id(), "project_id": studio["id"], "project_name": studio["name"], "vendor": "Delhi Marble House", "category": "Flooring", "amount": 385000, "status": "pending", "created_at": now_iso()},
        {"id": gen_id(), "project_id": kohli["id"], "project_name": kohli["name"], "vendor": "Modula Kitchens", "category": "Kitchen", "amount": 320000, "status": "approved", "created_at": now_iso()},
        {"id": gen_id(), "project_id": kohli["id"], "project_name": kohli["name"], "vendor": "Woodcraft Interiors", "category": "Furniture", "amount": 812000, "status": "pending", "created_at": now_iso()},
        {"id": gen_id(), "project_id": bansal["id"], "project_name": bansal["name"], "vendor": "Prism Paints & Polish", "category": "Painting", "amount": 168000, "status": "approved", "created_at": now_iso()},
        {"id": gen_id(), "project_id": res24["id"], "project_name": res24["name"], "vendor": "SunLite Electricals", "category": "Electrical", "amount": 214000, "status": "pending", "created_at": now_iso()},
    ]
    for q in quotations_seed:
        await db.quotations.insert_one(q)

    # Continue working (for demo user + everyone)
    cw_seed = [
        {"id": gen_id(), "kind": "boq", "title": "Residence 24 — Interior BOQ", "project_name": res24["name"], "subtitle": "BOQ draft in progress", "progress": 68, "updated_at": (base - timedelta(minutes=25)).isoformat(), "cta_label": "Continue BOQ", "cta_href": "/boq"},
        {"id": gen_id(), "kind": "site_reki", "title": "Site Reki — The House Within", "project_name": house_within["name"], "subtitle": "Reki report drafting", "progress": 40, "updated_at": (base - timedelta(hours=2)).isoformat(), "cta_label": "Resume Reki", "cta_href": "/projects"},
        {"id": gen_id(), "kind": "quotation", "title": "Vendor Quotation — Flooring Package", "project_name": studio["name"], "subtitle": "Comparing 3 vendor quotes", "progress": 55, "updated_at": (base - timedelta(hours=5)).isoformat(), "cta_label": "Review Quotes", "cta_href": "/quotations"},
    ]
    for c in cw_seed:
        await db.continue_working.insert_one(c)


# ============================================================
# ================== PHASE 2 — BOQ MODULE ====================
# ============================================================

BOQ_STATUS = {"draft", "in_progress", "awaiting_approval", "returned", "approved", "final", "archived"}


def compute_item_amount(item: dict) -> float:
    if item.get("calc_type") == "L":
        return float(item.get("amount") or 0)
    q = float(item.get("quantity") or 0)
    r = float(item.get("rate") or 0)
    return round(q * r, 2)


def compute_formula_qty(detail: dict) -> Optional[float]:
    if not detail:
        return None
    formula = detail.get("formula")
    L = float(detail.get("length") or 0)
    W = float(detail.get("width") or 0)
    H = float(detail.get("height") or 0)
    D = float(detail.get("depth") or 0)
    R = float(detail.get("repetitions") or 1)
    ded = float(detail.get("deduction") or 0)
    waste_pct = float(detail.get("wastage_pct") or 0)
    count = float(detail.get("count") or 0)
    std = float(detail.get("std_qty") or 0)
    base = None
    if formula == "LxW":
        base = L * W
    elif formula == "LxWxH":
        base = L * W * H
    elif formula == "LxWxD":
        base = L * W * D
    elif formula == "CountxStd":
        base = count * std
    elif formula == "Running":
        base = L
    else:
        return None
    if base is None:
        return None
    q = base * R - ded
    if waste_pct:
        q = q * (1 + waste_pct / 100.0)
    return round(q, 2)


async def _recompute_totals(boq_id: str) -> dict:
    """Recompute category subtotals + boq project total + final total."""
    boq = await db.boqs.find_one({"id": boq_id}, {"_id": 0})
    if not boq:
        raise HTTPException(404, "BOQ not found")
    categories = await db.boq_categories.find({"boq_id": boq_id}, {"_id": 0}).sort("order", 1).to_list(200)
    items = await db.boq_items.find({"boq_id": boq_id}, {"_id": 0}).to_list(2000)

    by_cat: Dict[str, float] = {}
    for it in items:
        amt = compute_item_amount(it)
        by_cat[it["category_id"]] = by_cat.get(it["category_id"], 0) + amt

    project_total = 0.0
    for c in categories:
        sub = round(by_cat.get(c["id"], 0), 2)
        project_total += sub
        await db.boq_categories.update_one({"id": c["id"]}, {"$set": {"subtotal": sub}})

    misc_pct = float(boq.get("misc_pct") or 0)
    misc_amount = round(project_total * misc_pct / 100.0, 2)

    additional = boq.get("additional_charges") or []
    add_total = 0.0
    for a in additional:
        if a.get("type") == "percent":
            add_total += project_total * float(a.get("value") or 0) / 100.0
        else:
            add_total += float(a.get("value") or 0)

    design_fee = boq.get("design_fee") or {}
    def _fee_amt(cfg):
        if not cfg or cfg.get("mode") == "N/A": return 0.0
        v = float(cfg.get("value") or 0)
        if cfg.get("mode") == "percent":
            return project_total * v / 100.0
        return v
    design_amt = _fee_amt(boq.get("design_fee"))
    exec_amt = _fee_amt(boq.get("execution_fee"))
    supervisor_amt = _fee_amt(boq.get("supervisor_cost"))

    final_total = round(project_total + misc_amount + add_total + design_amt + exec_amt + supervisor_amt, 2)

    await db.boqs.update_one({"id": boq_id}, {"$set": {
        "project_total": round(project_total, 2),
        "misc_amount": misc_amount,
        "additional_total": round(add_total, 2),
        "design_amount": round(design_amt, 2),
        "execution_amount": round(exec_amt, 2),
        "supervisor_amount": round(supervisor_amt, 2),
        "final_total": final_total,
        "total_amount": final_total,
        "updated_at": now_iso(),
    }})
    return {"project_total": round(project_total, 2), "final_total": final_total}


async def _full_boq(boq_id: str) -> dict:
    boq = await db.boqs.find_one({"id": boq_id}, {"_id": 0})
    if not boq:
        raise HTTPException(404, "BOQ not found")
    cats = await db.boq_categories.find({"boq_id": boq_id}, {"_id": 0}).sort("order", 1).to_list(500)
    items = await db.boq_items.find({"boq_id": boq_id}, {"_id": 0}).sort("order", 1).to_list(5000)
    proj = await db.projects.find_one({"id": boq.get("project_id")}, {"_id": 0}) if boq.get("project_id") else None
    boq["categories"] = cats
    boq["items"] = items
    boq["project"] = proj
    return boq


def _ensure_editable(boq: dict):
    if boq.get("locked") or boq.get("status") in ("approved", "final", "archived"):
        raise HTTPException(423, "This BOQ is locked. Duplicate to create a new editable version.")


DEFAULT_TERMS_HTML = (
    "<ol>"
    "<li>All rates are inclusive of material, labour, transportation and applicable taxes unless stated otherwise.</li>"
    "<li>Any additional work outside the scope of this BOQ shall be charged separately as per mutually agreed rates.</li>"
    "<li>Payment schedule: 30% advance, 40% at execution stage, 20% at completion, 10% at handover.</li>"
    "<li>Validity of this BOQ is 30 days from the date of issue.</li>"
    "</ol>"
)


# ---------- BOQ endpoints ----------
@api.get("/boqs")
async def boqs_list(
    project_id: Optional[str] = None,
    status: Optional[str] = None,
    q: Optional[str] = None,
    current=Depends(get_current_user),
):
    query: dict = {}
    if project_id: query["project_id"] = project_id
    if status: query["status"] = status
    if q: query["project_name"] = {"$regex": q, "$options": "i"}
    rows = await db.boqs.find(query, {"_id": 0}).sort("updated_at", -1).to_list(500)
    for r in rows:
        r["category_count"] = await db.boq_categories.count_documents({"boq_id": r["id"]})
        r["item_count"] = await db.boq_items.count_documents({"boq_id": r["id"]})
    return rows


@api.get("/boqs/summary")
async def boqs_summary(current=Depends(get_current_user)):
    total = await db.boqs.count_documents({})
    drafts = await db.boqs.count_documents({"status": "draft"})
    in_prog = await db.boqs.count_documents({"status": "in_progress"})
    awaiting = await db.boqs.count_documents({"status": "awaiting_approval"})
    approved = await db.boqs.count_documents({"status": {"$in": ["approved", "final"]}})
    templates = await db.boq_templates.count_documents({})
    return {"total": total, "drafts": drafts + in_prog, "awaiting": awaiting, "approved": approved, "templates": templates}


@api.get("/boqs/{boq_id}")
async def boq_get(boq_id: str, current=Depends(get_current_user)):
    return await _full_boq(boq_id)


class BoqCreateIn(BaseModel):
    project_id: str
    title: Optional[str] = None
    template_id: Optional[str] = None
    version: Optional[str] = "V1"
    prepared_by: Optional[str] = None


@api.post("/boqs")
async def boq_create(payload: BoqCreateIn, current=Depends(get_current_user)):
    await enforce_free_trial_cap(current, "boqs")
    project = await db.projects.find_one({"id": payload.project_id}, {"_id": 0})
    if not project:
        raise HTTPException(404, "Project not found")
    boq_id = gen_id()
    title = payload.title or f"{project['name']} — BOQ"
    now = now_iso()
    boq = {
        "id": boq_id,
        "project_id": project["id"],
        "project_name": project["name"],
        "client_name": project.get("client_name"),
        "location": project.get("location"),
        "title": title,
        "version": payload.version or "V1",
        "status": "draft",
        "prepared_by": payload.prepared_by or (current.get("name") if current else "INOS Studio"),
        "date": now[:10],
        "misc_pct": 10.0,
        "design_fee": {"mode": "N/A", "value": 0},
        "execution_fee": {"mode": "N/A", "value": 0},
        "supervisor_cost": {"mode": "N/A", "value": 0},
        "additional_charges": [],
        "terms_html": DEFAULT_TERMS_HTML,
        "created_by": current.get("name") if current else "System",
        "created_by_id": current.get("id") if current else None,
        "created_at": now, "updated_at": now,
        "locked": False, "parent_version_id": None,
        "project_total": 0, "misc_amount": 0, "final_total": 0, "total_amount": 0,
    }
    await db.boqs.insert_one(boq)

    # If template provided, copy categories/items
    if payload.template_id:
        tmpl = await db.boq_templates.find_one({"id": payload.template_id}, {"_id": 0})
        if tmpl:
            for i, tc in enumerate(tmpl.get("categories") or []):
                new_cat_id = gen_id()
                await db.boq_categories.insert_one({
                    "id": new_cat_id, "boq_id": boq_id,
                    "code": tc.get("code"), "name": tc.get("name"),
                    "order": i, "collapsed": False, "subtotal": 0,
                })
                for j, ti in enumerate(tc.get("items") or []):
                    it = {**ti, "id": gen_id(), "boq_id": boq_id, "category_id": new_cat_id, "order": j}
                    it["amount"] = compute_item_amount(it)
                    await db.boq_items.insert_one(it)
            await _recompute_totals(boq_id)

    await db.activity.insert_one({
        "id": gen_id(), "actor": current.get("name"), "actor_initials": current.get("avatar_initials"),
        "action": "created", "target": f"BOQ — {project['name']}",
        "project_id": project["id"], "project_name": project["name"],
        "at": now, "status": "info",
    })
    return await _full_boq(boq_id)


class BoqPatchIn(BaseModel):
    model_config = ConfigDict(extra="allow")
    title: Optional[str] = None
    status: Optional[str] = None
    misc_pct: Optional[float] = None
    design_fee: Optional[dict] = None
    execution_fee: Optional[dict] = None
    supervisor_cost: Optional[dict] = None
    additional_charges: Optional[list] = None
    terms_html: Optional[str] = None
    prepared_by: Optional[str] = None
    date: Optional[str] = None


@api.patch("/boqs/{boq_id}")
async def boq_patch(boq_id: str, payload: BoqPatchIn, current=Depends(get_current_user)):
    boq = await db.boqs.find_one({"id": boq_id}, {"_id": 0})
    if not boq: raise HTTPException(404, "BOQ not found")
    _ensure_editable(boq)
    upd = {k: v for k, v in payload.model_dump(exclude_none=True).items()}
    if upd:
        upd["updated_at"] = now_iso()
        await db.boqs.update_one({"id": boq_id}, {"$set": upd})
    await _recompute_totals(boq_id)
    return await _full_boq(boq_id)


@api.delete("/boqs/{boq_id}")
async def boq_delete(boq_id: str, current=Depends(get_current_user)):
    boq = await db.boqs.find_one({"id": boq_id}, {"_id": 0})
    if not boq: raise HTTPException(404, "BOQ not found")
    if boq.get("status") not in ("draft", "in_progress"):
        raise HTTPException(423, "This BOQ is locked and cannot be deleted. Duplicate to create a new editable version.")
    await db.boqs.delete_one({"id": boq_id})
    await db.boq_categories.delete_many({"boq_id": boq_id})
    await db.boq_items.delete_many({"boq_id": boq_id})
    return {"ok": True}


# ---------- Categories ----------
class CategoryIn(BaseModel):
    code: Optional[str] = None
    name: Optional[str] = None
    catalog_code: Optional[str] = None  # add from catalog
    include_items: Optional[bool] = True
    item_codes: Optional[List[str]] = None  # subset of catalog items


@api.post("/boqs/{boq_id}/categories")
async def category_add(boq_id: str, payload: CategoryIn, current=Depends(get_current_user)):
    boq = await db.boqs.find_one({"id": boq_id}, {"_id": 0})
    if not boq: raise HTTPException(404, "BOQ not found")
    _ensure_editable(boq)
    order = await db.boq_categories.count_documents({"boq_id": boq_id})
    cid = gen_id()

    if payload.catalog_code:
        catalog = await db.boq_catalog_categories.find_one({"code": payload.catalog_code}, {"_id": 0})
        if not catalog:
            raise HTTPException(404, "Catalog category not found")
        await db.boq_categories.insert_one({
            "id": cid, "boq_id": boq_id, "code": catalog["code"], "name": catalog["name"],
            "order": order, "collapsed": False, "subtotal": 0,
        })
        if payload.include_items:
            items = catalog.get("items") or []
            if payload.item_codes:
                items = [i for i in items if i.get("code") in payload.item_codes]
            for j, ti in enumerate(items):
                default_rate = ti.get("rate", ti.get("default_rate", 0))
                ct = ti.get("calc_type", "M")
                it = {
                    "id": gen_id(), "boq_id": boq_id, "category_id": cid, "order": j,
                    "description": ti.get("description"), "location": ti.get("location", ""),
                    "unit": ti.get("unit"), "quantity": ti.get("quantity", 0),
                    "rate": 0 if ct == "L" else default_rate,
                    "calc_type": ct, "notes": "",
                    "detail": ti.get("detail") or {}, "hide_from_client": False,
                    "amount": default_rate if ct == "L" else 0,
                }
                it["amount"] = compute_item_amount(it)
                await db.boq_items.insert_one(it)
    else:
        await db.boq_categories.insert_one({
            "id": cid, "boq_id": boq_id, "code": payload.code or f"C{order+1}", "name": payload.name or "New Category",
            "order": order, "collapsed": False, "subtotal": 0,
        })
    await _recompute_totals(boq_id)
    return await _full_boq(boq_id)


class CategoryPatchIn(BaseModel):
    code: Optional[str] = None
    name: Optional[str] = None
    collapsed: Optional[bool] = None


@api.patch("/boqs/{boq_id}/categories/{cid}")
async def category_patch(boq_id: str, cid: str, payload: CategoryPatchIn, current=Depends(get_current_user)):
    boq = await db.boqs.find_one({"id": boq_id}, {"_id": 0})
    if not boq: raise HTTPException(404, "BOQ not found")
    _ensure_editable(boq)
    upd = {k: v for k, v in payload.model_dump(exclude_none=True).items()}
    if upd:
        await db.boq_categories.update_one({"id": cid, "boq_id": boq_id}, {"$set": upd})
    return await _full_boq(boq_id)


@api.delete("/boqs/{boq_id}/categories/{cid}")
async def category_delete(boq_id: str, cid: str, current=Depends(get_current_user)):
    boq = await db.boqs.find_one({"id": boq_id}, {"_id": 0})
    if not boq: raise HTTPException(404, "BOQ not found")
    _ensure_editable(boq)
    await db.boq_categories.delete_one({"id": cid, "boq_id": boq_id})
    await db.boq_items.delete_many({"boq_id": boq_id, "category_id": cid})
    await _recompute_totals(boq_id)
    return await _full_boq(boq_id)


class ReorderIn(BaseModel):
    ordered_ids: List[str]


@api.post("/boqs/{boq_id}/categories/reorder")
async def categories_reorder(boq_id: str, payload: ReorderIn, current=Depends(get_current_user)):
    boq = await db.boqs.find_one({"id": boq_id}, {"_id": 0})
    if not boq: raise HTTPException(404, "BOQ not found")
    _ensure_editable(boq)
    for i, cid in enumerate(payload.ordered_ids):
        await db.boq_categories.update_one({"id": cid, "boq_id": boq_id}, {"$set": {"order": i}})
    return await _full_boq(boq_id)


# ---------- Items ----------
class ItemIn(BaseModel):
    description: Optional[str] = ""
    location: Optional[str] = ""
    unit: Optional[str] = "Nos."
    quantity: Optional[float] = 0
    rate: Optional[float] = 0
    calc_type: Optional[str] = "M"
    notes: Optional[str] = ""
    amount: Optional[float] = None
    detail: Optional[dict] = None
    hide_from_client: Optional[bool] = False


@api.post("/boqs/{boq_id}/categories/{cid}/items")
async def item_add(boq_id: str, cid: str, payload: ItemIn, current=Depends(get_current_user)):
    boq = await db.boqs.find_one({"id": boq_id}, {"_id": 0})
    if not boq: raise HTTPException(404, "BOQ not found")
    _ensure_editable(boq)
    order = await db.boq_items.count_documents({"boq_id": boq_id, "category_id": cid})
    it = payload.model_dump(exclude_none=True)
    it.update({"id": gen_id(), "boq_id": boq_id, "category_id": cid, "order": order})
    it["amount"] = compute_item_amount(it)
    await db.boq_items.insert_one(it)
    await _recompute_totals(boq_id)
    return await _full_boq(boq_id)


class ItemPatchIn(BaseModel):
    model_config = ConfigDict(extra="allow")


@api.patch("/boqs/{boq_id}/items/{iid}")
async def item_patch(boq_id: str, iid: str, payload: ItemPatchIn, current=Depends(get_current_user)):
    boq = await db.boqs.find_one({"id": boq_id}, {"_id": 0})
    if not boq: raise HTTPException(404, "BOQ not found")
    _ensure_editable(boq)
    doc = payload.model_dump(exclude_unset=True)
    # If detail with formula provided, compute qty
    if "detail" in doc and doc["detail"]:
        q = compute_formula_qty(doc["detail"])
        if q is not None:
            doc["quantity"] = q
    existing = await db.boq_items.find_one({"id": iid, "boq_id": boq_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Item not found")
    merged = {**existing, **doc}
    merged["amount"] = compute_item_amount(merged)
    await db.boq_items.update_one({"id": iid, "boq_id": boq_id}, {"$set": {**doc, "amount": merged["amount"]}})
    await _recompute_totals(boq_id)
    return await _full_boq(boq_id)


@api.delete("/boqs/{boq_id}/items/{iid}")
async def item_delete(boq_id: str, iid: str, current=Depends(get_current_user)):
    boq = await db.boqs.find_one({"id": boq_id}, {"_id": 0})
    if not boq: raise HTTPException(404, "BOQ not found")
    _ensure_editable(boq)
    await db.boq_items.delete_one({"id": iid, "boq_id": boq_id})
    await _recompute_totals(boq_id)
    return await _full_boq(boq_id)


class ItemsReorderIn(BaseModel):
    category_id: str
    ordered_ids: List[str]


@api.post("/boqs/{boq_id}/items/reorder")
async def items_reorder(boq_id: str, payload: ItemsReorderIn, current=Depends(get_current_user)):
    boq = await db.boqs.find_one({"id": boq_id}, {"_id": 0})
    if not boq: raise HTTPException(404, "BOQ not found")
    _ensure_editable(boq)
    for i, iid in enumerate(payload.ordered_ids):
        await db.boq_items.update_one({"id": iid, "boq_id": boq_id}, {"$set": {"order": i, "category_id": payload.category_id}})
    return await _full_boq(boq_id)


class BulkItemsIn(BaseModel):
    ids: List[str]
    op: str  # delete | change_unit | change_rate | move | duplicate | set_lump | set_measured
    value: Optional[Any] = None


@api.post("/boqs/{boq_id}/items/bulk")
async def items_bulk(boq_id: str, payload: BulkItemsIn, current=Depends(get_current_user)):
    boq = await db.boqs.find_one({"id": boq_id}, {"_id": 0})
    if not boq: raise HTTPException(404, "BOQ not found")
    _ensure_editable(boq)
    ids = payload.ids
    if payload.op == "delete":
        await db.boq_items.delete_many({"boq_id": boq_id, "id": {"$in": ids}})
    elif payload.op == "change_unit":
        await db.boq_items.update_many({"boq_id": boq_id, "id": {"$in": ids}}, {"$set": {"unit": payload.value}})
    elif payload.op == "change_rate":
        rate = float(payload.value or 0)
        items = await db.boq_items.find({"boq_id": boq_id, "id": {"$in": ids}}, {"_id": 0}).to_list(500)
        for it in items:
            it["rate"] = rate
            it["amount"] = compute_item_amount(it)
            await db.boq_items.update_one({"id": it["id"]}, {"$set": {"rate": rate, "amount": it["amount"]}})
    elif payload.op in ("set_lump", "set_measured"):
        ct = "L" if payload.op == "set_lump" else "M"
        items = await db.boq_items.find({"boq_id": boq_id, "id": {"$in": ids}}, {"_id": 0}).to_list(500)
        for it in items:
            it["calc_type"] = ct
            it["amount"] = compute_item_amount(it)
            await db.boq_items.update_one({"id": it["id"]}, {"$set": {"calc_type": ct, "amount": it["amount"]}})
    elif payload.op == "move":
        await db.boq_items.update_many({"boq_id": boq_id, "id": {"$in": ids}}, {"$set": {"category_id": payload.value}})
    elif payload.op == "duplicate":
        items = await db.boq_items.find({"boq_id": boq_id, "id": {"$in": ids}}, {"_id": 0}).to_list(500)
        for it in items:
            new = {**it, "id": gen_id()}
            new["description"] = (new.get("description") or "") + " (copy)"
            await db.boq_items.insert_one(new)
    else:
        raise HTTPException(400, "Unknown op")
    await _recompute_totals(boq_id)
    return await _full_boq(boq_id)


# ---------- Catalog ----------
@api.get("/boq-catalog")
async def boq_catalog(current=Depends(get_current_user)):
    return await db.boq_catalog_categories.find({}, {"_id": 0}).sort("code", 1).to_list(200)


@api.get("/boq-catalog/{code}")
async def boq_catalog_detail(code: str, current=Depends(get_current_user)):
    c = await db.boq_catalog_categories.find_one({"code": code}, {"_id": 0})
    if not c: raise HTTPException(404, "Catalog category not found")
    return c


# ---------- Versioning ----------
class DuplicateIn(BaseModel):
    new_version: Optional[str] = None
    reason: Optional[str] = "Revision"
    note: Optional[str] = ""
    copy_categories: bool = True
    copy_terms: bool = True
    copy_fees: bool = True


@api.post("/boqs/{boq_id}/duplicate-version")
async def boq_duplicate_version(boq_id: str, payload: DuplicateIn, current=Depends(get_current_user)):
    src = await db.boqs.find_one({"id": boq_id}, {"_id": 0})
    if not src: raise HTTPException(404, "BOQ not found")
    # Determine next version label
    def _next_v(v):
        try:
            n = int((v or "V1").lstrip("Vv"))
            return f"V{n+1}"
        except Exception:
            return "V2"
    new_id = gen_id()
    new_boq = {**src}
    new_boq.pop("_id", None)
    new_boq["id"] = new_id
    new_boq["version"] = payload.new_version or _next_v(src.get("version"))
    new_boq["status"] = "draft"
    new_boq["locked"] = False
    new_boq["parent_version_id"] = boq_id
    new_boq["revision_note"] = payload.note
    new_boq["revision_reason"] = payload.reason
    new_boq["created_at"] = now_iso()
    new_boq["updated_at"] = now_iso()
    if not payload.copy_terms: new_boq["terms_html"] = DEFAULT_TERMS_HTML
    if not payload.copy_fees:
        new_boq["design_fee"] = {"mode": "N/A", "value": 0}
        new_boq["execution_fee"] = {"mode": "N/A", "value": 0}
        new_boq["supervisor_cost"] = {"mode": "N/A", "value": 0}
        new_boq["additional_charges"] = []
    await db.boqs.insert_one(new_boq)

    if payload.copy_categories:
        cats = await db.boq_categories.find({"boq_id": boq_id}, {"_id": 0}).to_list(500)
        cat_map = {}
        for c in cats:
            nc = {**c, "id": gen_id(), "boq_id": new_id}
            cat_map[c["id"]] = nc["id"]
            await db.boq_categories.insert_one(nc)
        items = await db.boq_items.find({"boq_id": boq_id}, {"_id": 0}).to_list(5000)
        for it in items:
            ni = {**it, "id": gen_id(), "boq_id": new_id, "category_id": cat_map.get(it["category_id"], it["category_id"])}
            await db.boq_items.insert_one(ni)
    await _recompute_totals(new_id)
    return await _full_boq(new_id)


@api.get("/boqs/{boq_id}/versions")
async def boq_versions(boq_id: str, current=Depends(get_current_user)):
    boq = await db.boqs.find_one({"id": boq_id}, {"_id": 0})
    if not boq: raise HTTPException(404, "BOQ not found")
    # walk lineage: find root, then all children of root
    root_id = boq_id
    while True:
        parent = boq.get("parent_version_id")
        if not parent: break
        root_id = parent
        boq = await db.boqs.find_one({"id": parent}, {"_id": 0})
        if not boq: break
    versions = await db.boqs.find({"$or": [{"id": root_id}, {"parent_version_id": {"$ne": None}}]}, {"_id": 0}).to_list(500)
    # Collect only those in the same lineage
    lineage = []
    id_set = {root_id}
    changed = True
    all_boqs = await db.boqs.find({}, {"_id": 0}).to_list(1000)
    while changed:
        changed = False
        for b in all_boqs:
            if b.get("parent_version_id") in id_set and b["id"] not in id_set:
                id_set.add(b["id"])
                changed = True
    for b in all_boqs:
        if b["id"] in id_set:
            lineage.append(b)
    lineage.sort(key=lambda x: x.get("created_at", ""))
    return lineage


@api.get("/boqs/{boq_id}/compare")
async def boq_compare(boq_id: str, vs: str, current=Depends(get_current_user)):
    a = await _full_boq(boq_id)
    b = await _full_boq(vs)
    a_items = {(i["category_id"], i["description"]): i for i in a["items"]}
    b_items = {(i["category_id"], i["description"]): i for i in b["items"]}
    added, removed, updated = [], [], []
    for k, bi in b_items.items():
        if k not in a_items: added.append(bi)
        else:
            ai = a_items[k]
            if float(ai.get("amount") or 0) != float(bi.get("amount") or 0) or float(ai.get("rate") or 0) != float(bi.get("rate") or 0) or float(ai.get("quantity") or 0) != float(bi.get("quantity") or 0):
                updated.append({"before": ai, "after": bi})
    for k, ai in a_items.items():
        if k not in b_items: removed.append(ai)
    delta = float(b.get("final_total") or 0) - float(a.get("final_total") or 0)
    return {
        "a": {"id": a["id"], "version": a.get("version"), "final_total": a.get("final_total")},
        "b": {"id": b["id"], "version": b.get("version"), "final_total": b.get("final_total")},
        "added": added, "removed": removed, "updated": updated,
        "delta": round(delta, 2),
    }


# ---------- Approval ----------
class SubmitIn(BaseModel):
    reviewer_id: Optional[str] = None
    reviewer_name: Optional[str] = None
    deadline: Optional[str] = None
    note: Optional[str] = ""


@api.post("/boqs/{boq_id}/submit-for-approval")
async def boq_submit(boq_id: str, payload: SubmitIn, current=Depends(get_current_user)):
    boq = await db.boqs.find_one({"id": boq_id}, {"_id": 0})
    if not boq: raise HTTPException(404, "BOQ not found")
    _ensure_editable(boq)
    await db.boqs.update_one({"id": boq_id}, {"$set": {
        "status": "awaiting_approval",
        "review": {"reviewer_id": payload.reviewer_id, "reviewer_name": payload.reviewer_name, "deadline": payload.deadline, "note": payload.note, "submitted_at": now_iso(), "submitted_by": current.get("name")},
        "updated_at": now_iso(),
    }})
    await db.activity.insert_one({
        "id": gen_id(), "actor": current.get("name"), "actor_initials": current.get("avatar_initials"),
        "action": "submitted for approval", "target": f"BOQ {boq.get('version')} — {boq['project_name']}",
        "project_id": boq.get("project_id"), "project_name": boq.get("project_name"),
        "at": now_iso(), "status": "info",
    })
    return await _full_boq(boq_id)


class ApprovalIn(BaseModel):
    remarks: Optional[str] = ""


@api.post("/boqs/{boq_id}/approve")
async def boq_approve(boq_id: str, payload: ApprovalIn, current=Depends(get_current_user)):
    boq = await db.boqs.find_one({"id": boq_id}, {"_id": 0})
    if not boq: raise HTTPException(404, "BOQ not found")
    now = now_iso()
    await db.boqs.update_one({"id": boq_id}, {"$set": {
        "status": "approved", "locked": True,
        "approved_by": current.get("name"), "approved_at": now,
        "approval_remarks": payload.remarks, "updated_at": now,
    }})
    # Auto-attach BOQ into Phase B documents workspace (new schema) — with REAL PDF bytes
    existing_doc = await db.documents.find_one({"source_app": "boq", "source_id": boq_id})
    if not existing_doc:
        pdf_bytes = b""
        try:
            # Generate the internal-variant PDF right here so the workspace
            # has the actual approved file, not a 0-byte placeholder.
            from phase7_pdf_v2 import (
                _register_fonts, _masthead, _category_table, _totals_block,
                _terms_block, _signatures, _footer_maker, _FONT_REGISTERED as _fr,  # noqa: F401
            )
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import mm
            from reportlab.platypus import SimpleDocTemplate, Spacer
            import io as _io
            _register_fonts()
            full = await _full_boq(boq_id)
            proj = await db.projects.find_one({"id": full.get("project_id")}) or {}
            proj.pop("_id", None)
            buf = _io.BytesIO()
            _doc = SimpleDocTemplate(buf, pagesize=A4,
                leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=20*mm,
                title=full.get("boq_number") or f"BOQ-V{full.get('version',1)}")
            story = []
            story.extend(_masthead(full, proj, "internal", True))
            cols = ["sno", "desc", "loc", "unit", "qty", "rate", "type", "amt"]
            for c in (full.get("categories") or []):
                if not c.get("items"):
                    continue
                story.append(_category_table(c, cols, True, True))
                story.append(Spacer(1, 3*mm))
            story.extend(_totals_block(full, True))
            story.extend(_terms_block(full))
            story.extend(_signatures(proj))
            draw = _footer_maker(full.get("boq_number") or "BOQ", full.get("version", 1), "Confidential — For internal use")
            _doc.build(story, onFirstPage=draw, onLaterPages=draw)
            pdf_bytes = buf.getvalue()
        except Exception as _pdf_err:
            logger.warning(f"BOQ approve: could not build PDF ({_pdf_err})")
        await db.documents.insert_one({
            "id": gen_id(),
            "project_id": boq.get("project_id"),
            "category": "BOQs",
            "title": f"{boq.get('version') or 'BOQ'} — {boq['project_name']}",
            "filename": f"{boq.get('boq_number') or boq.get('version') or 'BOQ'}.pdf",
            "mime": "application/pdf",
            "size": len(pdf_bytes),
            "version": boq.get("version") or "V1",
            "revision": 1,
            "document_date": datetime.now(timezone.utc).date().isoformat(),
            "uploaded_by": current.get("email"),
            "uploaded_by_name": current.get("name"),
            "visibility": "internal",
            "status": "approved",
            "source_app": "boq",
            "source_id": boq_id,
            "connected_milestone": "BOQ Approval",
            "is_archived": False,
            "content_b64": base64.b64encode(pdf_bytes).decode() if pdf_bytes else "",
            "created_at": now, "updated_at": now,
        })
    await db.activity.insert_one({
        "id": gen_id(), "actor": current.get("name"), "actor_initials": current.get("avatar_initials"),
        "action": "approved", "target": f"BOQ {boq.get('version')} — {boq['project_name']}",
        "project_id": boq.get("project_id"), "project_name": boq.get("project_name"),
        "at": now, "status": "success",
    })
    # Mark BOQ milestone done if any
    await db.milestones.update_many(
        {"project_id": boq.get("project_id"), "title": {"$regex": "BOQ", "$options": "i"}},
        {"$set": {"status": "done"}},
    )
    return await _full_boq(boq_id)


@api.post("/boqs/{boq_id}/return")
async def boq_return(boq_id: str, payload: ApprovalIn, current=Depends(get_current_user)):
    boq = await db.boqs.find_one({"id": boq_id}, {"_id": 0})
    if not boq: raise HTTPException(404, "BOQ not found")
    await db.boqs.update_one({"id": boq_id}, {"$set": {"status": "returned", "review_remarks": payload.remarks, "updated_at": now_iso()}})
    return await _full_boq(boq_id)


@api.post("/boqs/{boq_id}/reject")
async def boq_reject(boq_id: str, payload: ApprovalIn, current=Depends(get_current_user)):
    boq = await db.boqs.find_one({"id": boq_id}, {"_id": 0})
    if not boq: raise HTTPException(404, "BOQ not found")
    await db.boqs.update_one({"id": boq_id}, {"$set": {"status": "archived", "review_remarks": payload.remarks, "updated_at": now_iso()}})
    return await _full_boq(boq_id)


# ---------- Templates ----------
class TemplateIn(BaseModel):
    name: str
    project_type: Optional[str] = ""
    description: Optional[str] = ""
    include_rates: bool = True
    include_terms: bool = True
    include_fees: bool = True
    scope: str = "personal"


@api.post("/boqs/{boq_id}/save-as-template")
async def boq_save_template(boq_id: str, payload: TemplateIn, current=Depends(get_current_user)):
    boq = await _full_boq(boq_id)
    cats = []
    items_by_cat: Dict[str, list] = {}
    for it in boq["items"]:
        items_by_cat.setdefault(it["category_id"], []).append({
            "description": it.get("description"), "unit": it.get("unit"),
            "quantity": it.get("quantity"), "rate": it.get("rate") if payload.include_rates else 0,
            "calc_type": it.get("calc_type", "M"),
        })
    for c in boq["categories"]:
        cats.append({"code": c.get("code"), "name": c.get("name"), "items": items_by_cat.get(c["id"], [])})
    tmpl = {
        "id": gen_id(), "name": payload.name, "project_type": payload.project_type,
        "description": payload.description, "scope": payload.scope,
        "categories": cats,
        "terms_html": boq.get("terms_html") if payload.include_terms else DEFAULT_TERMS_HTML,
        "misc_pct": boq.get("misc_pct") if payload.include_fees else 10.0,
        "created_by": current.get("name"), "created_at": now_iso(),
    }
    await db.boq_templates.insert_one(tmpl)
    return {"id": tmpl["id"], "name": tmpl["name"]}


@api.get("/boq-templates")
async def templates_list(current=Depends(get_current_user)):
    return await db.boq_templates.find({}, {"_id": 0}).to_list(200)


# ---------- Excel & PDF Export ----------
def _fmt_inr(n) -> str:
    try:
        n = int(round(float(n)))
    except Exception:
        return "0"
    s = str(abs(n))
    if len(s) <= 3:
        out = s
    else:
        last3 = s[-3:]
        rest = s[:-3]
        with_commas = ""
        for i, ch in enumerate(reversed(rest)):
            if i > 0 and i % 2 == 0:
                with_commas = "," + with_commas
            with_commas = ch + with_commas
        out = with_commas + "," + last3
    return ("-" if n < 0 else "") + out


@api.get("/boqs/{boq_id}/export/excel")
async def boq_export_excel(boq_id: str, variant: str = "full", current=Depends(get_current_user)):
    boq = await _full_boq(boq_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"

    thin = Side(border_style="thin", color="DDD8CE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F453B")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    cat_fill = PatternFill("solid", fgColor="F3F3F1")
    cat_font = Font(bold=True, color="1F453B", size=11)

    ws["A1"] = "INOS — BILL OF QUANTITIES"
    ws["A1"].font = Font(bold=True, size=16, color="EF7F1B")
    ws.merge_cells("A1:G1")

    ws["A3"] = "Project:"; ws["B3"] = boq.get("project_name")
    ws["A4"] = "Client:"; ws["B4"] = boq.get("client_name") or ""
    ws["A5"] = "Location:"; ws["B5"] = boq.get("location") or ""
    ws["A6"] = "Version:"; ws["B6"] = boq.get("version")
    ws["A7"] = "Status:"; ws["B7"] = boq.get("status")
    ws["A8"] = "Prepared By:"; ws["B8"] = boq.get("prepared_by") or ""
    ws["A9"] = "Date:"; ws["B9"] = boq.get("date") or ""

    row = 11
    headers = ["S. No.", "Description", "Unit", "Qty", "Rate (₹)", "Amount (₹)"]
    if variant == "quantity":
        headers = ["S. No.", "Description", "Unit", "Qty"]
    elif variant == "summary":
        headers = ["Category", "Items", "Subtotal (₹)"]

    if variant == "summary":
        for j, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=j, value=h)
            cell.fill = header_fill; cell.font = header_font; cell.border = border
        row += 1
        for c in boq["categories"]:
            cnt = len([i for i in boq["items"] if i["category_id"] == c["id"]])
            ws.cell(row=row, column=1, value=f"{c.get('code','')} — {c.get('name')}").border = border
            ws.cell(row=row, column=2, value=cnt).border = border
            ws.cell(row=row, column=3, value=_fmt_inr(c.get("subtotal", 0))).border = border
            row += 1
    else:
        for j, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=j, value=h)
            cell.fill = header_fill; cell.font = header_font; cell.border = border
        row += 1
        cat_by_id = {c["id"]: c for c in boq["categories"]}
        for c in boq["categories"]:
            ws.cell(row=row, column=1, value=f"{c.get('code','')} — {c.get('name')}").font = cat_font
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = cat_fill
                ws.cell(row=row, column=col).border = border
            row += 1
            sno = 1
            for it in [i for i in boq["items"] if i["category_id"] == c["id"]]:
                ws.cell(row=row, column=1, value=sno).border = border
                ws.cell(row=row, column=2, value=it.get("description")).border = border
                ws.cell(row=row, column=3, value=it.get("unit")).border = border
                if it.get("calc_type") == "L":
                    ws.cell(row=row, column=4, value="—").border = border
                    if len(headers) >= 5: ws.cell(row=row, column=5, value="Lump sum").border = border
                else:
                    ws.cell(row=row, column=4, value=float(it.get("quantity") or 0)).border = border
                    if len(headers) >= 5: ws.cell(row=row, column=5, value=float(it.get("rate") or 0)).border = border
                if len(headers) >= 6: ws.cell(row=row, column=6, value=float(it.get("amount") or 0)).border = border
                row += 1
                sno += 1
            if len(headers) >= 6:
                sc = ws.cell(row=row, column=5, value="SUBTOTAL")
                sc.font = Font(bold=True); sc.border = border
                sv = ws.cell(row=row, column=6, value=float(c.get("subtotal", 0)))
                sv.font = Font(bold=True); sv.border = border
                row += 2

    # Cost summary
    row += 1
    ws.cell(row=row, column=1, value="Cost Summary").font = Font(bold=True, size=12)
    row += 1
    for label, val in [
        ("Project Total", boq.get("project_total", 0)),
        (f"Miscellaneous ({boq.get('misc_pct',10)}%)", boq.get("misc_amount", 0)),
        ("Additional Charges", boq.get("additional_total", 0)),
        ("Design Fees", boq.get("design_amount", 0)),
        ("Execution Fees", boq.get("execution_amount", 0)),
        ("Supervisor", boq.get("supervisor_amount", 0)),
        ("FINAL TOTAL", boq.get("final_total", 0)),
    ]:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=float(val or 0))
        if label == "FINAL TOTAL":
            ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="EF7F1B")
            ws.cell(row=row, column=2).font = Font(bold=True, size=12, color="EF7F1B")
        row += 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    boq_number = boq.get('boq_number') or f"BOQ-{boq.get('version','V1')}"
    filename = f"{boq_number}.xlsx"
    # Update workbook title header (first cell) if present
    try:
        ws.title = boq_number[-31:]  # sheet name limit
    except Exception:
        pass
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


class PdfExportIn(BaseModel):
    variant: str = "full"
    show_rates: bool = True
    show_subtotals: bool = True
    include_terms: bool = True
    include_signatures: bool = True
    include_logo: bool = True
    include_version: bool = True
    include_approval: bool = True
    client_copy: bool = False


@api.post("/boqs/{boq_id}/export/pdf-legacy")
async def boq_export_pdf(boq_id: str, payload: PdfExportIn = Body(default=PdfExportIn()), current=Depends(get_current_user)):
    boq = await _full_boq(boq_id)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=14*mm, rightMargin=14*mm, topMargin=16*mm, bottomMargin=16*mm)
    styles = getSampleStyleSheet()
    orange = colors.HexColor("#EF7F1B")
    green = colors.HexColor("#1F453B")
    muted = colors.HexColor("#6B7280")
    border_c = colors.HexColor("#DDD8CE")
    cat_bg = colors.HexColor("#F3F3F1")

    story = []
    title_style = ParagraphStyle("t", parent=styles["Title"], textColor=orange, fontSize=18, spaceAfter=4)
    h_style = ParagraphStyle("h", parent=styles["Heading2"], textColor=green, fontSize=13, spaceAfter=4)
    meta_style = ParagraphStyle("m", parent=styles["Normal"], textColor=muted, fontSize=9)
    body_style = ParagraphStyle("b", parent=styles["Normal"], fontSize=9)

    story.append(Paragraph("INOS — BILL OF QUANTITIES", title_style))
    story.append(Paragraph(f"{boq.get('project_name','')} · {boq.get('version','V1')} · Status: {boq.get('status','').replace('_',' ').title()}", meta_style))
    story.append(Spacer(1, 8))

    meta = [
        ["Client:", boq.get("client_name") or "", "Prepared By:", boq.get("prepared_by") or ""],
        ["Location:", boq.get("location") or "", "Date:", boq.get("date") or ""],
    ]
    t = Table(meta, colWidths=[26*mm, 66*mm, 26*mm, 66*mm])
    t.setStyle(TableStyle([
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("TEXTCOLOR", (0,0), (0,-1), muted),
        ("TEXTCOLOR", (2,0), (2,-1), muted),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    story.append(t)
    story.append(Spacer(1, 10))

    for c in boq["categories"]:
        story.append(Paragraph(f"<b>{c.get('code','')} — {c.get('name','')}</b>", h_style))
        headers = ["S. No.", "Description", "Unit", "Qty"]
        if payload.show_rates: headers += ["Rate (₹)", "Amount (₹)"]
        data = [headers]
        items = [i for i in boq["items"] if i["category_id"] == c["id"]]
        for sno, it in enumerate(items, 1):
            row = [str(sno), Paragraph(it.get("description") or "", body_style), it.get("unit") or ""]
            if it.get("calc_type") == "L":
                row.append("—")
                if payload.show_rates:
                    row.append("Lump sum")
            else:
                row.append(f"{it.get('quantity') or 0:g}")
                if payload.show_rates:
                    row.append(_fmt_inr(it.get("rate") or 0))
            if payload.show_rates:
                row.append(_fmt_inr(it.get("amount") or 0))
            data.append(row)
        if payload.show_subtotals and payload.show_rates:
            data.append(["", "", "", "", "SUBTOTAL", _fmt_inr(c.get("subtotal", 0))])
        col_widths = [12*mm, 78*mm, 16*mm, 16*mm]
        if payload.show_rates: col_widths += [26*mm, 30*mm]
        tbl = Table(data, colWidths=col_widths, repeatRows=1)
        style = [
            ("BACKGROUND", (0,0), (-1,0), green),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("FONTSIZE", (0,0), (-1,-1), 8.5),
            ("GRID", (0,0), (-1,-1), 0.4, border_c),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
            ("BOTTOMPADDING", (0,0), (-1,-1), 3),
            ("TOPPADDING", (0,0), (-1,-1), 3),
        ]
        if payload.show_subtotals and payload.show_rates:
            style.append(("BACKGROUND", (0,-1), (-1,-1), cat_bg))
            style.append(("FONTNAME", (0,-1), (-1,-1), "Helvetica-Bold"))
        tbl.setStyle(TableStyle(style))
        story.append(tbl)
        story.append(Spacer(1, 8))

    # Cost summary
    story.append(Spacer(1, 6))
    story.append(Paragraph("<b>Cost Summary</b>", h_style))
    summary_rows = [
        ["Project Total", _fmt_inr(boq.get("project_total", 0))],
        [f"Miscellaneous ({boq.get('misc_pct',10)}%)", _fmt_inr(boq.get("misc_amount", 0))],
    ]
    if boq.get("additional_total"): summary_rows.append(["Additional Charges", _fmt_inr(boq.get("additional_total"))])
    if boq.get("design_amount"): summary_rows.append(["Design Fees", _fmt_inr(boq.get("design_amount"))])
    if boq.get("execution_amount"): summary_rows.append(["Execution Fees", _fmt_inr(boq.get("execution_amount"))])
    if boq.get("supervisor_amount"): summary_rows.append(["Supervisor", _fmt_inr(boq.get("supervisor_amount"))])
    summary_rows.append(["FINAL TOTAL", "₹ " + _fmt_inr(boq.get("final_total", 0))])
    st = Table(summary_rows, colWidths=[130*mm, 40*mm])
    st.setStyle(TableStyle([
        ("FONTSIZE", (0,0), (-1,-1), 10),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("LINEABOVE", (0,-1), (-1,-1), 0.6, green),
        ("FONTNAME", (0,-1), (-1,-1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0,-1), (-1,-1), orange),
        ("FONTSIZE", (0,-1), (-1,-1), 12),
        ("ALIGN", (1,0), (1,-1), "RIGHT"),
    ]))
    story.append(st)

    if payload.include_terms and boq.get("terms_html"):
        story.append(Spacer(1, 14))
        story.append(Paragraph("<b>Terms & Conditions</b>", h_style))
        # Convert basic HTML to reportlab Paragraph (already supports subset)
        story.append(Paragraph(boq.get("terms_html").replace("<ol>", "").replace("</ol>", "").replace("<li>", "• ").replace("</li>", "<br/>"), body_style))

    if payload.include_signatures:
        story.append(Spacer(1, 18))
        sig = [["For INOS Studio", "For Client"], ["_____________________", "_____________________"], ["Name / Signature / Date", "Name / Signature / Date"]]
        stbl = Table(sig, colWidths=[80*mm, 80*mm])
        stbl.setStyle(TableStyle([("FONTSIZE", (0,0), (-1,-1), 9), ("TEXTCOLOR", (0,0), (-1,0), muted), ("BOTTOMPADDING", (0,0), (-1,-1), 6)]))
        story.append(stbl)

    def _footer(c, d):
        c.setFont("Helvetica", 8)
        c.setFillColor(muted)
        c.drawString(14*mm, 10*mm, f"INOS ERP · BOQ {boq.get('version','V1')}")
        c.drawRightString(196*mm, 10*mm, f"Page {d.page}")

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    buf.seek(0)
    safe_name = "".join(ch if ord(ch) < 128 else "_" for ch in (boq.get('project_name') or 'Project')).replace(' ', '_')
    filename = f"BOQ_{boq.get('version','V1')}_{safe_name}.pdf"
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )



# ---------- Phase 4 — Quotations ----------
from phase4_quotations import register_phase4, seed_phase4

register_phase4(app, api, db, {
    "get_current_user": get_current_user,
    "require_internal": require_internal_user,
    "gen_id": gen_id,
    "now_iso": now_iso,
    "fmt_inr": _fmt_inr,
    "UPLOAD_ROOT": UPLOAD_ROOT,
    "enforce_free_trial_cap": enforce_free_trial_cap,
})

# ---------- Phase 5 — Projects + Client Portal ----------
from phase5_projects import register_phase5, seed_phase5

APP_BASE_URL = os.environ.get("APP_BASE_URL") or (os.environ.get("CORS_ORIGINS", "").split(",")[0] if os.environ.get("CORS_ORIGINS") else "")

register_phase5(app, api, db, {
    "get_current_user": get_current_user,
    "require_internal": require_internal_user,
    "gen_id": gen_id,
    "now_iso": now_iso,
    "fmt_inr": _fmt_inr,
    "UPLOAD_ROOT": UPLOAD_ROOT,
    "app_base_url": APP_BASE_URL,
    "enforce_free_trial_cap": enforce_free_trial_cap,
})

# Phase D — Dashboard widget system
from phase6_dashboards import register_dashboard_routes  # noqa: E402
register_dashboard_routes(api, db, get_current_user)

# Phase 4 — BOQ family/versioning + backfill
from phase7_boq_versions import register_boq_v2_routes, backfill as _boq_backfill  # noqa: E402
register_boq_v2_routes(api, db, get_current_user)

@app.on_event("startup")
async def _phase7_backfill():
    try:
        await _boq_backfill(db)
        logger.info("BOQ v2 backfill complete.")
    except Exception as e:
        logger.warning(f"BOQ backfill skipped: {e}")

import logging as _logging_early
logger = _logging_early.getLogger(__name__)
# Register Phase 4 PDF v2 exporter (5 variants w/ ₹ Noto Sans)
try:
    from phase7_pdf_v2 import register_pdf_v2
    register_pdf_v2(api, db, get_current_user, _full_boq)
    logger.info("Phase 4 PDF v2 exporter registered.")
except Exception as e:
    logger.warning(f"PDF v2 register skipped: {e}")

# Register Phase 8 dashboard data endpoints (project-wise widgets)
try:
    from phase8_dashboards_data import register_dashboards_data
    register_dashboards_data(api, db, get_current_user)
    logger.info("Phase 8 dashboard data endpoints registered.")
except Exception as e:
    logger.warning(f"Phase 8 dashboards data register skipped: {e}")

# Register Phase 9 idempotent backfills (vendor categories, quotation validity/variance, milestones)
try:
    from phase9_seed_backfill import register_seed_backfill
    register_seed_backfill(app, db)
    logger.info("Phase 9 seed backfill registered.")
except Exception as e:
    logger.warning(f"Phase 9 backfill register skipped: {e}")

# Register Phase 10 visual dashboard endpoints (hero + donut + bar chart data)
try:
    from phase10_dashboards_visuals import register_dashboards_visuals
    register_dashboards_visuals(api, db, get_current_user)
    logger.info("Phase 10 visual dashboard endpoints registered.")
except Exception as e:
    logger.warning(f"Phase 10 visuals register skipped: {e}")

# Register Phase 11 generic app pages (settings, activity, roles, cross-project rollups, BOQ mgmt)
try:
    from phase11_generic_pages import register_phase11
    register_phase11(api, db, get_current_user)
    logger.info("Phase 11 generic page endpoints registered.")
except Exception as e:
    logger.warning(f"Phase 11 register skipped: {e}")

# Register Phase B — Documents app (uploads, forms, drawings, auto-hooks)
try:
    from phase12_documents import register_documents
    register_documents(api, db, get_current_user, enforce_free_trial_cap)
    logger.info("Phase B Documents endpoints registered.")
except Exception as e:
    logger.warning(f"Phase B Documents register skipped: {e}")

# Register Phase C/D + Handover ZIP (quotation aliases, vendor import, PDF export, ZIP)
try:
    from phase13_c_d_zip import register_phase13
    register_phase13(api, db, get_current_user, require_internal_user, _full_boq)
    logger.info("Phase C/D/ZIP endpoints registered.")
except Exception as e:
    logger.warning(f"Phase C/D register skipped: {e}")

# Register Phase F — Calendar, Notes, Tasks + BOQ PDF thumbnail
try:
    from phase14_calendar_notes_tasks import register_phase_f
    _phase_f_seed = register_phase_f(api, db, get_current_user)
    logger.info("Phase F endpoints registered.")
except Exception as e:
    logger.warning(f"Phase F register skipped: {e}")
    _phase_f_seed = None

app.include_router(api)


@app.on_event("startup")
async def _startup_phase_f_seed():
    if _phase_f_seed:
        try:
            await _phase_f_seed()
        except Exception as e:
            logger.warning(f"Phase F seed skipped: {e}")


# ---------- Comprehensive BOQ seed ----------
async def _seed_boq_catalog_and_kohli():
    """Seed BOQ catalog categories + rebuild BOQs with full item detail."""
    # Nuke any legacy simple BOQs
    await db.boqs.delete_many({})
    await db.boq_categories.delete_many({})
    await db.boq_items.delete_many({})

    # Catalog
    catalog = [
        {"code": "A", "name": "Demolition & Civil Works", "project_types": ["Residential Interior", "Commercial Interior"], "items": [
            {"code": "A1", "description": "Dismantling of existing walls", "unit": "Sq.ft.", "default_rate": 80, "calc_type": "M"},
            {"code": "A2", "description": "Removal of existing flooring", "unit": "Sq.ft.", "default_rate": 60, "calc_type": "M"},
            {"code": "A3", "description": "Demolition of false ceiling", "unit": "Sq.ft.", "default_rate": 40, "calc_type": "M"},
            {"code": "A4", "description": "Removal of electrical fixtures", "unit": "Nos.", "default_rate": 200, "calc_type": "M"},
            {"code": "A5", "description": "Debris removal & disposal", "unit": "Lump", "default_rate": 25000, "calc_type": "L"},
            {"code": "A6", "description": "Brick masonry work (4in wall)", "unit": "Sq.ft.", "default_rate": 380, "calc_type": "M"},
            {"code": "A7", "description": "Plastering (12mm thk, 1:4 CM)", "unit": "Sq.ft.", "default_rate": 105, "calc_type": "M"},
        ]},
        {"code": "B", "name": "Wall & Flooring", "project_types": ["Residential Interior", "Commercial Interior"], "items": [
            {"code": "B1", "description": "Vitrified tile flooring (600x600 mm)", "unit": "Sq.ft.", "default_rate": 180, "calc_type": "M"},
            {"code": "B2", "description": "Marble flooring — Statuario", "unit": "Sq.ft.", "default_rate": 550, "calc_type": "M"},
            {"code": "B3", "description": "Wooden laminate flooring — 8mm", "unit": "Sq.ft.", "default_rate": 220, "calc_type": "M"},
            {"code": "B4", "description": "Skirting — 4in matching tile", "unit": "Rft.", "default_rate": 120, "calc_type": "M"},
            {"code": "B5", "description": "Wall cladding — texture paint", "unit": "Sq.ft.", "default_rate": 180, "calc_type": "M"},
            {"code": "B6", "description": "Wallpaper — feature walls (imported)", "unit": "Sq.ft.", "default_rate": 320, "calc_type": "M"},
            {"code": "B7", "description": "Wall panelling — MDF with veneer", "unit": "Sq.ft.", "default_rate": 550, "calc_type": "M"},
            {"code": "B8", "description": "Waterproofing — bathrooms (2-coat)", "unit": "Sq.ft.", "default_rate": 50, "calc_type": "M"},
        ]},
        {"code": "C", "name": "Door & Window", "project_types": ["Residential Interior"], "items": [
            {"code": "C1", "description": "Main door — teakwood flush 40mm", "unit": "Nos.", "default_rate": 41600, "calc_type": "M"},
            {"code": "C2", "description": "Internal doors — laminated flush 32mm", "unit": "Nos.", "default_rate": 18500, "calc_type": "M"},
            {"code": "C3", "description": "Sliding wardrobe doors — with fittings", "unit": "Sq.ft.", "default_rate": 550, "calc_type": "M"},
            {"code": "C4", "description": "Window frames — powder-coated aluminium", "unit": "Rft.", "default_rate": 480, "calc_type": "M"},
        ]},
        {"code": "D", "name": "Services", "project_types": ["Residential Interior", "Commercial Interior"], "items": [
            {"code": "D1", "description": "Electrical wiring — full rewiring", "unit": "Lump", "default_rate": 145000, "calc_type": "L"},
            {"code": "D2", "description": "Modular switches & sockets — premium", "unit": "Nos.", "default_rate": 850, "calc_type": "M"},
            {"code": "D3", "description": "Ceiling lights — LED panels 15W", "unit": "Nos.", "default_rate": 2200, "calc_type": "M"},
            {"code": "D4", "description": "Chandelier — living room decorative", "unit": "Nos.", "default_rate": 38000, "calc_type": "M"},
            {"code": "D5", "description": "Plumbing works — bathrooms & kitchen", "unit": "Lump", "default_rate": 92000, "calc_type": "L"},
            {"code": "D6", "description": "CCTV installation — 6 cameras + DVR", "unit": "Set", "default_rate": 42000, "calc_type": "M"},
            {"code": "D7", "description": "Home automation — lighting & curtains", "unit": "Lump", "default_rate": 23350, "calc_type": "L"},
        ]},
        {"code": "E", "name": "Furniture & Fixtures", "project_types": ["Residential Interior"], "items": [
            {"code": "E1", "description": "Wardrobes — master bedroom (veneer finish)", "unit": "Sq.ft.", "default_rate": 1850, "calc_type": "M"},
            {"code": "E2", "description": "Wardrobes — kids room (laminate finish)", "unit": "Sq.ft.", "default_rate": 1650, "calc_type": "M"},
            {"code": "E3", "description": "TV unit — living room with backlit panel", "unit": "Sq.ft.", "default_rate": 1900, "calc_type": "M"},
            {"code": "E4", "description": "Study table + storage — solid wood top", "unit": "Nos.", "default_rate": 68000, "calc_type": "M"},
            {"code": "E5", "description": "Sofa — 3+2+1 upholstered fabric", "unit": "Set", "default_rate": 148000, "calc_type": "M"},
            {"code": "E6", "description": "Dining table — 6 seater, solid wood", "unit": "Set", "default_rate": 92000, "calc_type": "M"},
            {"code": "E7", "description": "Bed — king size with storage & headboard", "unit": "Nos.", "default_rate": 78000, "calc_type": "M"},
            {"code": "E8", "description": "Bed — queen (kids) with storage", "unit": "Nos.", "default_rate": 52000, "calc_type": "M"},
            {"code": "E9", "description": "Curtains + blinds — all rooms", "unit": "Lump", "default_rate": 45500, "calc_type": "L"},
            {"code": "E10", "description": "Loose furniture accents (chairs, side tables)", "unit": "Lump", "default_rate": 35000, "calc_type": "L"},
        ]},
        {"code": "F", "name": "Kitchen Works", "project_types": ["Residential Interior"], "items": [
            {"code": "F1", "description": "Modular kitchen base cabinets (BWP ply)", "unit": "Rft.", "default_rate": 10500, "calc_type": "M"},
            {"code": "F2", "description": "Wall cabinets — acrylic finish", "unit": "Rft.", "default_rate": 6800, "calc_type": "M"},
            {"code": "F3", "description": "Countertop — Granite polished", "unit": "Sq.ft.", "default_rate": 900, "calc_type": "M"},
            {"code": "F4", "description": "Backsplash tiles — mosaic finish", "unit": "Sq.ft.", "default_rate": 450, "calc_type": "M"},
            {"code": "F5", "description": "Chimney + hob combo — branded", "unit": "Set", "default_rate": 29200, "calc_type": "M"},
        ]},
    ]
    # Additional named-only categories
    for extra in ["Waterproofing", "False Ceiling", "Painting", "Electrical", "Plumbing", "Sanitaryware", "HVAC", "Metal Work", "Glass Work", "Modular Furniture", "Loose Furniture", "Lighting", "Landscape", "External Development", "Miscellaneous Works"]:
        code = "".join([w[0] for w in extra.split()])[:3].upper()
        catalog.append({"code": code, "name": extra, "project_types": [], "items": []})

    for c in catalog:
        await db.boq_catalog_categories.insert_one({**c})

    # Kohli item quantities aligned to subtotals
    kohli_items = {
        "A": [
            ("Dismantling of existing walls", "Sq.ft.", 200, 80, "M"),
            ("Removal of existing flooring", "Sq.ft.", 850, 60, "M"),
            ("Demolition of false ceiling", "Sq.ft.", 700, 40, "M"),
            ("Removal of electrical fixtures", "Nos.", 45, 200, "M"),
            ("Debris removal & disposal", "Lump", 0, 0, "L", 24000),
            ("Brick masonry work (4in wall)", "Sq.ft.", 180, 380, "M"),
            ("Plastering (12mm thk, 1:4 CM)", "Sq.ft.", 320, 105, "M"),
        ],
        "B": [
            ("Vitrified tile flooring (600x600 mm)", "Sq.ft.", 850, 180, "M"),
            ("Marble flooring — Statuario", "Sq.ft.", 220, 550, "M"),
            ("Wooden laminate flooring — 8mm", "Sq.ft.", 380, 220, "M"),
            ("Skirting — 4in matching tile", "Rft.", 340, 120, "M"),
            ("Wall cladding — texture paint", "Sq.ft.", 620, 180, "M"),
            ("Wallpaper — feature walls (imported)", "Sq.ft.", 240, 320, "M"),
            ("Wall panelling — MDF with veneer", "Sq.ft.", 180, 550, "M"),
            ("Waterproofing — bathrooms (2-coat)", "Sq.ft.", 620, 50, "M"),
        ],
        "C": [
            ("Main door — teakwood flush 40mm", "Nos.", 1, 41600, "M"),
            ("Internal doors — laminated flush 32mm", "Nos.", 4, 18500, "M"),
            ("Sliding wardrobe doors — with fittings", "Sq.ft.", 80, 550, "M"),
            ("Window frames — powder-coated aluminium", "Rft.", 60, 480, "M"),
        ],
        "D": [
            ("Electrical wiring — full rewiring", "Lump", 0, 0, "L", 145000),
            ("Modular switches & sockets — premium", "Nos.", 65, 850, "M"),
            ("Ceiling lights — LED panels 15W", "Nos.", 32, 2200, "M"),
            ("Chandelier — living room decorative", "Nos.", 1, 38000, "M"),
            ("Plumbing works — bathrooms & kitchen", "Lump", 0, 0, "L", 92000),
            ("CCTV installation — 6 cameras + DVR", "Set", 1, 42000, "M"),
            ("Home automation — lighting & curtains", "Lump", 0, 0, "L", 23350),
        ],
        "E": [
            ("Wardrobes — master bedroom (veneer finish)", "Sq.ft.", 120, 1850, "M"),
            ("Wardrobes — kids room (laminate finish)", "Sq.ft.", 80, 1650, "M"),
            ("TV unit — living room with backlit panel", "Sq.ft.", 45, 1900, "M"),
            ("Study table + storage — solid wood top", "Nos.", 1, 68000, "M"),
            ("Sofa — 3+2+1 upholstered fabric", "Set", 1, 148000, "M"),
            ("Dining table — 6 seater, solid wood", "Set", 1, 92000, "M"),
            ("Bed — king size with storage & headboard", "Nos.", 1, 78000, "M"),
            ("Bed — queen (kids) with storage", "Nos.", 1, 52000, "M"),
            ("Curtains + blinds — all rooms", "Lump", 0, 0, "L", 45500),
            ("Loose furniture accents (chairs, side tables)", "Lump", 0, 0, "L", 35000),
        ],
        "F": [
            ("Modular kitchen base cabinets (BWP ply)", "Rft.", 12, 10500, "M"),
            ("Wall cabinets — acrylic finish", "Rft.", 10, 6800, "M"),
            ("Countertop — Granite polished", "Sq.ft.", 42, 900, "M"),
            ("Backsplash tiles — mosaic finish", "Sq.ft.", 28, 450, "M"),
            ("Chimney + hob combo — branded", "Set", 1, 29200, "M"),
        ],
    }
    cat_names = {"A": "Demolition & Civil Works", "B": "Wall & Flooring", "C": "Door & Window", "D": "Services", "E": "Furniture & Fixtures", "F": "Kitchen Works"}

    projects = await db.projects.find({}, {"_id": 0}).to_list(200)
    by_name = {p["name"]: p for p in projects}
    now = now_iso()

    async def _create_boq(project, version, status, prepared_by, misc_pct=10.0, add_items_from=None, add_misc_charges=None, locked=False, approved=False):
        boq_id = gen_id()
        boq = {
            "id": boq_id, "project_id": project["id"], "project_name": project["name"],
            "client_name": project.get("client_name"), "location": project.get("location"),
            "title": f"{project['name']} — BOQ",
            "version": version, "status": status,
            "prepared_by": prepared_by, "date": "2026-06-26",
            "misc_pct": misc_pct,
            "design_fee": {"mode": "N/A", "value": 0},
            "execution_fee": {"mode": "N/A", "value": 0},
            "supervisor_cost": {"mode": "N/A", "value": 0},
            "additional_charges": add_misc_charges or [],
            "terms_html": DEFAULT_TERMS_HTML,
            "created_by": "Deepak Rao",
            "created_at": now, "updated_at": now,
            "locked": locked, "parent_version_id": None,
            "project_total": 0, "misc_amount": 0, "final_total": 0, "total_amount": 0,
        }
        if approved:
            boq["approved_by"] = "Mr. Kohli"; boq["approved_at"] = now
        await db.boqs.insert_one(boq)
        if add_items_from:
            for i, code in enumerate(["A", "B", "C", "D", "E", "F"]):
                cid = gen_id()
                await db.boq_categories.insert_one({
                    "id": cid, "boq_id": boq_id, "code": code, "name": cat_names[code],
                    "order": i, "collapsed": False, "subtotal": 0,
                })
                items = kohli_items[code]
                for j, row in enumerate(items):
                    desc, unit, qty, rate, ct = row[:5]
                    amount = row[5] if len(row) > 5 else (round(qty * rate, 2) if ct == "M" else 0)
                    item = {
                        "id": gen_id(), "boq_id": boq_id, "category_id": cid, "order": j,
                        "description": desc, "location": "", "unit": unit,
                        "quantity": qty, "rate": rate, "calc_type": ct,
                        "amount": amount, "notes": "",
                        "detail": {}, "hide_from_client": False,
                    }
                    await db.boq_items.insert_one(item)
        await _recompute_totals(boq_id)
        return boq_id

    # Kohli V1 — Approved, locked, matches exact totals
    kohli = by_name.get("Kohli Residence — Interior Renovation")
    if kohli:
        kid_v1 = await _create_boq(kohli, "V1", "approved", "INOS Studio",
                                    misc_pct=10.0, add_items_from=True, locked=True, approved=True)
        # Kohli V2 — Draft (duplicate of V1, editable)
        kid_v2 = await _create_boq(kohli, "V2", "draft", "INOS Studio",
                                    misc_pct=10.0, add_items_from=True)
        await db.boqs.update_one({"id": kid_v2}, {"$set": {"parent_version_id": kid_v1, "revision_reason": "Client requested additions to Furniture", "revision_note": "Adding accent chair & console"}})

    # Other BOQs
    other_seeds = [
        ("Residence 24", "V1", "draft", "INOS Studio", True),
        ("Jain Art Press", "V1", "draft", "INOS Studio", False),
        ("The House Within", "V1", "awaiting_approval", "INOS Studio", True),
        ("Studio Office", "V2", "approved", "INOS Studio", True),
        ("Bansal Villa", "V3", "approved", "INOS Studio", True),
    ]
    for pname, version, status, prep, seed_items in other_seeds:
        proj = by_name.get(pname)
        if not proj: continue
        locked = status in ("approved", "final")
        await _create_boq(proj, version, status, prep, add_items_from=seed_items, locked=locked, approved=locked)




app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


@app.on_event("startup")
async def startup():
    try:
        await seed_if_empty()
        await _seed_catalog_item_presets()
        await _seed_vendors_phase3()
        await seed_phase4(db, gen_id, now_iso)
        await seed_phase5(db, gen_id, now_iso, APP_BASE_URL)
        await _migrate_user_plans()
        logger.info("Seed check complete.")
    except Exception as e:
        logger.exception(f"Seed failed: {e}")


async def _migrate_user_plans():
    """Phase J — ensure every user has plan + is_super_admin. Idempotent."""
    await db.users.update_many({"plan": {"$exists": False}}, {"$set": {"plan": "free_trial"}})
    await db.users.update_many({"is_super_admin": {"$exists": False}}, {"$set": {"is_super_admin": False}})
    await db.users.update_one(
        {"email": "admin@buildcon.in"},
        {"$set": {"is_super_admin": True, "plan": "super_admin"}},
    )


async def _seed_vendors_phase3():
    """Idempotently seed 24 vendors + shortlists + ratings + saved searches."""
    if await db.vendor_shortlists.count_documents({}) > 0:
        return  # Phase 3 already seeded
    # Reset vendors and quotations to full schema
    await db.vendors.delete_many({})
    await db.quotations.delete_many({})
    await db.vendor_ratings.delete_many({})
    await db.vendor_documents.delete_many({})
    await db.vendor_shortlist_items.delete_many({})
    await db.vendor_saved_searches.delete_many({})

    now = now_iso()
    projects = await db.projects.find({}, {"_id": 0}).to_list(50)
    proj_by_name = {p["name"]: p for p in projects}
    proj_list = list(proj_by_name.values())

    VENDORS = [
        # (name, company, vendor_type, primary_category, city, state, rating, completed, price_range, availability_status, verified, preferred, current_assignments, brands, materials, specializations, project_types, phone_last4)
        ("Rakesh Sharma", "Sharma Civil Contractors", "Civil Contractor", "Civil & Masonry", "New Delhi", "Delhi", 4.6, 42, "Mid", "available", True, True, 2, ["Ultratech","Ambuja"], ["Concrete","Bricks","Sand"], ["Turnkey civil","High-rise"], ["Residential","Turnkey"], "4211"),
        ("Vikram Malhotra", "Malhotra Interiors", "Civil Contractor", "Interior Fit-Out", "Gurugram", "Haryana", 4.8, 38, "Premium", "available", True, True, 3, ["Kajaria","Asian Paints"], ["Plywood","Veneer","Marble"], ["Luxury residential","Villa"], ["Residential","Interior Fit-Out"], "8823"),
        ("Sunil Aggarwal", "Aggarwal Construction Co.", "Civil Contractor", "Civil & Masonry", "Noida", "UP", 4.2, 27, "Mid", "busy", True, False, 4, ["JK Cement"], ["Cement","Aggregate"], ["Commercial fit-out"], ["Commercial","Retail"], "5561"),
        ("Rohan Kapoor", "Kapoor & Sons Interiors", "Civil Contractor", "Interior Fit-Out", "New Delhi", "Delhi", 3.9, 18, "Budget", "available", True, False, 1, ["Berger"], ["MDF","Plywood"], ["Compact apartments"], ["Residential"], "1102"),
        ("Anil Verma", "SunLite Electricals", "Electrical Contractor", "Electrical", "New Delhi", "Delhi", 4.4, 33, "Mid", "available", True, True, 2, ["Havells","Legrand","Anchor"], ["Copper wire","MCB"], ["Home automation","Panel wiring"], ["Residential","Commercial"], "9091"),
        ("Deepak Yadav", "PowerGrid Solutions", "Electrical Contractor", "Electrical", "Gurugram", "Haryana", 4.7, 51, "Premium", "available", True, True, 3, ["Schneider","ABB","Siemens"], ["LT panels","Bus ducts"], ["MEP","Panel building"], ["Commercial","Office","Turnkey"], "3212"),
        ("Manish Goel", "Goel Electricals", "Electrical Contractor", "Electrical", "Noida", "UP", 3.7, 12, "Budget", "not_available", False, False, 0, ["Anchor"], ["PVC wire"], ["Retail rewiring"], ["Retail","Residential"], "7783"),
        ("Sanjay Bhatia", "AquaFit Plumbing", "Plumbing Contractor", "Plumbing", "Noida", "UP", 4.3, 22, "Mid", "available", True, False, 1, ["Jaguar","Astral","Ashirvad"], ["CPVC","PVC","Copper"], ["Bathroom","Kitchen"], ["Residential","Interior Fit-Out"], "5544"),
        ("Ramesh Iyer", "Iyer Plumbing Works", "Plumbing Contractor", "Plumbing", "Mumbai", "Maharashtra", 4.1, 19, "Mid", "busy", True, False, 2, ["Jaquar","Kohler"], ["Copper"], ["High-rise"], ["Residential","Commercial"], "2277"),
        ("Naveen Chawla", "Delhi Marble House", "Tile Supplier", "Flooring", "New Delhi", "Delhi", 4.6, 38, "Premium", "available", True, True, 2, ["Kajaria","Somany","Nitco"], ["Marble","Vitrified","Ceramic"], ["Luxury flooring"], ["Residential","Hospitality"], "6633"),
        ("Ajay Bansal", "Bansal Stone Traders", "Stone Supplier", "Flooring", "Gurugram", "Haryana", 4.3, 24, "Mid", "available", True, False, 1, ["Rajasthan Marbles"], ["Marble","Granite","Kota"], ["Natural stone"], ["Residential"], "9922"),
        ("Ravi Mittal", "Prism Sanitaryware", "Sanitaryware Supplier", "Sanitaryware", "New Delhi", "Delhi", 4.5, 29, "Premium", "available", True, True, 2, ["Kohler","Jaquar","American Standard"], ["Ceramic"], ["Luxury bath"], ["Residential","Hospitality"], "1188"),
        ("Prakash Nair", "Coastal Paints & Polish", "Paint Contractor", "Painting", "Mumbai", "Maharashtra", 4.4, 34, "Mid", "available", True, False, 1, ["Asian Paints","Berger","Dulux"], ["Emulsion","Enamel"], ["Texture finish","Metallic"], ["Residential","Commercial"], "7744"),
        ("Suresh Reddy", "Bangalore Paint House", "Paint Contractor", "Painting", "Bangalore", "Karnataka", 4.0, 21, "Budget", "busy", True, False, 3, ["Berger"], ["Emulsion"], ["Volume painting"], ["Residential"], "4455"),
        ("Gaurav Sethi", "Gypcore False Ceiling", "False Ceiling Contractor", "False Ceiling", "Gurugram", "Haryana", 4.5, 26, "Mid", "available", True, True, 1, ["Gyproc","USG Boral"], ["Gypsum","POP"], ["Designer ceiling","Cove"], ["Residential","Commercial"], "3399"),
        ("Kunal Arora", "Arora Ceilings & Panels", "False Ceiling Contractor", "False Ceiling", "New Delhi", "Delhi", 3.6, 9, "Budget", "available", False, False, 0, ["Gyproc"], ["Gypsum"], ["Basic ceiling"], ["Residential"], "0011"),
        ("Harish Kumar", "Woodcraft Interiors", "Furniture Manufacturer", "Furniture & Joinery", "Gurugram", "Haryana", 4.7, 45, "Premium", "available", True, True, 3, ["Century","Greenlam"], ["Plywood","MDF","Veneer"], ["Custom furniture","Wardrobes"], ["Residential","Hospitality","Interior Fit-Out"], "6677"),
        ("Ashok Mehta", "Modula Kitchens", "Modular Kitchen Vendor", "Kitchen", "New Delhi", "Delhi", 4.8, 48, "Premium", "available", True, True, 2, ["Hafele","Blum","Hettich"], ["BWP ply","Acrylic"], ["Luxury modular"], ["Residential"], "2244"),
        ("Vikas Chopra", "Chopra Carpentry", "Carpenter", "Furniture & Joinery", "Noida", "UP", 4.2, 17, "Mid", "available", True, False, 1, ["Century"], ["Plywood"], ["Custom joinery"], ["Residential"], "8899"),
        ("Deepak Sinha", "Sinha Wood Works", "Carpenter", "Furniture & Joinery", "New Delhi", "Delhi", 3.8, 11, "Budget", "not_available", True, False, 0, ["Greenply"], ["MDF"], ["Onsite work"], ["Residential"], "5511"),
        ("Ankit Bhargava", "Lumen Studio Lighting", "Lighting Vendor", "Lighting", "New Delhi", "Delhi", 4.6, 31, "Premium", "available", True, True, 1, ["Philips","Wipro","Osram"], ["LED","Copper"], ["Decorative","Architectural"], ["Residential","Hospitality"], "3311"),
        ("Rajesh Khanna", "CoolBreeze HVAC", "HVAC Contractor", "HVAC", "Gurugram", "Haryana", 4.4, 24, "Premium", "busy", True, False, 3, ["Daikin","Blue Star","Voltas"], ["Copper piping"], ["VRF","Split AC"], ["Commercial","Office","Retail"], "9944"),
        ("Priya Nanda", "GreenScape Landscape", "Landscape Contractor", "Landscape", "Pune", "Maharashtra", 4.5, 28, "Mid", "available", True, False, 2, ["Grasim"], ["Turf","Plants"], ["Terrace garden","Vertical"], ["Residential","Hospitality"], "7799"),
        ("Rohit Sood", "Sood MEP Consultants", "MEP Consultant", "MEP", "New Delhi", "Delhi", 4.9, 36, "Premium", "available", True, True, 2, [], [], ["MEP design","BOQ review"], ["Commercial","Office","Turnkey"], "1155"),
    ]

    ATT_STATES = [
        # (docs_complete, docs_expired, verified, status)
        # blocked (2), attention (3)
    ]
    # Vendors 6 (Manish Goel, Kunal Arora) → blocked
    # Vendors requiring attention: 3 (Manish Goel already blocked doesn't count; use Kunal, Deepak Sinha, and one more)
    blocked_idx = [6]  # Manish Goel index (0-based)
    attention_idx = [15, 19, 6]  # Kunal Arora, Deepak Sinha, Manish Goel

    vendor_ids = []
    for i, row in enumerate(VENDORS):
        (name, company, vtype, cat, city, state, rating, completed, price_range, avail, verified, preferred, current_assign, brands, materials, specializations, project_types, phone_last4) = row
        vid = gen_id()
        vendor_ids.append(vid)
        code = f"BC-V-{str(i+1).zfill(3)}"
        docs_complete = i not in attention_idx
        docs_expired = i in [15]  # Kunal has expired doc
        status = "blocked" if i in blocked_idx else "active"
        doc = {
            "id": vid, "vendor_code": code, "status": status,
            "name": name, "company": company,
            "primary_contact": name, "designation": "Proprietor" if i%2==0 else "Director",
            "phone": f"+91 98{100+i:03d} {phone_last4}", "alt_phone": None,
            "email": f"{name.lower().split()[0]}@{company.lower().replace(' ','').replace('&','').replace('.','')[:16]}.in",
            "website": f"https://{company.lower().replace(' ','').replace('&','').replace('.','')[:20]}.in",
            "address": f"{100+i}, Sector {i+1}", "city": city, "state": state, "pin_code": f"{110000+i*11}",
            "vendor_type": vtype, "primary_category": cat, "relationship_manager": "Deepak Rao",
            "product_categories": [cat], "service_categories": [cat],
            "specializations": specializations, "brands": brands, "materials": materials,
            "equipment_available": ["Basic tools"] if i%2==0 else ["Cutter","Grinder","Welding"],
            "labour_strength": 5 + (i*3) % 50, "certifications": ["ISO 9001"] if verified else [],
            "operating_locations": [city, "Delhi NCR"], "service_radius_km": 50 + (i%5)*10,
            "project_types": project_types, "min_project_area": 500, "max_project_area": 5000+i*500,
            "min_project_value": 100000, "max_project_value": 5000000+i*250000,
            "preferred_project_type": project_types[0], "similar_projects_completed": max(1, completed-5),
            "years_of_experience": 5 + i, "major_clients": ["DLF","Godrej"] if i%3==0 else ["Individuals"],
            "price_range": price_range, "standard_rate_basis": "Per Sq.ft." if "Sq" in cat else "Per Unit",
            "payment_terms": "30-40-20-10", "advance_pct": 30, "credit_period_days": 15,
            "tax_applicability": "GST 18%", "gst_number": f"07ABCDE{1000+i}F1Z{i%10}",
            "pan": f"ABCDE{1000+i:04d}F", "bank_details": {"account_no": f"XXXXXX{1000+i}", "ifsc": "HDFC0000123", "bank": "HDFC Bank"},
            "warranty_available": verified, "warranty_duration": "12 months" if verified else None,
            "delivery_terms": "F.O.R", "installation_included": True, "lead_time_days": 7 + i%14,
            "availability_status": avail, "available_from": now, "available_until": None,
            "max_concurrent_projects": 5, "current_assignments": current_assign,
            "preferred_timeline": "3-6 months", "emergency_availability": i%4==0, "notes": "",
            "verified": verified, "gst_verified": verified, "insurance_available": verified,
            "documents_complete": docs_complete, "documents_expired": docs_expired,
            "preferred": preferred, "previously_used": completed > 5, "recommended": rating >= 4.4,
            "rating": rating, "completed_projects": completed, "on_time_pct": 85 + int(rating*3),
            "avatar_initials": "".join([w[0] for w in name.split()])[:2].upper(),
            "project_ids": [], "created_by": "Deepak Rao", "created_at": now, "updated_at": now,
        }
        await db.vendors.insert_one(doc)

    # Connect ~10 vendors to projects
    if proj_list:
        for i, vid in enumerate(vendor_ids[:10]):
            pid = proj_list[i % len(proj_list)]["id"]
            await db.vendors.update_one({"id": vid}, {"$addToSet": {"project_ids": pid}})

    # Ratings — ~30 across vendors
    import random
    random.seed(42)
    for i in range(30):
        vid = vendor_ids[i % len(vendor_ids)]
        v = await db.vendors.find_one({"id": vid}, {"_id": 0, "rating": 1})
        base = float(v.get("rating") or 4.0)
        def rrand(mu):
            return max(1.0, min(5.0, round(mu + random.uniform(-0.4, 0.4), 1)))
        r = {
            "id": gen_id(), "vendor_id": vid,
            "project_id": (proj_list[i % len(proj_list)]["id"] if proj_list else None),
            "work_quality": rrand(base), "timeline": rrand(base-0.1),
            "communication": rrand(base+0.05), "pricing": rrand(base-0.2),
            "material_quality": rrand(base), "after_sales": rrand(base-0.1),
            "comment": random.choice(["Excellent work.", "Delivered on time.", "Good communication.", "Reliable partner.", "Would recommend.", "Site management could improve.", "Meets expectations."]),
            "given_by": "Deepak Rao", "given_at": now,
        }
        keys = ["work_quality", "timeline", "communication", "pricing", "material_quality", "after_sales"]
        r["avg"] = round(sum(r[k] for k in keys)/len(keys), 2)
        await db.vendor_ratings.insert_one(r)

    # Seed quotations connected to vendors
    quotation_seeds = []
    for i in range(12):
        vid = vendor_ids[i]
        proj = proj_list[i % len(proj_list)] if proj_list else None
        quotation_seeds.append({
            "id": gen_id(), "project_id": proj["id"] if proj else None,
            "project_name": proj["name"] if proj else None,
            "vendor": VENDORS[i][1], "vendor_id": vid, "category": VENDORS[i][3],
            "amount": 150000 + i*45000, "status": ["pending","approved","pending","under_review","approved"][i%5],
            "created_at": now,
        })
    for q in quotation_seeds:
        await db.quotations.insert_one(q)

    # Shortlists
    the_house = proj_by_name.get("The House Within")
    kohli = proj_by_name.get("Kohli Residence — Interior Renovation")
    studio = proj_by_name.get("Studio Office")

    def shortlist(name, project, work_pkg, vendor_indices, remarks):
        sid = gen_id()
        doc = {"id": sid, "name": name,
               "project_id": project["id"] if project else None,
               "project_name": project["name"] if project else None,
               "work_package": work_pkg, "notes": "", "created_by": "Deepak Rao",
               "created_at": now, "updated_at": now}
        return sid, doc, vendor_indices, remarks

    sls = [
        shortlist("The House Within — Electrical", the_house, "Electrical", [4,5,6], ["Best rates","Preferred partner","Backup"]),
        shortlist("Kohli Residence — Painting", kohli, "Painting", [12,13], ["Premium quality","Volume discount"]),
        shortlist("Studio Office — Furniture", studio, "Furniture & Joinery", [16,18,19,17], ["Top pick","Backup","Cost option","Modular kitchen too"]),
    ]
    for sid, sldoc, indices, remarks in sls:
        await db.vendor_shortlists.insert_one(sldoc)
        for order, (vi, rem) in enumerate(zip(indices, remarks)):
            await db.vendor_shortlist_items.insert_one({
                "id": gen_id(), "shortlist_id": sid, "vendor_id": vendor_ids[vi],
                "order": order, "internal_remarks": rem, "added_by": "Deepak Rao", "added_at": now,
            })

    # Saved searches
    admin_user = await db.users.find_one({"email": "admin@buildcon.in"}, {"_id": 0})
    admin_id = admin_user["id"] if admin_user else "system"
    saved = [
        {"name": "Premium Residential Vendors", "filters": {"price_range": "Premium", "project_type": "Residential"}, "scope": "org"},
        {"name": "West Delhi Contractors", "filters": {"city": "New Delhi"}, "scope": "org"},
        {"name": "Available Electrical Vendors", "filters": {"vendor_type": "Electrical Contractor", "available_now": True}, "scope": "personal"},
        {"name": "Verified 4.5+ Rated", "filters": {"verified": True, "min_rating": 4.5}, "scope": "org"},
        {"name": "My Preferred Vendors", "filters": {"preferred": True}, "scope": "personal"},
    ]
    for s in saved:
        await db.vendor_saved_searches.insert_one({"id": gen_id(), "user_id": admin_id, "created_at": now, **s})

    logger.info(f"Phase 3 seed: {len(vendor_ids)} vendors, {len(sls)} shortlists, 30 ratings, {len(saved)} saved searches, {len(quotation_seeds)} quotations.")


async def _seed_catalog_item_presets():
    """Idempotently add preset items + stable unique codes to catalog categories."""
    presets = {
        "Painting": ("PT", [
            {"code": "PT1", "description": "Interior emulsion — 2 coats", "unit": "Sq.ft.", "default_rate": 32, "calc_type": "M"},
            {"code": "PT2", "description": "Exterior weatherproof paint", "unit": "Sq.ft.", "default_rate": 48, "calc_type": "M"},
            {"code": "PT3", "description": "Enamel paint on metal", "unit": "Sq.ft.", "default_rate": 65, "calc_type": "M"},
            {"code": "PT4", "description": "Melamine wood polish", "unit": "Sq.ft.", "default_rate": 85, "calc_type": "M"},
            {"code": "PT5", "description": "Texture finish — designer", "unit": "Sq.ft.", "default_rate": 180, "calc_type": "M"},
            {"code": "PT6", "description": "Primer coat", "unit": "Sq.ft.", "default_rate": 18, "calc_type": "M"},
        ]),
        "Electrical": ("EL", [
            {"code": "EL1", "description": "Concealed wiring point", "unit": "Nos.", "default_rate": 950, "calc_type": "M"},
            {"code": "EL2", "description": "Switchboard — 6 module modular", "unit": "Nos.", "default_rate": 1650, "calc_type": "M"},
            {"code": "EL3", "description": "LED panel light — 15W", "unit": "Nos.", "default_rate": 2200, "calc_type": "M"},
            {"code": "EL4", "description": "Fan point with regulator", "unit": "Nos.", "default_rate": 1250, "calc_type": "M"},
            {"code": "EL5", "description": "AC point — 20A", "unit": "Nos.", "default_rate": 2800, "calc_type": "M"},
            {"code": "EL6", "description": "MCB Distribution Board — 8 way", "unit": "Nos.", "default_rate": 6800, "calc_type": "M"},
            {"code": "EL7", "description": "Electrical rewiring", "unit": "Point", "default_rate": 780, "calc_type": "M"},
        ]),
        "Plumbing": ("PL", [
            {"code": "PL1", "description": "CPVC water supply line", "unit": "Rft.", "default_rate": 220, "calc_type": "M"},
            {"code": "PL2", "description": "PVC waste line", "unit": "Rft.", "default_rate": 180, "calc_type": "M"},
            {"code": "PL3", "description": "Water tank fitting kit", "unit": "Nos.", "default_rate": 3200, "calc_type": "M"},
            {"code": "PL4", "description": "Angle valve — brass CP finish", "unit": "Nos.", "default_rate": 450, "calc_type": "M"},
            {"code": "PL5", "description": "Concealed plumbing point", "unit": "Point", "default_rate": 1850, "calc_type": "M"},
            {"code": "PL6", "description": "Shower mixer with diverter", "unit": "Set", "default_rate": 6800, "calc_type": "M"},
        ]),
        "Sanitaryware": ("SW", [
            {"code": "SW1", "description": "WC with flush tank — branded", "unit": "Set", "default_rate": 12500, "calc_type": "M"},
            {"code": "SW2", "description": "Wash basin with pedestal", "unit": "Set", "default_rate": 6800, "calc_type": "M"},
            {"code": "SW3", "description": "Health faucet with hose", "unit": "Nos.", "default_rate": 850, "calc_type": "M"},
            {"code": "SW4", "description": "Towel rail — SS 304", "unit": "Nos.", "default_rate": 1200, "calc_type": "M"},
            {"code": "SW5", "description": "Soap dispenser — wall mounted", "unit": "Nos.", "default_rate": 650, "calc_type": "M"},
            {"code": "SW6", "description": "Shower head — overhead rain", "unit": "Set", "default_rate": 3800, "calc_type": "M"},
        ]),
        "False Ceiling": ("FC", [
            {"code": "FC1", "description": "Gypsum board false ceiling", "unit": "Sq.ft.", "default_rate": 95, "calc_type": "M"},
            {"code": "FC2", "description": "POP false ceiling", "unit": "Sq.ft.", "default_rate": 85, "calc_type": "M"},
            {"code": "FC3", "description": "Grid ceiling — mineral fibre", "unit": "Sq.ft.", "default_rate": 110, "calc_type": "M"},
            {"code": "FC4", "description": "Cove lighting channel", "unit": "Rft.", "default_rate": 220, "calc_type": "M"},
            {"code": "FC5", "description": "Ceiling patch repair", "unit": "Sq.ft.", "default_rate": 45, "calc_type": "M"},
        ]),
        "Waterproofing": ("WP", [
            {"code": "WP1", "description": "Bathroom waterproofing — 2 coat", "unit": "Sq.ft.", "default_rate": 65, "calc_type": "M"},
            {"code": "WP2", "description": "Terrace waterproofing — APP membrane", "unit": "Sq.ft.", "default_rate": 95, "calc_type": "M"},
            {"code": "WP3", "description": "Wall waterproofing — external", "unit": "Sq.ft.", "default_rate": 55, "calc_type": "M"},
            {"code": "WP4", "description": "Underground sump waterproofing", "unit": "Lump", "default_rate": 28000, "calc_type": "L"},
            {"code": "WP5", "description": "Chemical treatment — anti-leak", "unit": "Sq.ft.", "default_rate": 42, "calc_type": "M"},
        ]),
        "Modular Furniture": ("MF", [
            {"code": "MF1", "description": "Wardrobe — laminate finish", "unit": "Sq.ft.", "default_rate": 1450, "calc_type": "M"},
            {"code": "MF2", "description": "TV unit with backlit panel", "unit": "Rft.", "default_rate": 2200, "calc_type": "M"},
            {"code": "MF3", "description": "Study table with drawer", "unit": "Nos.", "default_rate": 32000, "calc_type": "M"},
            {"code": "MF4", "description": "Bed with hydraulic storage", "unit": "Nos.", "default_rate": 58000, "calc_type": "M"},
            {"code": "MF5", "description": "Shoe rack — tall unit", "unit": "Rft.", "default_rate": 1650, "calc_type": "M"},
        ]),
        "Lighting": ("LT", [
            {"code": "LT1", "description": "Decorative pendant light", "unit": "Nos.", "default_rate": 4500, "calc_type": "M"},
            {"code": "LT2", "description": "Recessed downlight — LED 12W", "unit": "Nos.", "default_rate": 1200, "calc_type": "M"},
            {"code": "LT3", "description": "Cove LED strip — warm white", "unit": "Rft.", "default_rate": 380, "calc_type": "M"},
            {"code": "LT4", "description": "Wall sconce — designer", "unit": "Nos.", "default_rate": 2800, "calc_type": "M"},
            {"code": "LT5", "description": "Chandelier — living room", "unit": "Nos.", "default_rate": 38000, "calc_type": "M"},
        ]),
    }
    # Stable codes for remaining named-only categories
    other_codes = {
        "HVAC": "HV", "Metal Work": "MW", "Glass Work": "GW",
        "Loose Furniture": "LF", "Landscape": "LS",
        "External Development": "ED", "Miscellaneous Works": "MS",
    }
    added_total = 0
    for name, (code, items) in presets.items():
        cat = await db.boq_catalog_categories.find_one({"name": name})
        if not cat:
            continue
        upd = {"code": code}
        if not cat.get("items"):
            upd["items"] = items
            added_total += len(items)
            logger.info(f"Seeded {len(items)} preset items for catalog category '{name}' [{code}]")
        await db.boq_catalog_categories.update_one({"name": name}, {"$set": upd})
    for name, code in other_codes.items():
        await db.boq_catalog_categories.update_one({"name": name}, {"$set": {"code": code}})
    if added_total:
        logger.info(f"Total preset items added across catalog categories: {added_total}")


@app.on_event("shutdown")
async def shutdown_db_client():
    mongo_client.close()
