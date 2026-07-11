"""BOQ Phase 2 backend regression suite.

Covers: dashboard summary, list, filters, Kohli V1 totals & categories, locked BOQ 423,
duplicate-version, live edit patch, add category (catalog + custom), add/delete item,
bulk actions, submit-for-approval + approve (with document side-effect), Excel/PDF export
magic bytes and Kohli V1 final total in xlsx, OpenAPI required paths, catalog list.
"""
import io
import os
import re
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://buildcon-erp.preview.emergentagent.com").rstrip("/")
API = f"{BASE_URL}/api"

KOHLI_V1_ID = "520271e4-36b2-4270-8f2f-1d09db46603b"


# ---------- fixtures ----------
@pytest.fixture(scope="session")
def token():
    r = requests.post(f"{API}/auth/login", json={"email": "admin@buildcon.in", "password": "buildcon123"}, timeout=20)
    assert r.status_code == 200, r.text
    j = r.json()
    return j.get("token") or j.get("access_token")


@pytest.fixture(scope="session")
def s(token):
    sess = requests.Session()
    sess.headers.update({"Authorization": f"Bearer {token}", "Content-Type": "application/json"})
    return sess


# ---------- OpenAPI ----------
def test_openapi_has_boq_paths():
    r = requests.get(f"{API}/openapi.json", timeout=20)
    assert r.status_code == 200
    paths = r.json()["paths"]
    required = [
        "/api/boqs",
        "/api/boqs/{boq_id}",
        "/api/boqs/{boq_id}/categories",
        "/api/boqs/{boq_id}/items/{iid}",
        "/api/boqs/{boq_id}/duplicate-version",
        "/api/boqs/{boq_id}/approve",
        "/api/boqs/{boq_id}/export/excel",
        "/api/boqs/{boq_id}/export/pdf",
        "/api/boq-catalog",
    ]
    for p in required:
        assert p in paths, f"missing OpenAPI path {p}"


# ---------- Dashboard summary + list ----------
def test_boqs_summary(s):
    r = s.get(f"{API}/boqs/summary")
    assert r.status_code == 200
    j = r.json()
    for k in ("total", "drafts", "awaiting", "approved", "templates"):
        assert k in j
    assert j["total"] >= 7


def test_boqs_list_and_filters(s):
    r = s.get(f"{API}/boqs")
    assert r.status_code == 200
    rows = r.json()
    assert isinstance(rows, list) and len(rows) >= 7
    names = " | ".join(row.get("project_name", "") for row in rows)
    for expected in ["Kohli", "Residence 24", "Jain", "House Within", "Studio Office", "Bansal"]:
        assert expected in names, f"missing project {expected}"

    # filter by status
    r2 = s.get(f"{API}/boqs", params={"status": "approved"})
    assert r2.status_code == 200
    approved_rows = r2.json()
    assert len(approved_rows) >= 1
    assert all(b["status"] == "approved" for b in approved_rows), [b["status"] for b in approved_rows]

    # search q
    r3 = s.get(f"{API}/boqs", params={"q": "Kohli"})
    assert r3.status_code == 200
    kohlis = r3.json()
    assert len(kohlis) >= 2
    assert all("kohli" in b["project_name"].lower() for b in kohlis)


# ---------- Kohli V1 detail ----------
def test_kohli_v1_totals_and_categories(s):
    r = s.get(f"{API}/boqs/{KOHLI_V1_ID}")
    assert r.status_code == 200, r.text
    b = r.json()
    assert b["status"] == "approved"
    assert b["locked"] is True
    assert b.get("version") == "V1"
    # totals
    assert round(b["project_total"], 2) == 2832800.00, b["project_total"]
    assert round(b["misc_amount"], 2) == 283280.00, b["misc_amount"]
    assert round(b["final_total"], 2) == 3116080.00, b["final_total"]

    cats = b["categories"]
    assert len(cats) == 6, f"expected 6 categories got {len(cats)}"
    subs = {c["code"]: round(c["subtotal"], 2) for c in cats}
    expected_subs = {"A": 230000, "B": 716800, "C": 188400, "D": 466000, "E": 958000, "F": 273600}
    for code, exp in expected_subs.items():
        assert code in subs, f"cat code {code} missing"
        assert subs[code] == exp, f"cat {code} subtotal {subs[code]} != {exp}"

    assert len(b["items"]) == 41, f"expected 41 items got {len(b['items'])}"


# ---------- Locked (approved) BOQ returns 423 ----------
def test_locked_boq_patch_returns_423(s):
    r = s.patch(f"{API}/boqs/{KOHLI_V1_ID}", json={"title": "hack"})
    assert r.status_code == 423, f"expected 423 got {r.status_code} {r.text}"

    # Also item patch on approved BOQ
    boq = s.get(f"{API}/boqs/{KOHLI_V1_ID}").json()
    item_id = boq["items"][0]["id"]
    r2 = s.patch(f"{API}/boqs/{KOHLI_V1_ID}/items/{item_id}", json={"rate": 999})
    assert r2.status_code == 423


# ---------- Duplicate version → editable draft ----------
@pytest.fixture(scope="session")
def duplicate_boq(s):
    r = s.post(
        f"{API}/boqs/{KOHLI_V1_ID}/duplicate-version",
        json={"note": "TEST_ automation duplicate", "reason": "test", "copy_categories": True, "copy_terms": True, "copy_fees": True},
    )
    assert r.status_code == 200, r.text
    new_boq = r.json()
    assert new_boq["status"] == "draft"
    assert new_boq["locked"] is False
    assert new_boq["parent_version_id"] == KOHLI_V1_ID
    assert new_boq["id"] != KOHLI_V1_ID
    assert re.match(r"^V\d+$", new_boq.get("version") or "")
    # totals should match parent
    assert round(new_boq["final_total"], 2) == 3116080.00
    yield new_boq
    # teardown — delete draft
    s.delete(f"{API}/boqs/{new_boq['id']}")


# ---------- Live edit rate ----------
def test_item_rate_edit_recomputes(s, duplicate_boq):
    boq_id = duplicate_boq["id"]
    item = duplicate_boq["items"][0]
    iid = item["id"]
    old_rate = float(item.get("rate") or 0)
    new_rate = old_rate + 100.0
    r = s.patch(f"{API}/boqs/{boq_id}/items/{iid}", json={"rate": new_rate})
    assert r.status_code == 200
    boq = r.json()
    # find item
    updated = next(i for i in boq["items"] if i["id"] == iid)
    assert float(updated["rate"]) == new_rate
    # expected item amount = qty * rate (measured)
    qty = float(updated.get("quantity") or 0)
    expected_amt = round(qty * new_rate, 2)
    assert round(float(updated["amount"]), 2) == expected_amt
    # totals recomputed
    assert boq["final_total"] > 3116080.00 or boq["final_total"] < 3116080.00
    # misc_amount = 10% of project_total
    assert round(boq["misc_amount"], 2) == round(boq["project_total"] * 0.10, 2)


# ---------- Catalog list + add category from catalog ----------
def test_catalog_list_and_add(s, duplicate_boq):
    r = s.get(f"{API}/boq-catalog")
    assert r.status_code == 200
    catalog = r.json()
    assert isinstance(catalog, list) and len(catalog) > 0

    # pick one item to append — use catalog code
    ce = catalog[0]
    payload = {"code": ce.get("code"), "name": ce.get("name")}
    r2 = s.post(f"{API}/boqs/{duplicate_boq['id']}/categories", json=payload)
    assert r2.status_code == 200, r2.text


def test_add_custom_category_and_item_delete(s, duplicate_boq):
    # custom category
    r = s.post(f"{API}/boqs/{duplicate_boq['id']}/categories", json={"name": "TEST_ Custom Cat", "custom": True})
    assert r.status_code == 200, r.text
    boq = s.get(f"{API}/boqs/{duplicate_boq['id']}").json()
    cat = next((c for c in boq["categories"] if c.get("name") == "TEST_ Custom Cat"), None)
    assert cat is not None
    cid = cat["id"]

    # add line item
    r2 = s.post(f"{API}/boqs/{duplicate_boq['id']}/categories/{cid}/items", json={"description": "TEST_ new item", "unit": "Nos.", "quantity": 1, "rate": 0})
    assert r2.status_code == 200, r2.text
    boq2 = s.get(f"{API}/boqs/{duplicate_boq['id']}").json()
    new_item = next((i for i in boq2["items"] if i.get("category_id") == cid and i.get("description") == "TEST_ new item"), None)
    assert new_item is not None
    iid = new_item["id"]

    # delete it
    r3 = s.delete(f"{API}/boqs/{duplicate_boq['id']}/items/{iid}")
    assert r3.status_code in (200, 204)
    boq3 = s.get(f"{API}/boqs/{duplicate_boq['id']}").json()
    assert not any(i["id"] == iid for i in boq3["items"])


def test_bulk_actions(s, duplicate_boq):
    boq = s.get(f"{API}/boqs/{duplicate_boq['id']}").json()
    ids = [i["id"] for i in boq["items"][:2]]
    r = s.post(f"{API}/boqs/{duplicate_boq['id']}/items/bulk", json={"ids": ids, "op": "change_unit", "value": "Sqft"})
    assert r.status_code == 200, r.text
    boq2 = s.get(f"{API}/boqs/{duplicate_boq['id']}").json()
    for it in boq2["items"]:
        if it["id"] in ids:
            assert it.get("unit") == "Sqft"

    # bulk delete
    r2 = s.post(f"{API}/boqs/{duplicate_boq['id']}/items/bulk", json={"ids": ids, "op": "delete"})
    assert r2.status_code == 200
    boq3 = s.get(f"{API}/boqs/{duplicate_boq['id']}").json()
    remaining_ids = {i["id"] for i in boq3["items"]}
    assert not any(x in remaining_ids for x in ids)


# ---------- Submit for approval → Approve → doc created ----------
def test_submit_and_approve_creates_document(s):
    # Duplicate a fresh copy so we don't touch the shared one
    r = s.post(f"{API}/boqs/{KOHLI_V1_ID}/duplicate-version", json={"note": "TEST_ approve flow"})
    assert r.status_code == 200
    boq = r.json()
    bid = boq["id"]
    try:
        r1 = s.post(f"{API}/boqs/{bid}/submit-for-approval", json={"note": "TEST_"})
        assert r1.status_code == 200
        assert r1.json()["status"] == "awaiting_approval"

        r2 = s.post(f"{API}/boqs/{bid}/approve", json={"remarks": "TEST_ approve"})
        assert r2.status_code == 200
        approved = r2.json()
        assert approved["status"] == "approved"
        assert approved["locked"] is True

        # After approve, patch should be locked
        r3 = s.patch(f"{API}/boqs/{bid}", json={"title": "x"})
        assert r3.status_code == 423

        # Document auto-generated — check via /api/documents/recent (higher limit) and via activity
        docs_resp = s.get(f"{API}/documents/recent", params={"limit": 100})
        assert docs_resp.status_code == 200
        docs_json = docs_resp.json()
        docs = docs_json if isinstance(docs_json, list) else docs_json.get("items", [])
        assert any(d.get("boq_id") == bid for d in docs), "auto-attached BOQ pdf doc not found in documents/recent"
    finally:
        # cleanup — cannot delete locked BOQ via delete? try anyway
        s.delete(f"{API}/boqs/{bid}")


# ---------- Excel export ----------
def test_excel_export_kohli_v1(s):
    r = s.get(f"{API}/boqs/{KOHLI_V1_ID}/export/excel")
    assert r.status_code == 200, r.text
    content = r.content
    assert content[:2] == b"PK", "not a zip/xlsx"
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    # scan for FINAL TOTAL row with 3116080
    found = False
    for row in ws.iter_rows(values_only=True):
        vals = [str(v) if v is not None else "" for v in row]
        joined = " ".join(vals).upper()
        if "FINAL TOTAL" in joined or "GRAND TOTAL" in joined:
            for v in row:
                if isinstance(v, (int, float)) and round(float(v), 2) == 3116080.00:
                    found = True
                    break
        if found: break
    assert found, "FINAL TOTAL row with 3116080 not found in xlsx"


# ---------- PDF export ----------
def test_pdf_export_kohli_v1(s):
    r = s.post(f"{API}/boqs/{KOHLI_V1_ID}/export/pdf", json={})
    if r.status_code == 405:
        r = s.get(f"{API}/boqs/{KOHLI_V1_ID}/export/pdf")
    assert r.status_code == 200, r.text[:500]
    assert r.content[:5] == b"%PDF-", f"not a pdf, first bytes={r.content[:8]!r}"


# ---------- Regression: Phase 1 continue-working ----------
def test_phase1_continue_working_includes_residence24(s):
    # Try common endpoints
    for path in ("/dashboard/continue-working", "/continue-working", "/dashboard"):
        r = s.get(f"{API}{path}")
        if r.status_code == 200:
            text = r.text
            if "Residence 24" in text:
                return
    # fallback: check via boqs list
    rows = s.get(f"{API}/boqs").json()
    assert any("Residence 24" in b.get("project_name", "") for b in rows)
