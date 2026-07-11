"""Phase 4 — Quotations App backend module.

Exposes `register_phase4(app, api, db, deps)` which wires all quotation routes.

`deps` provides shared helpers from server.py:
  - get_current_user (auth dep)
  - gen_id, now_iso, _fmt_inr
  - UPLOAD_ROOT (Path), get_file (route already exists)
"""
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form, Body, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Any, Dict
from datetime import datetime, timezone, timedelta
import io
import copy
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

logger = logging.getLogger("phase4")

# ---------- Constants ----------
QUOTATION_STATUSES = [
    "draft", "requested", "received", "under_review", "awaiting_approval",
    "returned", "approved", "rejected", "selected", "not_selected",
    "expired", "archived",
]
EDITABLE_STATUSES = {"draft", "returned"}
LOCKED_STATUSES = {"approved", "selected", "archived"}


def register_phase4(app, api, db, deps: dict):
    get_current_user = deps["get_current_user"]
    require_internal = deps.get("require_internal") or get_current_user
    gen_id = deps["gen_id"]
    now_iso = deps["now_iso"]
    fmt_inr = deps["fmt_inr"]
    enforce_free_trial_cap = deps.get("enforce_free_trial_cap")

    # ---------- helpers ----------
    async def _next_quotation_number() -> str:
        year = datetime.now(timezone.utc).year
        prefix = f"QT-{year}-"
        latest = await db.quotations.find({"quotation_number": {"$regex": f"^{prefix}"}}, {"_id": 0, "quotation_number": 1}).sort("quotation_number", -1).limit(1).to_list(1)
        n = 1
        if latest:
            try:
                n = int(latest[0]["quotation_number"].split("-")[-1]) + 1
            except Exception:
                n = 1
        return f"{prefix}{n:04d}"

    async def _boq_estimate_for_item(boq_item_id: Optional[str]) -> Optional[dict]:
        if not boq_item_id:
            return None
        it = await db.boq_items.find_one({"id": boq_item_id}, {"_id": 0})
        if not it:
            return None
        return {"boq_rate": float(it.get("rate") or 0), "boq_qty": float(it.get("quantity") or 0), "boq_amount": float(it.get("amount") or 0), "boq_unit": it.get("unit"), "boq_description": it.get("description")}

    def _calc_item_amount(qty, rate, calc_type):
        if calc_type == "L":
            return float(rate or 0)
        return round(float(qty or 0) * float(rate or 0), 2)

    async def _recompute_totals(qid: str) -> dict:
        q = await db.quotations.find_one({"id": qid}, {"_id": 0})
        if not q:
            raise HTTPException(404, "Quotation not found")
        items = await db.quotation_items.find({"quotation_id": qid}, {"_id": 0}).to_list(2000)
        base = 0.0
        tax = 0.0
        boq_ref_total_quote = 0.0
        boq_ref_total_boq = 0.0
        for it in items:
            amt = _calc_item_amount(it.get("quantity"), it.get("rate"), it.get("calc_type"))
            base += amt
            tp = float(it.get("tax_pct") or 0)
            tax += round(amt * tp / 100.0, 2)
            if it.get("boq_item_id"):
                bi = await db.boq_items.find_one({"id": it["boq_item_id"]}, {"_id": 0})
                if bi:
                    boq_ref_total_boq += float(bi.get("amount") or 0)
                    boq_ref_total_quote += amt
        terms = q.get("commercial_terms") or {}
        transport = float(terms.get("transportation_amount") or 0)
        installation = float(terms.get("installation_amount") or 0)
        additional = sum(float(a.get("amount") or 0) for a in (q.get("additional_charges") or []))
        discount = float(q.get("discount") or 0)
        total = round(base + tax + transport + installation + additional - discount, 2)
        variation = None
        if boq_ref_total_boq > 0:
            variation = round((boq_ref_total_quote - boq_ref_total_boq) / boq_ref_total_boq * 100.0, 2)
        subtotals = {"base": round(base, 2), "tax": round(tax, 2), "transport": transport, "installation": installation, "additional": round(additional, 2), "discount": discount, "total": total}
        await db.quotations.update_one({"id": qid}, {"$set": {"subtotals": subtotals, "boq_variation_pct": variation, "updated_at": now_iso()}})
        return {"subtotals": subtotals, "boq_variation_pct": variation}

    async def _full_quotation(qid: str) -> dict:
        q = await db.quotations.find_one({"id": qid}, {"_id": 0})
        if not q:
            raise HTTPException(404, "Quotation not found")
        items = await db.quotation_items.find({"quotation_id": qid}, {"_id": 0}).sort("order", 1).to_list(2000)
        atts = await db.quotation_attachments.find({"quotation_id": qid}, {"_id": 0}).to_list(200)
        acts = await db.quotation_activity.find({"quotation_id": qid}, {"_id": 0}).sort("at", -1).to_list(200)
        vendor = await db.vendors.find_one({"id": q.get("vendor_id")}, {"_id": 0}) if q.get("vendor_id") else None
        project = await db.projects.find_one({"id": q.get("project_id")}, {"_id": 0}) if q.get("project_id") else None
        # Attach boq refs
        for it in items:
            if it.get("boq_item_id"):
                ref = await _boq_estimate_for_item(it["boq_item_id"])
                if ref:
                    it["boq_ref_data"] = ref
                    amt = _calc_item_amount(it.get("quantity"), it.get("rate"), it.get("calc_type"))
                    it["variation_pct"] = round((amt - ref["boq_amount"]) / ref["boq_amount"] * 100.0, 2) if ref["boq_amount"] > 0 else None
        q["items"] = items
        q["attachments"] = atts
        q["activity"] = acts
        q["vendor"] = vendor
        q["project"] = project
        # Validity days remaining
        vu = q.get("valid_until")
        if vu:
            try:
                d = datetime.fromisoformat(vu).date()
                today = datetime.now(timezone.utc).date()
                q["days_remaining"] = (d - today).days
            except Exception:
                q["days_remaining"] = None
        # Approval history
        q["approval_history"] = [a for a in acts if a.get("action") in ("submit", "send_to_reviewer", "send_to_vendor", "approve", "return", "reject", "request_clarification", "mark_selected", "mark_not_selected")]
        return q

    async def _log_activity(qid: str, action: str, actor: str, meta: dict = None):
        await db.quotation_activity.insert_one({"id": gen_id(), "quotation_id": qid, "action": action, "actor": actor, "at": now_iso(), "meta": meta or {}})

    def _ensure_editable(q: dict):
        if q.get("locked") or q.get("status") not in EDITABLE_STATUSES:
            raise HTTPException(423, f"Quotation is locked (status: {q.get('status')}). Duplicate a version to edit.")

    # ---------- Models ----------
    class QuotationCreateIn(BaseModel):
        model_config = ConfigDict(extra="allow")
        project_id: Optional[str] = None
        vendor_id: Optional[str] = None
        title: Optional[str] = None
        work_category: Optional[str] = None
        quotation_number: Optional[str] = None
        quotation_date: Optional[str] = None
        valid_until: Optional[str] = None
        currency: str = "INR"
        tax_config: Optional[str] = "GST 18%"
        source: str = "created"

    class QuotationPatchIn(BaseModel):
        model_config = ConfigDict(extra="allow")

    class ItemIn(BaseModel):
        model_config = ConfigDict(extra="allow")
        description: str
        unit: str = "Nos."
        quantity: float = 1.0
        rate: float = 0.0
        tax_pct: float = 18.0
        calc_type: str = "M"
        boq_item_id: Optional[str] = None
        boq_ref: Optional[str] = None
        location: Optional[str] = None
        remarks: Optional[str] = None
        category_id: Optional[str] = None

    class ItemPatchIn(BaseModel):
        model_config = ConfigDict(extra="allow")

    class ItemsReorderIn(BaseModel):
        order: List[str]

    class BulkItemsIn(BaseModel):
        items: List[ItemIn]

    class ImportBoqIn(BaseModel):
        boq_id: str
        item_ids: List[str]

    class RemarkIn(BaseModel):
        remarks: Optional[str] = ""

    class RFQIn(BaseModel):
        project_id: str
        work_category: str
        vendor_ids: List[str]
        deadline: Optional[str] = None

    class SaveComparisonIn(BaseModel):
        name: str
        project_id: Optional[str] = None
        work_category: Optional[str] = None
        quotation_ids: List[str]

    # ---------- Read endpoints ----------
    @api.get("/quotations/summary")
    async def quotations_summary(current=Depends(require_internal)):
        c = lambda **kw: db.quotations.count_documents(kw)
        today = datetime.now(timezone.utc).date()
        soon = (today + timedelta(days=7)).isoformat()
        expiring_count = await db.quotations.count_documents({"status": {"$in": ["approved", "selected", "under_review", "awaiting_approval", "received"]}, "valid_until": {"$gte": today.isoformat(), "$lte": soon}})
        return {
            "total": await c(),
            "drafts": await c(status="draft"),
            "requested": await c(status="requested"),
            "received": await c(status="received"),
            "under_review": await c(status="under_review"),
            "awaiting_approval": await c(status="awaiting_approval"),
            "approved": await c(status="approved"),
            "selected": await c(status="selected"),
            "rejected": await c(status="rejected"),
            "expired": await c(status="expired"),
            "expiring_soon": expiring_count,
        }

    @api.get("/quotations/awaiting-action")
    async def awaiting_action(current=Depends(require_internal)):
        today = datetime.now(timezone.utc).date()
        soon = (today + timedelta(days=7)).isoformat()
        needs_review = await db.quotations.count_documents({"status": "under_review"})
        returned = await db.quotations.count_documents({"status": "returned"})
        expiring = await db.quotations.count_documents({"status": {"$in": ["received", "under_review", "awaiting_approval", "approved"]}, "valid_until": {"$gte": today.isoformat(), "$lte": soon}})
        missing_response = await db.quotations.count_documents({"status": "requested"})
        above_boq = await db.quotations.count_documents({"boq_variation_pct": {"$gte": 10}})
        pending_approval = await db.quotations.count_documents({"status": "awaiting_approval"})
        return {
            "needs_review": needs_review, "returned": returned, "expiring": expiring,
            "missing_response": missing_response, "above_boq": above_boq, "pending_approval": pending_approval,
        }

    @api.get("/quotations/recent-comparisons")
    async def recent_comparisons(limit: int = 6, current=Depends(require_internal)):
        rows = await db.quotation_comparisons.find({}, {"_id": 0}).sort("saved_at", -1).limit(limit).to_list(limit)
        for r in rows:
            r["vendor_count"] = len(r.get("quotation_ids", []))
        return rows

    @api.get("/quotations/continue-working")
    async def continue_working(current=Depends(require_internal)):
        rows = await db.quotations.find({"status": "draft"}, {"_id": 0}).sort("updated_at", -1).limit(3).to_list(3)
        return rows

    @api.get("/quotations")
    async def quotations_list(
        q: Optional[str] = None,
        status: Optional[str] = None,
        project_id: Optional[str] = None,
        vendor_id: Optional[str] = None,
        work_category: Optional[str] = None,
        source: Optional[str] = None,
        quotation_type: Optional[str] = None,
        created_by: Optional[str] = None,
        reviewer_id: Optional[str] = None,
        min_amount: Optional[float] = None,
        max_amount: Optional[float] = None,
        min_variation: Optional[float] = None,
        max_variation: Optional[float] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        validity_from: Optional[str] = None,
        validity_to: Optional[str] = None,
        limit: int = 200,
        current=Depends(require_internal),
    ):
        query: dict = {}
        if status: query["status"] = status
        if project_id: query["project_id"] = project_id
        if vendor_id: query["vendor_id"] = vendor_id
        if work_category: query["work_category"] = work_category
        if source: query["source"] = source
        if created_by: query["created_by"] = created_by
        if reviewer_id: query["reviewer_id"] = reviewer_id
        if quotation_type: query["quotation_type"] = quotation_type
        if min_amount is not None or max_amount is not None:
            amt_q = {}
            if min_amount is not None: amt_q["$gte"] = min_amount
            if max_amount is not None: amt_q["$lte"] = max_amount
            query["subtotals.total"] = amt_q
        if min_variation is not None or max_variation is not None:
            vq = {}
            if min_variation is not None: vq["$gte"] = min_variation
            if max_variation is not None: vq["$lte"] = max_variation
            query["boq_variation_pct"] = vq
        if date_from or date_to:
            dq = {}
            if date_from: dq["$gte"] = date_from
            if date_to: dq["$lte"] = date_to
            query["quotation_date"] = dq
        if validity_from or validity_to:
            vq = {}
            if validity_from: vq["$gte"] = validity_from
            if validity_to: vq["$lte"] = validity_to
            query["valid_until"] = vq
        if q:
            ql = q.strip()
            query["$or"] = [
                {"title": {"$regex": ql, "$options": "i"}},
                {"quotation_number": {"$regex": ql, "$options": "i"}},
                {"vendor_name": {"$regex": ql, "$options": "i"}},
                {"project_name": {"$regex": ql, "$options": "i"}},
                {"work_category": {"$regex": ql, "$options": "i"}},
            ]
        rows = await db.quotations.find(query, {"_id": 0}).sort("updated_at", -1).limit(limit).to_list(limit)
        return rows

    @api.get("/quotations/{qid}")
    async def quotation_get(qid: str, current=Depends(require_internal)):
        return await _full_quotation(qid)

    @api.post("/quotations")
    async def quotation_create(payload: QuotationCreateIn, current=Depends(require_internal)):
        if enforce_free_trial_cap:
            await enforce_free_trial_cap(current, "quotations")
        data = payload.model_dump()
        qid = gen_id()
        num = data.get("quotation_number") or await _next_quotation_number()
        # Ensure unique
        while await db.quotations.find_one({"quotation_number": num}):
            num = await _next_quotation_number()
        vendor = await db.vendors.find_one({"id": data.get("vendor_id")}, {"_id": 0}) if data.get("vendor_id") else None
        project = await db.projects.find_one({"id": data.get("project_id")}, {"_id": 0}) if data.get("project_id") else None
        doc = {
            "id": qid,
            "quotation_number": num,
            "title": data.get("title") or "Untitled Quotation",
            "project_id": data.get("project_id"),
            "project_name": project.get("name") if project else None,
            "vendor_id": data.get("vendor_id"),
            "vendor_name": vendor.get("company") if vendor else (vendor.get("name") if vendor else None),
            "work_category": data.get("work_category"),
            "source": data.get("source") or "created",
            "status": "draft",
            "version": 1,
            "parent_version_id": None,
            "currency": data.get("currency") or "INR",
            "quotation_date": data.get("quotation_date") or datetime.now(timezone.utc).date().isoformat(),
            "valid_until": data.get("valid_until") or (datetime.now(timezone.utc).date() + timedelta(days=30)).isoformat(),
            "prepared_by": current.get("name"),
            "reviewer_id": data.get("reviewer_id"),
            "tax_config": data.get("tax_config") or "GST 18%",
            "commercial_terms": data.get("commercial_terms") or {
                "payment_terms": "30% advance, 40% on delivery, 20% on installation, 10% on handover",
                "advance_pct": 30, "credit_period_days": 15,
                "transportation_amount": 0, "transportation_included": True,
                "installation_amount": 0, "installation_included": True,
                "warranty_months": 12, "warranty_notes": "",
                "delivery_timeline_days": 15, "completion_timeline_days": 45,
                "exclusions": [], "special_conditions": "",
            },
            "additional_charges": data.get("additional_charges") or [],
            "discount": data.get("discount") or 0,
            "subtotals": {"base": 0, "tax": 0, "transport": 0, "installation": 0, "additional": 0, "discount": 0, "total": 0},
            "boq_variation_pct": None,
            "finalized": False, "selected": False, "locked": False,
            "quotation_type": data.get("quotation_type") or "Vendor",
            "created_by": current.get("name"),
            "created_by_id": current.get("id"),
            "created_at": now_iso(),
            "updated_at": now_iso(),
        }
        await db.quotations.insert_one(doc)
        await _log_activity(qid, "create", current.get("name"), {"quotation_number": num})
        return {"id": qid, "quotation_number": num}

    @api.patch("/quotations/{qid}")
    async def quotation_patch(qid: str, payload: QuotationPatchIn, current=Depends(require_internal)):
        q = await db.quotations.find_one({"id": qid}, {"_id": 0})
        if not q: raise HTTPException(404, "Quotation not found")
        _ensure_editable(q)
        data = payload.model_dump(exclude_unset=True)
        # Prevent editing protected fields
        for k in ("id", "quotation_number", "status", "locked", "version", "selected", "finalized", "created_at", "created_by"):
            data.pop(k, None)
        if data.get("vendor_id") and data["vendor_id"] != q.get("vendor_id"):
            v = await db.vendors.find_one({"id": data["vendor_id"]}, {"_id": 0})
            data["vendor_name"] = v.get("company") if v else None
        if data.get("project_id") and data["project_id"] != q.get("project_id"):
            p = await db.projects.find_one({"id": data["project_id"]}, {"_id": 0})
            data["project_name"] = p.get("name") if p else None
        data["updated_at"] = now_iso()
        await db.quotations.update_one({"id": qid}, {"$set": data})
        await _recompute_totals(qid)
        return await _full_quotation(qid)

    @api.delete("/quotations/{qid}")
    async def quotation_delete(qid: str, current=Depends(require_internal)):
        q = await db.quotations.find_one({"id": qid}, {"_id": 0})
        if not q: raise HTTPException(404, "Quotation not found")
        if q.get("status") != "draft":
            raise HTTPException(423, "Only draft quotations can be deleted.")
        await db.quotations.delete_one({"id": qid})
        await db.quotation_items.delete_many({"quotation_id": qid})
        await db.quotation_attachments.delete_many({"quotation_id": qid})
        await db.quotation_activity.delete_many({"quotation_id": qid})
        return {"ok": True}

    # ---------- Items ----------
    @api.post("/quotations/{qid}/items")
    async def item_add(qid: str, payload: ItemIn, current=Depends(require_internal)):
        q = await db.quotations.find_one({"id": qid}, {"_id": 0})
        if not q: raise HTTPException(404, "Quotation not found")
        _ensure_editable(q)
        cnt = await db.quotation_items.count_documents({"quotation_id": qid})
        d = payload.model_dump()
        it = {
            "id": gen_id(), "quotation_id": qid, "order": cnt,
            **d,
            "amount": _calc_item_amount(d.get("quantity"), d.get("rate"), d.get("calc_type")),
            "created_at": now_iso(),
        }
        await db.quotation_items.insert_one(it)
        await _recompute_totals(qid)
        return it

    @api.patch("/quotations/{qid}/items/{iid}")
    async def item_patch(qid: str, iid: str, payload: ItemPatchIn, current=Depends(require_internal)):
        q = await db.quotations.find_one({"id": qid}, {"_id": 0})
        if not q: raise HTTPException(404, "Quotation not found")
        _ensure_editable(q)
        cur = await db.quotation_items.find_one({"id": iid, "quotation_id": qid}, {"_id": 0})
        if not cur: raise HTTPException(404, "Item not found")
        d = payload.model_dump(exclude_unset=True)
        merged = {**cur, **d}
        merged["amount"] = _calc_item_amount(merged.get("quantity"), merged.get("rate"), merged.get("calc_type"))
        await db.quotation_items.update_one({"id": iid}, {"$set": {**d, "amount": merged["amount"]}})
        await _recompute_totals(qid)
        return {"ok": True, "amount": merged["amount"]}

    @api.delete("/quotations/{qid}/items/{iid}")
    async def item_delete(qid: str, iid: str, current=Depends(require_internal)):
        q = await db.quotations.find_one({"id": qid}, {"_id": 0})
        if not q: raise HTTPException(404, "Quotation not found")
        _ensure_editable(q)
        await db.quotation_items.delete_one({"id": iid, "quotation_id": qid})
        await _recompute_totals(qid)
        return {"ok": True}

    @api.post("/quotations/{qid}/items/reorder")
    async def items_reorder(qid: str, payload: ItemsReorderIn, current=Depends(require_internal)):
        q = await db.quotations.find_one({"id": qid}, {"_id": 0})
        if not q: raise HTTPException(404, "Quotation not found")
        _ensure_editable(q)
        for i, iid in enumerate(payload.order):
            await db.quotation_items.update_one({"id": iid, "quotation_id": qid}, {"$set": {"order": i}})
        return {"ok": True}

    @api.post("/quotations/{qid}/items/bulk")
    async def items_bulk(qid: str, payload: BulkItemsIn, current=Depends(require_internal)):
        q = await db.quotations.find_one({"id": qid}, {"_id": 0})
        if not q: raise HTTPException(404, "Quotation not found")
        _ensure_editable(q)
        cnt = await db.quotation_items.count_documents({"quotation_id": qid})
        for i, it in enumerate(payload.items):
            d = it.model_dump()
            doc = {
                "id": gen_id(), "quotation_id": qid, "order": cnt + i,
                **d,
                "amount": _calc_item_amount(d.get("quantity"), d.get("rate"), d.get("calc_type")),
                "created_at": now_iso(),
            }
            await db.quotation_items.insert_one(doc)
        await _recompute_totals(qid)
        return {"ok": True, "added": len(payload.items)}

    @api.post("/quotations/{qid}/items/import-from-boq")
    async def items_import_from_boq(qid: str, payload: ImportBoqIn, current=Depends(require_internal)):
        q = await db.quotations.find_one({"id": qid}, {"_id": 0})
        if not q: raise HTTPException(404, "Quotation not found")
        _ensure_editable(q)
        boq = await db.boqs.find_one({"id": payload.boq_id}, {"_id": 0})
        if not boq: raise HTTPException(404, "BOQ not found")
        items = await db.boq_items.find({"id": {"$in": payload.item_ids}, "boq_id": payload.boq_id}, {"_id": 0}).to_list(500)
        cnt = await db.quotation_items.count_documents({"quotation_id": qid})
        for i, bi in enumerate(items):
            doc = {
                "id": gen_id(), "quotation_id": qid, "order": cnt + i,
                "description": bi.get("description"),
                "unit": bi.get("unit"),
                "quantity": float(bi.get("quantity") or 0),
                "rate": float(bi.get("rate") or 0),
                "amount": float(bi.get("amount") or 0),
                "tax_pct": 18.0,
                "calc_type": bi.get("calc_type") or "M",
                "boq_item_id": bi.get("id"),
                "boq_ref": bi.get("code") or (bi.get("description") or "")[:40],
                "location": bi.get("location"),
                "remarks": "",
                "created_at": now_iso(),
            }
            await db.quotation_items.insert_one(doc)
        await _recompute_totals(qid)
        return {"ok": True, "added": len(items)}

    # ---------- Attachments ----------
    @api.post("/quotations/{qid}/attachments")
    async def qatt_upload(qid: str, kind: str = Form("attachment"), file: UploadFile = File(...), current=Depends(require_internal)):
        q = await db.quotations.find_one({"id": qid}, {"_id": 0})
        if not q: raise HTTPException(404, "Quotation not found")
        contents = await file.read()
        if len(contents) > 10 * 1024 * 1024:
            raise HTTPException(413, "File exceeds 10MB")
        UPLOAD_ROOT = deps["UPLOAD_ROOT"]
        qdir = UPLOAD_ROOT / "quotations" / qid
        qdir.mkdir(parents=True, exist_ok=True)
        aid = gen_id()
        safe = "".join(c if c.isalnum() or c in "._-" else "_" for c in (file.filename or "file"))
        dest = qdir / f"{aid}-{safe}"
        dest.write_bytes(contents)
        rel = f"quotations/{qid}/{aid}-{safe}"
        doc = {"id": aid, "quotation_id": qid, "kind": kind, "filename": file.filename, "path": rel, "size": len(contents), "mime": file.content_type or "application/octet-stream", "uploaded_by": current.get("name"), "uploaded_at": now_iso()}
        await db.quotation_attachments.insert_one(doc)
        await _log_activity(qid, "attach", current.get("name"), {"filename": file.filename})
        doc.pop("_id", None)
        return doc

    @api.delete("/quotations/{qid}/attachments/{aid}")
    async def qatt_delete(qid: str, aid: str, current=Depends(require_internal)):
        d = await db.quotation_attachments.find_one({"id": aid, "quotation_id": qid}, {"_id": 0})
        if d:
            UPLOAD_ROOT = deps["UPLOAD_ROOT"]
            p = UPLOAD_ROOT / d.get("path", "")
            if p.exists(): p.unlink(missing_ok=True)
            await db.quotation_attachments.delete_one({"id": aid})
        return {"ok": True}

    # ---------- Versioning ----------
    @api.post("/quotations/{qid}/duplicate-version")
    async def duplicate_version(qid: str, current=Depends(require_internal)):
        q = await db.quotations.find_one({"id": qid}, {"_id": 0})
        if not q: raise HTTPException(404, "Quotation not found")
        newq = copy.deepcopy(q)
        new_id = gen_id()
        newq["id"] = new_id
        newq["version"] = int(q.get("version") or 1) + 1
        newq["parent_version_id"] = qid
        newq["status"] = "draft"
        newq["locked"] = False
        newq["selected"] = False
        newq["finalized"] = False
        newq["created_at"] = now_iso()
        newq["updated_at"] = now_iso()
        newq.pop("_id", None)
        await db.quotations.insert_one(newq)
        items = await db.quotation_items.find({"quotation_id": qid}, {"_id": 0}).to_list(2000)
        for it in items:
            it2 = {**it, "id": gen_id(), "quotation_id": new_id}
            it2.pop("_id", None)
            await db.quotation_items.insert_one(it2)
        await _log_activity(new_id, "duplicate_version", current.get("name"), {"from": qid})
        return {"id": new_id, "version": newq["version"]}

    @api.get("/quotations/{qid}/versions")
    async def versions(qid: str, current=Depends(require_internal)):
        q = await db.quotations.find_one({"id": qid}, {"_id": 0})
        if not q: raise HTTPException(404, "Quotation not found")
        # Trace back to root
        root_id = qid
        cur = q
        while cur.get("parent_version_id"):
            p = await db.quotations.find_one({"id": cur["parent_version_id"]}, {"_id": 0})
            if not p: break
            root_id = p["id"]; cur = p
        # Find all descendants
        chain = [await db.quotations.find_one({"id": root_id}, {"_id": 0})]
        # BFS
        frontier = [root_id]
        while frontier:
            nxt = []
            for pid in frontier:
                kids = await db.quotations.find({"parent_version_id": pid}, {"_id": 0}).to_list(50)
                for k in kids:
                    chain.append(k); nxt.append(k["id"])
            frontier = nxt
        chain.sort(key=lambda x: x.get("version") or 0)
        return chain

    # ---------- Workflow ----------
    @api.post("/quotations/{qid}/send-to-reviewer")
    async def send_reviewer(qid: str, payload: RemarkIn = Body(default=RemarkIn()), current=Depends(require_internal)):
        q = await db.quotations.find_one({"id": qid}, {"_id": 0})
        if not q: raise HTTPException(404, "Quotation not found")
        if q.get("status") not in ("draft", "returned"): raise HTTPException(400, "Only draft or returned quotations can be sent to reviewer")
        await db.quotations.update_one({"id": qid}, {"$set": {"status": "under_review", "updated_at": now_iso()}})
        await _log_activity(qid, "send_to_reviewer", current.get("name"), {"remarks": payload.remarks})
        return {"ok": True, "status": "under_review"}

    @api.post("/quotations/{qid}/send-to-vendor")
    async def send_vendor(qid: str, payload: RemarkIn = Body(default=RemarkIn()), current=Depends(require_internal)):
        q = await db.quotations.find_one({"id": qid}, {"_id": 0})
        if not q: raise HTTPException(404, "Quotation not found")
        await db.quotations.update_one({"id": qid}, {"$set": {"status": "requested", "updated_at": now_iso()}})
        # Auto-create quotation_request
        rfq_num = f"RFQ-{datetime.now(timezone.utc).year}-{gen_id()[:6].upper()}"
        rid = gen_id()
        await db.quotation_requests.insert_one({
            "id": rid, "rfq_number": rfq_num,
            "project_id": q.get("project_id"), "project_name": q.get("project_name"),
            "work_category": q.get("work_category"),
            "vendor_ids": [q.get("vendor_id")] if q.get("vendor_id") else [],
            "status_per_vendor": {q.get("vendor_id"): "pending"} if q.get("vendor_id") else {},
            "quotation_ids": [qid],
            "sent_at": now_iso(),
            "deadline": (datetime.now(timezone.utc) + timedelta(days=7)).isoformat(),
            "created_by": current.get("name"),
        })
        await _log_activity(qid, "send_to_vendor", current.get("name"), {"rfq_number": rfq_num})
        return {"ok": True, "status": "requested", "rfq_id": rid}

    @api.post("/quotations/{qid}/approve")
    async def approve(qid: str, payload: RemarkIn = Body(default=RemarkIn()), current=Depends(require_internal)):
        q = await db.quotations.find_one({"id": qid}, {"_id": 0})
        if not q: raise HTTPException(404, "Quotation not found")
        if q.get("status") not in ("under_review", "awaiting_approval", "submitted"): raise HTTPException(400, "Cannot approve in current status")
        await db.quotations.update_one({"id": qid}, {"$set": {"status": "approved", "locked": True, "approved_by": current.get("name"), "approved_at": now_iso(), "updated_at": now_iso()}})
        await _log_activity(qid, "approve", current.get("name"), {"remarks": payload.remarks})
        return {"ok": True, "status": "approved"}

    @api.post("/quotations/{qid}/return")
    async def return_quotation(qid: str, payload: RemarkIn = Body(default=RemarkIn()), current=Depends(require_internal)):
        q = await db.quotations.find_one({"id": qid}, {"_id": 0})
        if not q: raise HTTPException(404, "Quotation not found")
        await db.quotations.update_one({"id": qid}, {"$set": {"status": "returned", "updated_at": now_iso()}})
        await _log_activity(qid, "return", current.get("name"), {"remarks": payload.remarks})
        return {"ok": True, "status": "returned"}

    @api.post("/quotations/{qid}/reject")
    async def reject(qid: str, payload: RemarkIn = Body(default=RemarkIn()), current=Depends(require_internal)):
        q = await db.quotations.find_one({"id": qid}, {"_id": 0})
        if not q: raise HTTPException(404, "Quotation not found")
        await db.quotations.update_one({"id": qid}, {"$set": {"status": "rejected", "rejected_at": now_iso(), "updated_at": now_iso()}})
        await _log_activity(qid, "reject", current.get("name"), {"remarks": payload.remarks})
        return {"ok": True, "status": "rejected"}

    @api.post("/quotations/{qid}/request-clarification")
    async def request_clarification(qid: str, payload: RemarkIn = Body(default=RemarkIn()), current=Depends(require_internal)):
        q = await db.quotations.find_one({"id": qid}, {"_id": 0})
        if not q: raise HTTPException(404, "Quotation not found")
        await _log_activity(qid, "request_clarification", current.get("name"), {"remarks": payload.remarks})
        return {"ok": True}

    @api.post("/quotations/{qid}/mark-selected")
    async def mark_selected(qid: str, payload: RemarkIn = Body(default=RemarkIn()), current=Depends(require_internal)):
        q = await db.quotations.find_one({"id": qid}, {"_id": 0})
        if not q: raise HTTPException(404, "Quotation not found")
        # Un-select others in same project+category
        if q.get("project_id") and q.get("work_category"):
            await db.quotations.update_many(
                {"project_id": q["project_id"], "work_category": q["work_category"], "id": {"$ne": qid}, "selected": True},
                {"$set": {"selected": False, "finalized": False, "status": "not_selected", "updated_at": now_iso()}},
            )
        await db.quotations.update_one({"id": qid}, {"$set": {"status": "selected", "selected": True, "finalized": True, "locked": True, "selected_at": now_iso(), "updated_at": now_iso()}})
        await _log_activity(qid, "mark_selected", current.get("name"), {"remarks": payload.remarks})
        # Add document entry
        await db.documents.insert_one({
            "id": gen_id(), "name": f"Quotation {q.get('quotation_number')} — {q.get('vendor_name')} (Selected)",
            "type": "PDF", "category": "Quotation", "project_id": q.get("project_id"),
            "project_name": q.get("project_name"), "uploaded_by": current.get("name"), "uploaded_at": now_iso(), "source": "generated",
        })
        # Project activity
        if q.get("project_id"):
            await db.activity.insert_one({
                "id": gen_id(), "project_id": q.get("project_id"),
                "project_name": q.get("project_name"),
                "description": f"Vendor {q.get('vendor_name')} finalized for {q.get('work_category')}",
                "actor": current.get("name"), "actor_initials": (current.get("avatar_initials") or "?"),
                "status": "success", "at": now_iso(), "created_at": now_iso(),
            })
        return {"ok": True, "status": "selected"}

    @api.post("/quotations/{qid}/mark-not-selected")
    async def mark_not_selected(qid: str, payload: RemarkIn = Body(default=RemarkIn()), current=Depends(require_internal)):
        q = await db.quotations.find_one({"id": qid}, {"_id": 0})
        if not q: raise HTTPException(404, "Quotation not found")
        await db.quotations.update_one({"id": qid}, {"$set": {"status": "not_selected", "selected": False, "finalized": False, "updated_at": now_iso()}})
        await _log_activity(qid, "mark_not_selected", current.get("name"), {"remarks": payload.remarks})
        return {"ok": True, "status": "not_selected"}

    # ---------- Comparison ----------
    @api.get("/quotations-compare")
    async def compare(ids: str = Query(...), current=Depends(require_internal)):
        idlist = [i for i in ids.split(",") if i]
        if len(idlist) < 2: raise HTTPException(400, "Provide at least 2 quotation ids")
        quotes = []
        for qid in idlist:
            fq = await _full_quotation(qid)
            quotes.append(fq)
        # Build line-item matrix keyed by boq_item_id or description
        item_keys = {}
        for q in quotes:
            for it in q.get("items", []):
                key = it.get("boq_item_id") or it.get("description")
                item_keys.setdefault(key, {"description": it.get("description"), "unit": it.get("unit"), "boq_item_id": it.get("boq_item_id"), "boq_rate": (it.get("boq_ref_data") or {}).get("boq_rate"), "quotes": {}})
                item_keys[key]["quotes"][q["id"]] = {"rate": it.get("rate"), "amount": it.get("amount"), "quantity": it.get("quantity")}
        # Enrich vendor metrics
        for q in quotes:
            v = q.get("vendor") or {}
            q["vendor_metrics"] = {
                "rating": v.get("rating"), "completed_projects": v.get("completed_projects"),
                "on_time_pct": v.get("on_time_pct"), "similar_projects_completed": v.get("similar_projects_completed"),
                "availability_status": v.get("availability_status"), "current_assignments": v.get("current_assignments"),
                "verified": v.get("verified"), "preferred": v.get("preferred"),
            }
        lowest_id = None
        try:
            lowest_id = min(quotes, key=lambda q: q.get("subtotals", {}).get("total", 999999999))["id"]
        except Exception:
            pass
        return {"quotations": quotes, "line_items": list(item_keys.values()), "lowest_id": lowest_id}

    @api.post("/quotation-comparisons")
    async def save_comparison(payload: SaveComparisonIn, current=Depends(require_internal)):
        cid = gen_id()
        proj = await db.projects.find_one({"id": payload.project_id}, {"_id": 0}) if payload.project_id else None
        doc = {"id": cid, "name": payload.name, "project_id": payload.project_id,
               "project_name": proj.get("name") if proj else None,
               "work_category": payload.work_category, "quotation_ids": payload.quotation_ids,
               "saved_by": current.get("name"), "saved_at": now_iso()}
        await db.quotation_comparisons.insert_one(doc)
        doc.pop("_id", None)
        return doc

    @api.get("/quotation-comparisons/{cid}")
    async def get_comparison(cid: str, current=Depends(require_internal)):
        c = await db.quotation_comparisons.find_one({"id": cid}, {"_id": 0})
        if not c: raise HTTPException(404, "Comparison not found")
        return c

    # ---------- RFQ ----------
    @api.get("/quotation-requests")
    async def rfq_list(current=Depends(require_internal)):
        rows = await db.quotation_requests.find({}, {"_id": 0}).sort("sent_at", -1).to_list(200)
        # Enrich vendor names
        for r in rows:
            names = []
            for vid in r.get("vendor_ids", []):
                v = await db.vendors.find_one({"id": vid}, {"_id": 0, "company": 1, "name": 1})
                if v: names.append({"id": vid, "name": v.get("company") or v.get("name"), "status": r.get("status_per_vendor", {}).get(vid, "pending")})
            r["vendors"] = names
        return rows

    @api.post("/quotation-requests")
    async def rfq_create(payload: RFQIn, current=Depends(require_internal)):
        rid = gen_id()
        rfq_num = f"RFQ-{datetime.now(timezone.utc).year}-{rid[:6].upper()}"
        p = await db.projects.find_one({"id": payload.project_id}, {"_id": 0}) if payload.project_id else None
        doc = {
            "id": rid, "rfq_number": rfq_num,
            "project_id": payload.project_id, "project_name": p.get("name") if p else None,
            "work_category": payload.work_category,
            "vendor_ids": payload.vendor_ids,
            "status_per_vendor": {vid: "pending" for vid in payload.vendor_ids},
            "quotation_ids": [],
            "sent_at": now_iso(),
            "deadline": payload.deadline or (datetime.now(timezone.utc) + timedelta(days=7)).isoformat(),
            "created_by": current.get("name"),
        }
        await db.quotation_requests.insert_one(doc)
        return {"id": rid, "rfq_number": rfq_num}

    @api.post("/quotation-requests/{rid}/remind")
    async def rfq_remind(rid: str, current=Depends(require_internal)):
        r = await db.quotation_requests.find_one({"id": rid}, {"_id": 0})
        if not r: raise HTTPException(404, "RFQ not found")
        await db.quotation_activity.insert_one({"id": gen_id(), "quotation_id": None, "rfq_id": rid, "action": "reminder_sent", "actor": current.get("name"), "at": now_iso(), "meta": {}})
        return {"ok": True, "message": "Reminder logged (email sending is stubbed)"}

    # ---------- Export ----------
    @api.get("/quotations/{qid}/export/excel")
    async def export_excel(qid: str, current=Depends(require_internal)):
        q = await _full_quotation(qid)
        wb = Workbook()
        ws = wb.active
        ws.title = "Quotation"
        thin = Side(border_style="thin", color="DDD8CE")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="1F453B")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        ws["A1"] = f"INOS — QUOTATION {q.get('quotation_number')}"
        ws["A1"].font = Font(bold=True, size=16, color="EF7F1B")
        ws.merge_cells("A1:H1")
        ws["A3"] = "Title:"; ws["B3"] = q.get("title")
        ws["A4"] = "Project:"; ws["B4"] = q.get("project_name") or ""
        ws["A5"] = "Vendor:"; ws["B5"] = q.get("vendor_name") or ""
        ws["A6"] = "Category:"; ws["B6"] = q.get("work_category") or ""
        ws["A7"] = "Date:"; ws["B7"] = q.get("quotation_date") or ""
        ws["A8"] = "Valid Until:"; ws["B8"] = q.get("valid_until") or ""
        ws["A9"] = "Status:"; ws["B9"] = q.get("status")
        row = 11
        headers = ["S.No", "Particulars", "BOQ Ref", "Unit", "Qty", "Rate (₹)", "Amount (₹)", "Tax %"]
        for j, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=j, value=h)
            c.fill = header_fill; c.font = header_font; c.border = border
        row += 1
        for i, it in enumerate(q.get("items", []), 1):
            ws.cell(row=row, column=1, value=i).border = border
            ws.cell(row=row, column=2, value=it.get("description")).border = border
            ws.cell(row=row, column=3, value=it.get("boq_ref") or "").border = border
            ws.cell(row=row, column=4, value=it.get("unit")).border = border
            ws.cell(row=row, column=5, value=float(it.get("quantity") or 0)).border = border
            ws.cell(row=row, column=6, value=float(it.get("rate") or 0)).border = border
            ws.cell(row=row, column=7, value=float(it.get("amount") or 0)).border = border
            ws.cell(row=row, column=8, value=float(it.get("tax_pct") or 0)).border = border
            row += 1
        row += 1
        sub = q.get("subtotals") or {}
        for label, val in [("Base", sub.get("base", 0)), ("Tax", sub.get("tax", 0)), ("Transport", sub.get("transport", 0)), ("Installation", sub.get("installation", 0)), ("Additional", sub.get("additional", 0)), ("Discount", -sub.get("discount", 0)), ("TOTAL", sub.get("total", 0))]:
            ws.cell(row=row, column=6, value=label).font = Font(bold=(label == "TOTAL"))
            ws.cell(row=row, column=7, value=float(val or 0)).font = Font(bold=(label == "TOTAL"), color="EF7F1B" if label == "TOTAL" else "252525")
            row += 1
        ws.column_dimensions["B"].width = 44
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        fn = f"Quotation_{q.get('quotation_number')}.xlsx"
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{fn}"'})

    @api.post("/quotations/{qid}/export/pdf")
    async def export_pdf(qid: str, current=Depends(require_internal)):
        q = await _full_quotation(qid)
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=14*mm, rightMargin=14*mm, topMargin=16*mm, bottomMargin=16*mm, compress=0)
        styles = getSampleStyleSheet()
        orange = colors.HexColor("#EF7F1B"); green = colors.HexColor("#1F453B"); border_c = colors.HexColor("#DDD8CE"); muted = colors.HexColor("#6B7280")
        story = []
        # Logo / Wordmark
        header_tbl = Table([[Paragraph('<font color="#EF7F1B" size="20"><b>INOS</b></font><br/><font color="#6B7280" size="7">ERP · Quotation</font>', styles["Normal"]),
                             Paragraph(f'<para align="right"><font color="#1F453B" size="10"><b>QUOTATION</b></font><br/><font size="14" color="#EF7F1B"><b>{q.get("quotation_number")}</b></font><br/><font size="8" color="#6B7280">Version {q.get("version")} · {q.get("status","").upper()}</font></para>', styles["Normal"])]],
                            colWidths=[95*mm, 85*mm])
        header_tbl.setStyle(TableStyle([("LINEBELOW",(0,0),(-1,-1),1.5,orange), ("BOTTOMPADDING",(0,0),(-1,-1),6)]))
        story.append(header_tbl); story.append(Spacer(1, 6))

        title_st = ParagraphStyle('t', parent=styles['Title'], textColor=colors.HexColor("#202735"), fontSize=15, spaceAfter=4, alignment=0)
        story.append(Paragraph(f"{q.get('title') or 'Quotation'}", title_st))
        story.append(Paragraph(f'<font color="#6B7280" size="9">{q.get("work_category") or ""} · {q.get("quotation_date") or ""} · Valid until {q.get("valid_until") or ""}</font>', styles["Normal"]))
        story.append(Spacer(1, 10))

        vendor = q.get("vendor") or {}
        project = q.get("project") or {}
        # Vendor + Project blocks
        def _block(label, lines):
            para = f'<font color="#9A9388" size="8"><b>{label}</b></font><br/>'
            para += "<br/>".join(f'<font size="9">{l}</font>' for l in lines if l)
            return Paragraph(para, styles["Normal"])
        vendor_block = _block("VENDOR", [
            f"<b>{vendor.get('company') or vendor.get('name') or q.get('vendor_name') or '—'}</b>",
            vendor.get("primary_contact") or vendor.get("name") or "",
            vendor.get("phone") or "",
            vendor.get("email") or "",
            f"GST: {vendor.get('gst_number') or '—'}",
        ])
        project_block = _block("PROJECT", [
            f"<b>{project.get('name') or q.get('project_name') or '—'}</b>",
            project.get("client_name") or "",
            project.get("location") or "",
            f"Category: {q.get('work_category') or '—'}",
            f"Currency: {q.get('currency') or 'INR'} · Tax: {q.get('tax_config') or 'GST 18%'}",
        ])
        blocks_tbl = Table([[vendor_block, project_block]], colWidths=[90*mm, 90*mm])
        blocks_tbl.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#FAF8F5")),("BOX",(0,0),(-1,-1),0.4,border_c),("LEFTPADDING",(0,0),(-1,-1),8),("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8)]))
        story.append(blocks_tbl); story.append(Spacer(1, 10))

        # Items table
        items = q.get("items", [])
        rows = [["S.No", "Particulars", "Unit", "Qty", "Rate", "Amount", "Tax %"]]
        for i, it in enumerate(items, 1):
            rows.append([str(i), Paragraph(it.get("description") or "", styles["BodyText"]), it.get("unit") or "", str(it.get("quantity") or 0), fmt_inr(it.get("rate") or 0), fmt_inr(it.get("amount") or 0), f"{it.get('tax_pct') or 0}%"])
        items_t = Table(rows, colWidths=[12*mm, 72*mm, 15*mm, 18*mm, 24*mm, 26*mm, 15*mm], repeatRows=1)
        items_t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), green),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("FONT", (0,0), (-1,0), "Helvetica-Bold", 9),
            ("GRID", (0,0), (-1,-1), 0.3, border_c),
            ("FONT", (0,1), (-1,-1), "Helvetica", 8.5),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#FAF8F5")]),
            ("ALIGN", (3,1), (6,-1), "RIGHT"),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
        ]))
        story.append(items_t); story.append(Spacer(1, 10))

        # Cost summary
        sub = q.get("subtotals") or {}
        summary_rows = [
            ["Base", fmt_inr(sub.get("base", 0))],
            ["Tax", fmt_inr(sub.get("tax", 0))],
            ["Transportation", fmt_inr(sub.get("transport", 0))],
            ["Installation", fmt_inr(sub.get("installation", 0))],
            ["Additional Charges", fmt_inr(sub.get("additional", 0))],
            ["Discount", "-" + fmt_inr(sub.get("discount", 0))],
            ["QUOTATION TOTAL", fmt_inr(sub.get("total", 0))],
        ]
        st = Table(summary_rows, colWidths=[110*mm, 40*mm], hAlign="RIGHT")
        st.setStyle(TableStyle([
            ("GRID", (0,0), (-1,-1), 0.3, border_c),
            ("BACKGROUND", (0,-1), (-1,-1), colors.HexColor("#FEF2E6")),
            ("FONT", (0,-1), (-1,-1), "Helvetica-Bold", 11),
            ("TEXTCOLOR", (0,-1), (-1,-1), orange),
            ("ALIGN", (1,0), (1,-1), "RIGHT"),
            ("FONT", (0,0), (0,-2), "Helvetica", 9),
        ]))
        story.append(st); story.append(Spacer(1, 12))

        # Commercial terms
        ct = q.get("commercial_terms") or {}
        story.append(Paragraph('<font color="#1F453B" size="11"><b>Commercial Terms</b></font>', styles["Normal"]))
        story.append(Spacer(1, 4))
        terms_rows = [
            ["Payment Terms", ct.get("payment_terms") or "—"],
            ["Advance %", str(ct.get("advance_pct") or 0)],
            ["Credit Period", f"{ct.get('credit_period_days') or 0} days"],
            ["Delivery Timeline", f"{ct.get('delivery_timeline_days') or 0} days"],
            ["Completion Timeline", f"{ct.get('completion_timeline_days') or 0} days"],
            ["Warranty", f"{ct.get('warranty_months') or 0} months — {ct.get('warranty_notes') or ''}"],
            ["Transportation", "Included" if ct.get("transportation_included") else f"₹{ct.get('transportation_amount') or 0}"],
            ["Installation", "Included" if ct.get("installation_included") else f"₹{ct.get('installation_amount') or 0}"],
        ]
        tt = Table(terms_rows, colWidths=[50*mm, 130*mm])
        tt.setStyle(TableStyle([("GRID",(0,0),(-1,-1),0.3,border_c),("BACKGROUND",(0,0),(0,-1),colors.HexColor("#F3F3F1")),("FONT",(0,0),(-1,-1),"Helvetica",9),("VALIGN",(0,0),(-1,-1),"TOP"),("LEFTPADDING",(0,0),(-1,-1),6),("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4)]))
        story.append(tt); story.append(Spacer(1, 8))

        # Exclusions
        exclusions = ct.get("exclusions") or []
        if exclusions:
            story.append(Paragraph('<font color="#1F453B" size="10"><b>Exclusions</b></font>', styles["Normal"]))
            story.append(Paragraph(", ".join(exclusions), styles["Normal"]))
            story.append(Spacer(1, 6))
        if ct.get("special_conditions"):
            story.append(Paragraph('<font color="#1F453B" size="10"><b>Special Conditions</b></font>', styles["Normal"]))
            story.append(Paragraph(ct.get("special_conditions"), styles["Normal"]))
            story.append(Spacer(1, 6))

        # Terms & Conditions boilerplate — always renders
        story.append(Spacer(1, 8))
        story.append(Paragraph('<font color="#1F453B" size="10"><b>Standard Terms &amp; Conditions</b></font>', styles["Normal"]))
        tc_text = (
            '<font size="8" color="#4B5158">'
            '1. This quotation is valid until the date specified above and is subject to revision if not accepted within the validity period.<br/>'
            '2. All rates are in Indian Rupees (INR) unless otherwise stated. Applicable taxes as per prevailing GST rates in India.<br/>'
            '3. Payment terms: Advance payment as specified above. Balance payments as per milestones defined. All payments to be made via bank transfer (RTGS/NEFT/UPI) in favour of the firm.<br/>'
            '4. Any additional work outside the scope of this quotation will be charged separately at mutually agreed rates via a written change order.<br/>'
            '5. The vendor shall be responsible for all statutory approvals, licenses, and clearances required for the work under this quotation.<br/>'
            '6. Site conditions and access should be as per the site visit conducted. Any material changes may impact the quoted rates and timelines.<br/>'
            '7. Warranty period commences from the date of successful handover and installation completion signed off by INOS project manager.<br/>'
            '8. In case of any dispute, jurisdiction shall be in New Delhi, India. Both parties agree to first attempt mutual resolution before legal recourse.<br/>'
            '9. Force majeure clauses apply for events beyond reasonable control including but not limited to natural calamities, government restrictions, and pandemics.<br/>'
            '10. Confidentiality: All drawings, specifications, and commercial terms shared under this quotation are confidential to INOS and the client.<br/>'
            '</font>'
        )
        story.append(Paragraph(tc_text, styles["Normal"]))
        story.append(Spacer(1, 6))

        # Notes block
        story.append(Paragraph('<font color="#1F453B" size="10"><b>Notes to Vendor</b></font>', styles["Normal"]))
        notes_text = (
            '<font size="8" color="#4B5158">'
            'Please ensure all materials are as per approved samples and specifications. Deliveries should be scheduled 48 hours in advance with the site supervisor. '
            'GST invoices must be raised in the name of the firm with correct GSTIN. Delayed deliveries beyond the agreed timeline may attract liquidated damages at 0.5% per week subject to a maximum of 5% of the quotation value. '
            'Site protection, dust barriers, and safety measures during execution are the responsibility of the vendor. '
            'Any deviation from approved drawings must be pre-approved in writing by the INOS project manager.'
            '</font>'
        )
        story.append(Paragraph(notes_text, styles["Normal"]))

        # Signature area
        story.append(Spacer(1, 14))
        sig = Table([[Paragraph('<font size="8" color="#6B7280">For INOS</font><br/><br/><br/>___________________________<br/><font size="8">Authorized Signatory</font>', styles["Normal"]),
                       Paragraph(f'<font size="8" color="#6B7280">For {vendor.get("company") or q.get("vendor_name") or "Vendor"}</font><br/><br/><br/>___________________________<br/><font size="8">Vendor Signature & Stamp</font>', styles["Normal"])]],
                     colWidths=[90*mm, 90*mm])
        sig.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),("LINEABOVE",(0,0),(-1,0),0.3,border_c),("TOPPADDING",(0,0),(-1,-1),20)]))
        story.append(sig)

        # Footer with page number
        def _footer(canvas, d):
            canvas.saveState()
            canvas.setFont("Helvetica", 7)
            canvas.setFillColor(muted)
            canvas.drawString(14*mm, 8*mm, f"INOS ERP · Quotation {q.get('quotation_number')} · Prepared by {q.get('prepared_by') or '—'}")
            canvas.drawRightString(196*mm, 8*mm, f"Page {d.page}")
            canvas.restoreState()

        doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
        buf.seek(0)
        fn = f"Quotation_{q.get('quotation_number')}.pdf"
        return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{fn}"'})

    @api.get("/quotations-export-comparison-pdf")
    async def export_comparison_pdf(ids: str = Query(...), current=Depends(require_internal)):
        idlist = [i for i in ids.split(",") if i]
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=12*mm, rightMargin=12*mm, topMargin=14*mm, bottomMargin=14*mm)
        styles = getSampleStyleSheet()
        orange = colors.HexColor("#EF7F1B"); green = colors.HexColor("#1F453B"); border_c = colors.HexColor("#DDD8CE")
        story = [Paragraph("INOS — QUOTATION COMPARISON", ParagraphStyle('t', parent=styles['Title'], textColor=orange, fontSize=18))]
        quotes = []
        for qid in idlist:
            q = await db.quotations.find_one({"id": qid}, {"_id": 0})
            if q: quotes.append(q)
        header = ["Attribute"] + [q.get("vendor_name") or "" for q in quotes]
        attrs = [
            ("Quotation #", [q.get("quotation_number") for q in quotes]),
            ("Version", [str(q.get("version")) for q in quotes]),
            ("Status", [q.get("status") for q in quotes]),
            ("Date", [q.get("quotation_date") for q in quotes]),
            ("Valid Until", [q.get("valid_until") for q in quotes]),
            ("Base", [fmt_inr((q.get("subtotals") or {}).get("base", 0)) for q in quotes]),
            ("Tax", [fmt_inr((q.get("subtotals") or {}).get("tax", 0)) for q in quotes]),
            ("TOTAL", [fmt_inr((q.get("subtotals") or {}).get("total", 0)) for q in quotes]),
            ("BOQ Variation", [f"{q.get('boq_variation_pct')}%" if q.get('boq_variation_pct') is not None else "—" for q in quotes]),
        ]
        rows = [header] + [[a] + v for a, v in attrs]
        t = Table(rows, repeatRows=1)
        t.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,0), green), ("TEXTCOLOR", (0,0), (-1,0), colors.white), ("GRID", (0,0), (-1,-1), 0.3, border_c), ("FONT", (0,0), (-1,-1), "Helvetica", 9)]))
        story.append(t); doc.build(story); buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": 'attachment; filename="quotation_comparison.pdf"'})


async def seed_phase4(db, gen_id, now_iso):
    """Idempotent Phase 4 seed."""
    if await db.quotation_comparisons.count_documents({}) > 0:
        return
    await db.quotations.delete_many({})  # remove Phase 3 stub quotations
    await db.quotation_items.delete_many({})
    await db.quotation_attachments.delete_many({})
    await db.quotation_activity.delete_many({})
    await db.quotation_requests.delete_many({})
    await db.quotation_comparisons.delete_many({})

    projects = await db.projects.find({}, {"_id": 0}).to_list(50)
    if not projects: return
    proj_by_name = {p["name"]: p for p in projects}
    vendors = await db.vendors.find({}, {"_id": 0}).to_list(50)
    if not vendors: return
    boqs = await db.boqs.find({}, {"_id": 0}).to_list(50)
    admin = await db.users.find_one({"email": "admin@buildcon.in"}, {"_id": 0})
    admin_name = admin.get("name") if admin else "Deepak Rao"

    kohli = proj_by_name.get("Kohli Residence — Interior Renovation")
    house = proj_by_name.get("The House Within")
    studio = proj_by_name.get("Studio Office")
    jain = proj_by_name.get("Jain Art Press")
    r24 = proj_by_name.get("Residence 24")
    bansal = proj_by_name.get("Bansal Villa")

    kohli_boq = next((b for b in boqs if kohli and b.get("project_id") == kohli.get("id")), None)
    kohli_boq_items = []
    if kohli_boq:
        kohli_boq_items = await db.boq_items.find({"boq_id": kohli_boq["id"]}, {"_id": 0}).limit(20).to_list(20)

    def v_by_cat(cat, price=None, verified=None):
        vs = [v for v in vendors if v.get("primary_category") == cat]
        if price:
            vs = [v for v in vs if v.get("price_range") == price]
        if verified is not None:
            vs = [v for v in vs if v.get("verified") == verified]
        return vs

    def _calc_amt(qty, rate, calc):
        if calc == "L": return float(rate)
        return round(float(qty) * float(rate), 2)

    counter = [1]
    def next_num():
        n = f"QT-2026-{counter[0]:04d}"
        counter[0] += 1
        return n

    async def create_q(project, vendor, category, status, items_data, days_valid=30, use_boq=None, source="created", version=1):
        qid = gen_id()
        num = next_num()
        base = 0.0; tax = 0.0
        item_docs = []
        boq_ref_qamt = 0.0; boq_ref_bamt = 0.0
        for i, it in enumerate(items_data):
            qty = it.get("quantity", 1)
            rate = it.get("rate", 0)
            calc = it.get("calc_type", "M")
            amt = _calc_amt(qty, rate, calc)
            tp = it.get("tax_pct", 18)
            base += amt; tax += amt * tp / 100
            boq_iid = it.get("boq_item_id")
            if boq_iid:
                bi = await db.boq_items.find_one({"id": boq_iid}, {"_id": 0})
                if bi:
                    boq_ref_bamt += float(bi.get("amount") or 0)
                    boq_ref_qamt += amt
            item_docs.append({
                "id": gen_id(), "quotation_id": qid, "order": i,
                "description": it["description"], "unit": it.get("unit", "Nos."),
                "quantity": qty, "rate": rate, "amount": amt, "tax_pct": tp,
                "calc_type": calc, "boq_item_id": boq_iid, "boq_ref": it.get("boq_ref"),
                "location": it.get("location"), "remarks": it.get("remarks", ""),
                "created_at": now_iso(),
            })
        additional = 0.0
        transport = it.get("transport", 0) if items_data else 0
        install = 0.0
        discount = 0.0
        total = round(base + tax + transport + install + additional - discount, 2)
        variation = round((boq_ref_qamt - boq_ref_bamt) / boq_ref_bamt * 100.0, 2) if boq_ref_bamt > 0 else None
        today = datetime.now(timezone.utc).date()
        vu = (today + timedelta(days=days_valid)).isoformat()
        if status == "expired":
            vu = (today - timedelta(days=5)).isoformat()
        doc = {
            "id": qid, "quotation_number": num, "title": f"{category} — {project['name']}",
            "project_id": project["id"], "project_name": project["name"],
            "vendor_id": vendor["id"], "vendor_name": vendor.get("company") or vendor.get("name"),
            "work_category": category, "source": source, "status": status,
            "version": version, "parent_version_id": None,
            "currency": "INR", "quotation_date": today.isoformat(), "valid_until": vu,
            "prepared_by": admin_name, "reviewer_id": None,
            "tax_config": "GST 18%",
            "commercial_terms": {"payment_terms": "30-40-20-10", "advance_pct": 30, "credit_period_days": 15, "transportation_amount": transport, "transportation_included": transport == 0, "installation_amount": 0, "installation_included": True, "warranty_months": 12, "warranty_notes": "Manufacturer warranty", "delivery_timeline_days": 15, "completion_timeline_days": 45, "exclusions": ["Civil work", "Client-supplied material"], "special_conditions": ""},
            "additional_charges": [], "discount": 0,
            "subtotals": {"base": round(base, 2), "tax": round(tax, 2), "transport": transport, "installation": 0, "additional": 0, "discount": 0, "total": total},
            "boq_variation_pct": variation,
            "finalized": status == "selected", "selected": status == "selected", "locked": status in ("approved", "selected", "archived"),
            "quotation_type": "Vendor", "created_by": admin_name,
            "created_at": now_iso(), "updated_at": now_iso(),
        }
        if status == "approved" or status == "selected":
            doc["approved_by"] = admin_name; doc["approved_at"] = now_iso()
        if status == "selected":
            doc["selected_at"] = now_iso()
        if status == "rejected":
            doc["rejected_at"] = now_iso()
        await db.quotations.insert_one(doc)
        for it in item_docs: await db.quotation_items.insert_one(it)
        await db.quotation_activity.insert_one({"id": gen_id(), "quotation_id": qid, "action": "create", "actor": admin_name, "at": now_iso(), "meta": {}})
        if status in ("under_review", "awaiting_approval", "approved", "selected", "rejected", "returned"):
            await db.quotation_activity.insert_one({"id": gen_id(), "quotation_id": qid, "action": "send_to_reviewer", "actor": admin_name, "at": now_iso(), "meta": {}})
        if status in ("approved", "selected"):
            await db.quotation_activity.insert_one({"id": gen_id(), "quotation_id": qid, "action": "approve", "actor": admin_name, "at": now_iso(), "meta": {"remarks": "Approved after review"}})
        if status == "selected":
            await db.quotation_activity.insert_one({"id": gen_id(), "quotation_id": qid, "action": "mark_selected", "actor": admin_name, "at": now_iso(), "meta": {}})
        if status == "rejected":
            await db.quotation_activity.insert_one({"id": gen_id(), "quotation_id": qid, "action": "reject", "actor": admin_name, "at": now_iso(), "meta": {"remarks": "Above budget"}})
        return qid

    # Vendor pools by category
    flooring = v_by_cat("Flooring") or vendors[:3]
    electrical = v_by_cat("Electrical") or vendors[3:6]
    painting = v_by_cat("Painting") or vendors[6:8]
    furniture = v_by_cat("Furniture & Joinery") or vendors[8:12]
    plumbing = v_by_cat("Plumbing") or vendors[12:14]
    civil = v_by_cat("Civil & Masonry") or vendors[:2]
    kitchen = v_by_cat("Kitchen") or vendors[:1]
    hvac = v_by_cat("HVAC") or vendors[:1]

    # Reusable BOQ-linked items sample
    def sample_boq_items(n):
        if not kohli_boq_items: return []
        out = []
        for bi in kohli_boq_items[:n]:
            out.append({"description": bi.get("description"), "unit": bi.get("unit"), "quantity": float(bi.get("quantity") or 1), "rate": float(bi.get("rate") or 0) * 1.08, "boq_item_id": bi.get("id"), "boq_ref": bi.get("code"), "calc_type": bi.get("calc_type") or "M", "tax_pct": 18})
        return out

    def sample_boq_items_priced(n, price_mult):
        base = sample_boq_items(n)
        for b in base: b["rate"] = round(b["rate"] * price_mult / 1.08, 2)
        return base

    ids = {"drafts": [], "requested": [], "received": [], "under_review": [], "awaiting_approval": [], "approved": [], "selected": [], "rejected": [], "expired": []}

    # ---- 4 Drafts ----
    ids["drafts"].append(await create_q(r24, furniture[0], "Furniture & Joinery", "draft", [
        {"description": "Modular wardrobe — 6ft laminate", "unit": "Sq.ft.", "quantity": 96, "rate": 1450, "tax_pct": 18},
        {"description": "TV unit with backlit panel", "unit": "Rft.", "quantity": 12, "rate": 2200, "tax_pct": 18},
        {"description": "Study table with drawer", "unit": "Nos.", "quantity": 2, "rate": 32000, "tax_pct": 18},
    ]))
    ids["drafts"].append(await create_q(house, electrical[0], "Electrical", "draft", [
        {"description": "Concealed wiring point", "unit": "Nos.", "quantity": 45, "rate": 950, "tax_pct": 18},
        {"description": "LED panel light — 15W", "unit": "Nos.", "quantity": 22, "rate": 2200, "tax_pct": 18},
    ]))
    ids["drafts"].append(await create_q(bansal, painting[0], "Painting", "draft", [
        {"description": "Interior emulsion — 2 coats", "unit": "Sq.ft.", "quantity": 3800, "rate": 32, "tax_pct": 18},
        {"description": "Texture finish — designer", "unit": "Sq.ft.", "quantity": 350, "rate": 185, "tax_pct": 18},
    ]))
    ids["drafts"].append(await create_q(jain, civil[0] if civil else vendors[0], "Civil & Masonry", "draft", [
        {"description": "Brick wall — 4.5\"", "unit": "Sq.ft.", "quantity": 1200, "rate": 95, "tax_pct": 18},
    ]))

    # ---- 3 Requested (RFQ sent, no response yet) ----
    ids["requested"].append(await create_q(studio, hvac[0] if hvac else vendors[1], "HVAC", "requested", [
        {"description": "VRF outdoor unit — 8HP", "unit": "Nos.", "quantity": 2, "rate": 285000, "tax_pct": 18},
        {"description": "Indoor cassette 2 ton", "unit": "Nos.", "quantity": 6, "rate": 68000, "tax_pct": 18},
    ]))
    ids["requested"].append(await create_q(house, plumbing[0] if plumbing else vendors[2], "Plumbing", "requested", [
        {"description": "CPVC water supply line", "unit": "Rft.", "quantity": 180, "rate": 220, "tax_pct": 18},
        {"description": "Concealed plumbing point", "unit": "Point", "quantity": 14, "rate": 1850, "tax_pct": 18},
    ]))
    ids["requested"].append(await create_q(r24, kitchen[0] if kitchen else vendors[3], "Kitchen", "requested", [
        {"description": "Modular kitchen — L-shape, acrylic finish", "unit": "Lump", "quantity": 1, "rate": 485000, "calc_type": "L", "tax_pct": 18},
    ]))

    # ---- 3 Received (vendor responded, awaiting internal review) ----
    ids["received"].append(await create_q(kohli, flooring[1] if len(flooring)>1 else flooring[0], "Flooring", "received",
        sample_boq_items_priced(4, 1.05) or [{"description": "Vitrified tile 800x800", "unit": "Sq.ft.", "quantity": 1600, "rate": 145, "tax_pct": 18}]))
    ids["received"].append(await create_q(studio, furniture[1] if len(furniture)>1 else furniture[0], "Furniture & Joinery", "received", [
        {"description": "Workstation — 4 seater", "unit": "Nos.", "quantity": 12, "rate": 42000, "tax_pct": 18},
        {"description": "Executive cabin table", "unit": "Nos.", "quantity": 4, "rate": 68000, "tax_pct": 18},
    ]))
    ids["received"].append(await create_q(house, electrical[1] if len(electrical)>1 else electrical[0], "Electrical", "received", [
        {"description": "Concealed wiring point", "unit": "Nos.", "quantity": 45, "rate": 1050, "tax_pct": 18},
        {"description": "LED panel light — 15W", "unit": "Nos.", "quantity": 22, "rate": 2400, "tax_pct": 18},
        {"description": "MCB Distribution Board — 8 way", "unit": "Nos.", "quantity": 2, "rate": 7200, "tax_pct": 18},
    ]))

    # ---- 3 Under Review ----
    ids["under_review"].append(await create_q(kohli, flooring[0], "Flooring", "under_review",
        sample_boq_items_priced(4, 1.12) or [{"description": "Marble slab — Italian", "unit": "Sq.ft.", "quantity": 800, "rate": 480, "tax_pct": 18}]))
    ids["under_review"].append(await create_q(jain, painting[0], "Painting", "under_review", [
        {"description": "Interior emulsion — 2 coats", "unit": "Sq.ft.", "quantity": 4200, "rate": 34, "tax_pct": 18},
        {"description": "Enamel paint on metal", "unit": "Sq.ft.", "quantity": 260, "rate": 68, "tax_pct": 18},
    ]))
    ids["under_review"].append(await create_q(studio, electrical[0], "Electrical", "under_review", [
        {"description": "Cable tray — 300mm", "unit": "Rft.", "quantity": 240, "rate": 480, "tax_pct": 18},
        {"description": "LT panel — 800A", "unit": "Nos.", "quantity": 1, "rate": 385000, "tax_pct": 18},
    ]))

    # ---- 2 Awaiting Approval ----
    ids["awaiting_approval"].append(await create_q(bansal, civil[0], "Civil & Masonry", "awaiting_approval", [
        {"description": "RCC roof slab casting", "unit": "Sq.ft.", "quantity": 1800, "rate": 220, "tax_pct": 18},
    ]))
    ids["awaiting_approval"].append(await create_q(house, furniture[0], "Furniture & Joinery", "awaiting_approval", [
        {"description": "Full wardrobe with sliding doors", "unit": "Sq.ft.", "quantity": 140, "rate": 1650, "tax_pct": 18},
    ]))

    # ---- 3 Approved (1 with selected=true) ----
    ids["approved"].append(await create_q(studio, hvac[0] if hvac else vendors[0], "HVAC", "approved", [
        {"description": "VRF system — turnkey", "unit": "Lump", "quantity": 1, "rate": 2200000, "calc_type": "L", "tax_pct": 18},
    ]))
    ids["approved"].append(await create_q(bansal, painting[0], "Painting", "approved", [
        {"description": "Full interior painting", "unit": "Lump", "quantity": 1, "rate": 285000, "calc_type": "L", "tax_pct": 18},
    ]))
    kohli_selected_qid = await create_q(kohli, flooring[2] if len(flooring)>2 else flooring[0], "Flooring", "selected",
        sample_boq_items_priced(4, 0.98) or [{"description": "Vitrified tile 600x600", "unit": "Sq.ft.", "quantity": 1600, "rate": 128, "tax_pct": 18}])
    ids["selected"].append(kohli_selected_qid)

    # ---- 1 Rejected ----
    ids["rejected"].append(await create_q(kohli, flooring[0] if flooring else vendors[0], "Flooring", "rejected",
        sample_boq_items_priced(4, 1.28) or [{"description": "Premium marble", "unit": "Sq.ft.", "quantity": 800, "rate": 640, "tax_pct": 18}]))

    # ---- 1 Expired ----
    ids["expired"].append(await create_q(jain, plumbing[0] if plumbing else vendors[0], "Plumbing", "expired", [
        {"description": "Concealed plumbing point", "unit": "Point", "quantity": 8, "rate": 1750, "tax_pct": 18},
    ]))

    # ---- Comparisons ----
    kohli_flooring_ids = ids["under_review"][:1] + ids["received"][:1] + [kohli_selected_qid] + ids["rejected"][:1]
    kohli_flooring_ids = [q for q in kohli_flooring_ids if q]
    if kohli_flooring_ids:
        await db.quotation_comparisons.insert_one({
            "id": gen_id(), "name": "Kohli Residence — Flooring Package Comparison",
            "project_id": kohli["id"] if kohli else None, "project_name": kohli["name"] if kohli else None,
            "work_category": "Flooring", "quotation_ids": kohli_flooring_ids[:3],
            "saved_by": admin_name, "saved_at": now_iso(),
        })
    house_elec_ids = [ids["drafts"][1], ids["received"][2]] if len(ids["received"]) >= 3 else ids["drafts"][:1]
    await db.quotation_comparisons.insert_one({
        "id": gen_id(), "name": "The House Within — Electrical Comparison",
        "project_id": house["id"] if house else None, "project_name": house["name"] if house else None,
        "work_category": "Electrical", "quotation_ids": house_elec_ids,
        "saved_by": admin_name, "saved_at": now_iso(),
    })
    studio_furn_ids = [ids["received"][1], ids["under_review"][2] if len(ids["under_review"]) >= 3 else ids["drafts"][0]]
    await db.quotation_comparisons.insert_one({
        "id": gen_id(), "name": "Studio Office — Furniture Comparison",
        "project_id": studio["id"] if studio else None, "project_name": studio["name"] if studio else None,
        "work_category": "Furniture & Joinery", "quotation_ids": studio_furn_ids,
        "saved_by": admin_name, "saved_at": now_iso(),
    })

    # ---- RFQs ----
    for i, (proj, cat, vids) in enumerate([
        (house, "Plumbing", [v["id"] for v in plumbing[:3]] or [vendors[0]["id"]]),
        (studio, "HVAC", [v["id"] for v in hvac[:2]] or [vendors[0]["id"]]),
        (r24, "Kitchen", [v["id"] for v in kitchen[:2]] or [vendors[0]["id"]]),
        (jain, "False Ceiling", [v["id"] for v in v_by_cat("False Ceiling")[:2]] or [vendors[0]["id"]]),
    ]):
        rid = gen_id()
        await db.quotation_requests.insert_one({
            "id": rid, "rfq_number": f"RFQ-2026-{i+1:04d}",
            "project_id": proj["id"] if proj else None, "project_name": proj["name"] if proj else None,
            "work_category": cat, "vendor_ids": vids,
            "status_per_vendor": {vid: ["pending", "received", "declined"][j % 3] for j, vid in enumerate(vids)},
            "quotation_ids": [],
            "sent_at": now_iso(),
            "deadline": (datetime.now(timezone.utc) + timedelta(days=(i+3))).isoformat(),
            "created_by": admin_name,
        })

    logger.info(f"Phase 4 seed complete: {sum(len(v) for v in ids.values())} quotations, 3 comparisons, 4 RFQs.")
