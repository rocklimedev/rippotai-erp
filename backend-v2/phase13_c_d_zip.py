"""Phase C (Quotations & Vendors aliases + Excel import),
Phase D (BOQ PDF completeness + auto-attach real bytes),
Handover-package ZIP endpoint.

Registers additional/aliased endpoints on top of existing phase4/phase7 modules
without removing anything the current implementation exposes.
"""
import io
import csv
import uuid
import base64
import zipfile
from datetime import datetime, timezone
from typing import Optional, List, Dict, Any

from fastapi import Depends, HTTPException, Body, UploadFile, File, Query
from fastapi.responses import StreamingResponse, Response
from pydantic import BaseModel

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill


def _now():
    return datetime.now(timezone.utc).isoformat()


VENDOR_TEMPLATE_COLUMNS = [
    "name", "company", "vendor_type", "primary_category", "phone", "email",
    "gst", "city", "state", "price_range", "rating", "specializations", "brands",
    "materials", "verified", "preferred", "availability_status",
]


class QuotationRemark(BaseModel):
    remarks: Optional[str] = ""


def register_phase13(api, db, get_current_user, require_internal_user, _full_boq):

    # =====================================================================
    # Phase C — Quotation status aliases: submit / decline / resubmit
    # =====================================================================
    @api.post("/quotations/{qid}/submit")
    async def quotation_submit(qid: str, payload: QuotationRemark = Body(default=QuotationRemark()), current=Depends(require_internal_user)):
        q = await db.quotations.find_one({"id": qid}, {"_id": 0})
        if not q:
            raise HTTPException(404, "Quotation not found")
        if q.get("status") not in ("draft", "returned"):
            raise HTTPException(400, f"Cannot submit from status '{q.get('status')}'")
        # Block if vendor is blacklisted
        if q.get("vendor_id"):
            v = await db.vendors.find_one({"id": q["vendor_id"]}, {"_id": 0}) or {}
            if v.get("status") == "blocked" or v.get("blacklisted"):
                raise HTTPException(403, "Vendor is blacklisted — cannot submit quotation")
        await db.quotations.update_one({"id": qid}, {"$set": {
            "status": "submitted", "submitted_at": _now(), "updated_at": _now(),
        }})
        await db.quotation_activity.insert_one({
            "id": str(uuid.uuid4()), "quotation_id": qid, "type": "submit",
            "actor": current.get("name"), "at": _now(), "meta": {"remarks": payload.remarks},
        })
        return {"ok": True, "status": "submitted"}

    @api.post("/quotations/{qid}/decline")
    async def quotation_decline(qid: str, payload: QuotationRemark = Body(default=QuotationRemark()), current=Depends(require_internal_user)):
        q = await db.quotations.find_one({"id": qid}, {"_id": 0})
        if not q:
            raise HTTPException(404, "Quotation not found")
        await db.quotations.update_one({"id": qid}, {"$set": {
            "status": "declined", "declined_at": _now(), "updated_at": _now(),
        }})
        await db.quotation_activity.insert_one({
            "id": str(uuid.uuid4()), "quotation_id": qid, "type": "decline",
            "actor": current.get("name"), "at": _now(), "meta": {"remarks": payload.remarks},
        })
        return {"ok": True, "status": "declined"}

    @api.post("/quotations/{qid}/resubmit")
    async def quotation_resubmit(qid: str, payload: QuotationRemark = Body(default=QuotationRemark()), current=Depends(require_internal_user)):
        q = await db.quotations.find_one({"id": qid}, {"_id": 0})
        if not q:
            raise HTTPException(404, "Quotation not found")
        if q.get("status") not in ("returned", "declined"):
            raise HTTPException(400, f"Cannot resubmit from status '{q.get('status')}'")
        # Append prior state to versions[]
        versions = q.get("versions") or []
        versions.append({
            "at": _now(), "actor": current.get("name"),
            "prior_status": q.get("status"), "revision": len(versions) + 1,
            "remarks": payload.remarks,
        })
        await db.quotations.update_one({"id": qid}, {"$set": {
            "status": "submitted", "resubmitted_at": _now(),
            "updated_at": _now(), "versions": versions,
        }})
        await db.quotation_activity.insert_one({
            "id": str(uuid.uuid4()), "quotation_id": qid, "type": "resubmit",
            "actor": current.get("name"), "at": _now(),
            "meta": {"remarks": payload.remarks, "revision": len(versions)},
        })
        return {"ok": True, "status": "submitted", "revision": len(versions)}

    # =====================================================================
    # Phase C — Vendors: blacklist-exclusion picker + Excel/CSV import
    # =====================================================================
    @api.get("/vendors-picker")
    async def vendors_picker(
        q: Optional[str] = None, category: Optional[str] = None,
        current=Depends(require_internal_user),
    ):
        """Returns only active vendors (excludes blocked/blacklisted/inactive) for pickers."""
        query: Dict[str, Any] = {
            "status": {"$nin": ["blocked", "inactive", "archived"]},
            "blacklisted": {"$ne": True},
        }
        if category:
            query["primary_category"] = category
        if q:
            query["$or"] = [
                {"name": {"$regex": q, "$options": "i"}},
                {"company": {"$regex": q, "$options": "i"}},
                {"primary_category": {"$regex": q, "$options": "i"}},
            ]
        rows = await db.vendors.find(query, {"_id": 0}).sort("name", 1).to_list(300)
        return rows

    @api.get("/vendors-template")
    async def vendors_template(fmt: str = "csv", current=Depends(require_internal_user)):
        if fmt == "xlsx":
            wb = Workbook()
            ws = wb.active
            ws.title = "Vendors"
            ws.append(VENDOR_TEMPLATE_COLUMNS)
            head_font = Font(bold=True, color="FFFFFF")
            fill = PatternFill("solid", fgColor="1F453B")
            for cell in ws[1]:
                cell.font = head_font
                cell.fill = fill
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return StreamingResponse(
                buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": 'attachment; filename="vendors-template.xlsx"'},
            )
        # CSV
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(VENDOR_TEMPLATE_COLUMNS)
        w.writerow([
            "Sample Vendor Pvt Ltd", "Sample Vendor Pvt Ltd", "Contractor", "Painting",
            "9876500000", "sample@vendor.in", "07AAACS1234R1Z1", "New Delhi", "Delhi",
            "mid", "4.2", "Interior painting;Texture", "Asian Paints;Berger",
            "PU;Distemper", "true", "false", "available",
        ])
        return Response(
            content=buf.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": 'attachment; filename="vendors-template.csv"'},
        )

    @api.post("/vendors-import")
    async def vendors_import(file: UploadFile = File(...), current=Depends(require_internal_user)):
        data = await file.read()
        rows: List[Dict[str, Any]] = []
        fname = (file.filename or "").lower()
        try:
            if fname.endswith((".xlsx", ".xlsm")):
                wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
                ws = wb.active
                header = None
                for r in ws.iter_rows(values_only=True):
                    if header is None:
                        header = [str(c or "").strip().lower() for c in r]
                        continue
                    if not any(r):
                        continue
                    rows.append({header[i]: r[i] for i in range(min(len(header), len(r)))})
            else:
                text = data.decode("utf-8-sig", errors="ignore")
                reader = csv.DictReader(io.StringIO(text))
                rows = list(reader)
        except Exception as e:
            raise HTTPException(400, f"Could not parse file: {e}")

        imported, skipped, errors = 0, 0, []
        for i, row in enumerate(rows, start=2):
            name = (row.get("name") or row.get("company") or "").strip()
            phone = (row.get("phone") or "").strip()
            if not name and not phone:
                skipped += 1
                errors.append({"row": i, "reason": "missing name and phone"})
                continue
            existing = await db.vendors.find_one({"$or": [
                {"phone": phone} if phone else {"_never": True},
                {"name": name, "company": (row.get("company") or name).strip()},
            ]}, {"_id": 0, "id": 1})
            if existing:
                skipped += 1
                continue
            doc = {
                "id": str(uuid.uuid4()),
                "name": name,
                "company": (row.get("company") or name).strip(),
                "vendor_type": (row.get("vendor_type") or "Contractor").strip(),
                "primary_category": (row.get("primary_category") or "General").strip(),
                "phone": phone,
                "email": (row.get("email") or "").strip(),
                "gst": (row.get("gst") or "").strip(),
                "city": (row.get("city") or "").strip(),
                "state": (row.get("state") or "").strip(),
                "price_range": (row.get("price_range") or "mid").strip(),
                "rating": float(row.get("rating") or 0) or None,
                "specializations": [s.strip() for s in str(row.get("specializations") or "").split(";") if s.strip()],
                "brands": [s.strip() for s in str(row.get("brands") or "").split(";") if s.strip()],
                "materials": [s.strip() for s in str(row.get("materials") or "").split(";") if s.strip()],
                "verified": str(row.get("verified") or "").strip().lower() in ("1", "true", "yes"),
                "preferred": str(row.get("preferred") or "").strip().lower() in ("1", "true", "yes"),
                "availability_status": (row.get("availability_status") or "available").strip(),
                "status": "active",
                "blacklisted": False,
                "created_at": _now(),
                "updated_at": _now(),
                "source": "imported",
            }
            await db.vendors.insert_one(doc)
            imported += 1
        return {"ok": True, "imported": imported, "skipped": skipped, "errors": errors[:20], "total_rows": len(rows)}

    # =====================================================================
    # ZIP — Handover Package download
    # =====================================================================
    @api.get("/projects/{project_id}/handover-package/download")
    async def handover_package_zip(project_id: str, current=Depends(get_current_user)):
        proj = await db.projects.find_one({"id": project_id}, {"_id": 0}) or {}
        rows = await db.documents.find({
            "project_id": project_id, "is_archived": {"$ne": True},
            "status": {"$in": ["approved", "gfc", "final", "signed"]},
        }).to_list(500)
        if not rows:
            raise HTTPException(404, "No handover-eligible documents for this project")
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
            manifest = []
            for r in rows:
                fname = r.get("filename") or f"{r.get('id')}.bin"
                content = r.get("content_b64")
                if content:
                    z.writestr(f"{r.get('category','Other')}/{fname}", base64.b64decode(content))
                manifest.append({
                    "category": r.get("category"),
                    "title": r.get("title"),
                    "filename": fname,
                    "version": r.get("version"),
                    "status": r.get("status"),
                    "uploaded_by": r.get("uploaded_by_name") or r.get("uploaded_by"),
                    "date": r.get("document_date"),
                })
            z.writestr("MANIFEST.txt", "Handover Package — " + (proj.get("name") or "Project") + "\n" +
                       "Generated " + _now() + "\n\n" +
                       "\n".join(f"[{m['category']}] {m['filename']} · v{m['version']} · {m['status']} · {m['date']} · {m['uploaded_by']}" for m in manifest))
        buf.seek(0)
        data = buf.getvalue()
        safe_name = (proj.get("name") or "project").replace(" ", "-")
        return Response(
            content=data,
            media_type="application/zip",
            headers={
                "Content-Disposition": f'attachment; filename="Handover-{safe_name}.zip"',
                "X-Doc-Count": str(len(rows)),
            },
        )
